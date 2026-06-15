# -*- coding: utf-8 -*-
from PySide6.QtWidgets import (
    QApplication, QDialog, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFileDialog, QToolBar, QStatusBar, QMessageBox, QTableWidget,
    QTableWidgetItem, QSpinBox, QComboBox, QLineEdit, QColorDialog,
    QTabWidget, QFrame, QAbstractItemView, QHeaderView, QGraphicsScene, QGraphicsView, QProgressBar,
    QScrollArea,
    QPlainTextEdit, QTextBrowser
)
from PySide6.QtCore import Qt, QThreadPool, QTimer
from PySide6.QtGui import QAction, QColor, QPen, QBrush, QFont
import os
import math
import csv
from datetime import datetime
from html import escape
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill, Border, Side

# 版本号：优先从当前工程的 __init__ 里取，取不到就用 "dev"
try:
    from __init__ import __version__
except Exception:
    __version__ = "dev"

# Worker & tickets：直接从当前工程内部模块导入
from infra.threads import Worker
from core import tickets
from core.analysis import (
    apply_time_window_analysis as core_apply_time_window_analysis,
    compute_car_capacity_results,
)
from core.input_parser import parse_multi_project_inputs as core_parse_multi_project_inputs


class ExportTicketWindow(QMainWindow):
    """
    M-Line 混流节拍仿真系统（独立于数据校对）
    v2 岗位矩阵字段：
      序号 / 工程名称 / 设备数量 / 所属线别 / 岗位设备 / A工时 / B工时 / C工时
    说明：
      - 设备数量：1 表示单资源；2 表示双线双资源。
      - 所属线别：1号线 / 2号线 / 双线 / 双线共用。
      - A/B/C 工时 > 0：该车型在该岗位作业；工时 = 0：经过该岗位但不作业。
      - 参与投车的车型，工时不能为空；未参与投车的车型，工时可以为空。
    """
    COL_C_TIME = 7
    MAX_SINGLE_STEPS = 23  # 单工程组合票：新版模板固定支持 23 行

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("M-Line 混流节拍仿真系统 v2.9")
        self.resize(1180, 760)
        self.setMinimumSize(1080, 700)

        self.thread_pool = QThreadPool.globalInstance()
        self.dst_path = None
        self._frozen_vehicle_sequence = None
        self._frozen_vehicle_sequence_signature = None
        self._frozen_vehicle_sequence_hash = ""
        self._frozen_vehicle_sequence_generated_at = ""

        self._build_ui()
        self._connect_signals()
    def _on_tab_changed(self, index: int):
        """
        Tab 切换时，控制多工程页内“添加步骤 / 删除步骤 / 填入示例”按钮：
        - 仅在『多工程组合票』页签（第 0 个 Tab）启用；
        - 在『单工程组合票』页签禁用，避免误点影响单工程表。
        """
        is_multi = (index == 0)
        if hasattr(self, "btn_add_row"):
            self.btn_add_row.setEnabled(is_multi)
        if hasattr(self, "btn_del_row"):
            self.btn_del_row.setEnabled(is_multi)
        if hasattr(self, "btn_fill_sample"):
            self.btn_fill_sample.setEnabled(is_multi)
    # ---------------- UI ---------------- #
    def _build_ui(self):
        tb = QToolBar("Ticket")
        self.addToolBar(tb)

        self.act_help = QAction("帮助", self)
        tb.addAction(self.act_help)
        tb.addSeparator()

        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(10)

        # ====== Tab 控件 ======
        self.tabs = QTabWidget(self)
        root.addWidget(self.tabs)

        # ---------- Tab1：多车组合票 ----------
        self.page_multi = QWidget(self)
        page_multi_layout = QVBoxLayout(self.page_multi)
        page_multi_layout.setContentsMargins(8, 8, 8, 8)
        page_multi_layout.setSpacing(10)
        self.tabs.addTab(self.page_multi, "多工程组合票")

        # 多工程内部二级页：第一页负责录入，第二页负责分析/导出/后续动画预留
        self.multi_tabs = QTabWidget(self.page_multi)
        page_multi_layout.addWidget(self.multi_tabs, 1)

        self.page_multi_input = QWidget(self.page_multi)
        self.page_multi_result_scroll = QScrollArea(self.page_multi)
        self.page_multi_result_scroll.setWidgetResizable(True)
        self.page_multi_result_scroll.setFrameShape(QFrame.NoFrame)
        self.page_multi_result_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.page_multi_result = QWidget(self.page_multi_result_scroll)
        self.page_multi_result_scroll.setWidget(self.page_multi_result)

        input_layout = QVBoxLayout(self.page_multi_input)
        input_layout.setContentsMargins(0, 0, 0, 0)
        input_layout.setSpacing(10)

        result_layout = QVBoxLayout(self.page_multi_result)
        result_layout.setContentsMargins(0, 0, 0, 0)
        result_layout.setSpacing(10)


        self.multi_tabs.addTab(self.page_multi_input, "参数与岗位")
        self.multi_tabs.addTab(self.page_multi_result_scroll, "分析与导出")


        def _make_block(title: str):
            frame = QFrame(self.page_multi)
            frame.setFrameShape(QFrame.StyledPanel)
            frame.setObjectName("ticketBlock")
            frame.setStyleSheet(
                "QFrame#ticketBlock {"
                "background: #ffffff;"
                "border: 1px solid #d9dee7;"
                "border-radius: 10px;"
                "}"
            )
            lay = QVBoxLayout(frame)
            lay.setContentsMargins(12, 12, 12, 12)
            lay.setSpacing(10)
            title_label = QLabel(title, frame)
            title_label.setStyleSheet("font-size: 14px; font-weight: 700; color: #223042;")
            lay.addWidget(title_label)
            return frame, lay

        top_split = QVBoxLayout()
        top_split.setSpacing(10)
        input_layout.addLayout(top_split)

        # 左侧：运行参数区块
        params_frame, params_layout = _make_block("运行参数")
        top_split.addWidget(params_frame, 0)

        row_top_1 = QHBoxLayout()
        row_top_1.setSpacing(8)
        params_layout.addLayout(row_top_1)
        row_top_1.addWidget(QLabel("工程名称："))
        self.ed_project = QLineEdit()
        self.ed_project.setPlaceholderText("例如：L2++")
        self.ed_project.setFixedWidth(220)
        row_top_1.addWidget(self.ed_project)

        row_top_1.addSpacing(16)
        row_top_1.addWidget(QLabel("投车模式："))
        self.cmb_launch_mode = QComboBox()
        self.cmb_launch_mode.addItems(["按数量投车", "按比例投车"])
        self.cmb_launch_mode.setCurrentIndex(0)
        row_top_1.addWidget(self.cmb_launch_mode)
        row_top_1.addStretch()

        row_top_2 = QHBoxLayout()
        row_top_2.setSpacing(8)
        params_layout.addLayout(row_top_2)
        self.lbl_a_cars = QLabel("A数量：")
        row_top_2.addWidget(self.lbl_a_cars)
        self.spn_a_cars = QSpinBox()
        self.spn_a_cars.setRange(0, 9999)
        self.spn_a_cars.setValue(4)
        row_top_2.addWidget(self.spn_a_cars)

        row_top_2.addSpacing(10)
        self.lbl_b_cars = QLabel("B数量：")
        row_top_2.addWidget(self.lbl_b_cars)
        self.spn_b_cars = QSpinBox()
        self.spn_b_cars.setRange(0, 9999)
        self.spn_b_cars.setValue(0)
        row_top_2.addWidget(self.spn_b_cars)

        row_top_2.addSpacing(10)
        self.lbl_c_cars = QLabel("C数量：")
        row_top_2.addWidget(self.lbl_c_cars)
        self.spn_c_cars = QSpinBox()
        self.spn_c_cars.setRange(0, 9999)
        self.spn_c_cars.setValue(0)
        row_top_2.addWidget(self.spn_c_cars)
        row_top_2.addStretch()

        row_ratio = QHBoxLayout()
        row_ratio.setSpacing(8)
        params_layout.addLayout(row_ratio)
        self.lbl_total_cars = QLabel("分析时间：")
        row_ratio.addWidget(self.lbl_total_cars)
        self.spn_total_cars = QSpinBox()
        self.spn_total_cars.setRange(1, 9999)
        self.spn_total_cars.setValue(60)
        self.spn_total_cars.setSuffix(" 分钟")
        self.spn_total_cars.setToolTip(
            "按比例投车模式下使用；按分析时间和目标节拍计算理论投车台数，"
            "并额外生成50台仿真缓冲车辆观察窗口尾部趋势。"
        )
        row_ratio.addWidget(self.spn_total_cars)
        self.lbl_total_cars.hide()
        self.spn_total_cars.hide()

        # 旧比例文本框保留但隐藏，后续可删除；当前按比例模式改用 A/B/C 数值框作为比例。
        self.ed_ratio = QLineEdit()
        self.ed_ratio.hide()

        row_ratio.addStretch()

        row_top_3 = QHBoxLayout()
        row_top_3.setSpacing(8)
        params_layout.addLayout(row_top_3)
        row_top_3.addWidget(QLabel("时间格刻度："))
        self.cmb_grid = QComboBox()
        self.cmb_grid.addItems(["1.0", "0.5", "2.0"])
        self.cmb_grid.setCurrentIndex(0)
        row_top_3.addWidget(self.cmb_grid)

        # 等待分配先隐藏，底层默认按“开始前等待”
        self.cmb_wait = QComboBox()
        self.cmb_wait.addItems(["开始前等待", "末尾等待"])
        self.cmb_wait.setCurrentIndex(0)
        self.cmb_wait.hide()

        row_top_3.addSpacing(12)
        row_top_3.addWidget(QLabel("目标节拍："))
        self.spn_target_takt = QSpinBox()
        self.spn_target_takt.setRange(0, 9999)
        self.spn_target_takt.setValue(118)
        self.spn_target_takt.setToolTip("0表示不进行节拍判定；大于0时按各岗位A/B/C实际工时判断OK/NG")
        row_top_3.addWidget(self.spn_target_takt)
        
        row_top_3.addSpacing(12)
        self.lbl_sequence_mode = QLabel("排列方式：")
        row_top_3.addWidget(self.lbl_sequence_mode)
        self.cmb_seq = QComboBox()
        self.cmb_seq.addItems(["顺排(A→B→C)", "交替混流"])
        self.cmb_seq.setCurrentIndex(0)
        row_top_3.addWidget(self.cmb_seq)

        row_top_3.addSpacing(12)
        self.lbl_max_run = QLabel("最大连续台数：")
        row_top_3.addWidget(self.lbl_max_run)
        self.spn_max_run = QSpinBox()
        self.spn_max_run.setRange(1, 9999)
        self.spn_max_run.setValue(10)
        self.spn_max_run.setToolTip("默认10台；填1表示尽量强制交替")
        row_top_3.addWidget(self.spn_max_run)

        row_top_3.addSpacing(8)
        self.btn_freeze_sequence = QPushButton("排列冻结", self.page_multi_input)
        self.btn_freeze_sequence.setFixedWidth(82)
        self.btn_freeze_sequence.setToolTip("生成并冻结当前按数量交替混流的完整投车顺序")
        row_top_3.addWidget(self.btn_freeze_sequence)
        self.lbl_sequence_freeze_status = QLabel("未冻结")
        self.lbl_sequence_freeze_status.setStyleSheet("color: #8a5a00; font-size: 12px;")
        row_top_3.addWidget(self.lbl_sequence_freeze_status)
        row_top_3.addStretch()

        self.params_tip = QLabel(
            "顺排/交替混流：A/B/C 填数量；按比例投车：A/B/C 填比例，并填写分析时间。下方岗位矩阵用于逐行录入步骤。"
        )
        self.params_tip.setWordWrap(True)
        self.params_tip.setStyleSheet("color: #5f6b7a; font-size: 12px;")
        params_layout.addWidget(self.params_tip)

        # 中部：岗位矩阵区块
        table_frame, table_layout = _make_block("岗位矩阵")

        table_action_row = QHBoxLayout()
        table_action_row.setSpacing(8)
        table_layout.addLayout(table_action_row)
        table_action_row.addStretch()
        self.btn_import_matrix = QPushButton("导入矩阵", self.page_multi_input)
        self.btn_export_matrix = QPushButton("导出矩阵", self.page_multi_input)
        self.btn_add_row = QPushButton("添加步骤", self.page_multi_input)
        self.btn_del_row = QPushButton("删除步骤", self.page_multi_input)
        self.btn_fill_sample = QPushButton("填入示例", self.page_multi_input)
        table_action_row.addWidget(self.btn_import_matrix)
        table_action_row.addWidget(self.btn_export_matrix)
        table_action_row.addWidget(self.btn_add_row)
        table_action_row.addWidget(self.btn_del_row)
        table_action_row.addWidget(self.btn_fill_sample)

        input_layout.addWidget(table_frame, 1)

        input_next_row = QHBoxLayout()
        input_next_row.setSpacing(8)
        input_next_row.addStretch()
        self.btn_export = QPushButton("生成并导出组合票", self.page_multi_input)
        input_next_row.addWidget(self.btn_export)
        self.btn_go_result_page = QPushButton("下一步：分析与导出", self.page_multi_input)
        input_next_row.addWidget(self.btn_go_result_page)
        input_layout.addLayout(input_next_row)

        self.tbl = QTableWidget(0, 8, self)
        self.tbl.setHorizontalHeaderLabels([
            "序号", "工程名称", "设备数量", "所属线别",
            "岗位设备", "A工时", "B工时", "C工时"
        ])
        self.tbl.horizontalHeader().setStretchLastSection(False)
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setEditTriggers(
            QAbstractItemView.DoubleClicked
            | QAbstractItemView.EditKeyPressed
        )
        self.tbl.setTabKeyNavigation(False)
        self.tbl.setAlternatingRowColors(True)
        table_layout.addWidget(self.tbl, 1)

        # 分析页顶部：只保留一个小按钮行，避免占用纵向空间。
        result_nav_row = QHBoxLayout()
        result_nav_row.setSpacing(8)
        result_nav_row.addStretch()
        self.btn_analyze = QPushButton("分析当前排程", self.page_multi_result)
        result_nav_row.addWidget(self.btn_analyze)
        result_layout.addLayout(result_nav_row)

        # 以下控件仅作为旧逻辑兼容容器保留，不再占用分析页版面空间。
        self.lbl_analysis = QLabel(
            "结果分析：点击『分析当前排程』后显示总车数、总完成时间、总等待时间、平均等待时间与节拍判定。",
            self.page_multi_result,
        )
        self.lbl_analysis.setWordWrap(True)
        self.lbl_analysis.hide()

        self.tbl_station_analysis = QTableWidget(0, 8, self.page_multi_result)
        self.tbl_station_analysis.setHorizontalHeaderLabels([
            "岗位", "经过台数", "累计工时", "累计等待", "平均工时", "平均等待", "节拍判定", "超节拍车型"
        ])
        self.tbl_station_analysis.horizontalHeader().setStretchLastSection(False)
        self.tbl_station_analysis.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tbl_station_analysis.verticalHeader().setVisible(False)
        self.tbl_station_analysis.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl_station_analysis.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl_station_analysis.setAlternatingRowColors(True)
        self.tbl_station_analysis.hide()

        # 车型数据摘要区：只承载基础排程摘要与模型结果。
        vehicle_summary_frame, vehicle_summary_layout = _make_block("车型数据摘要区")
        result_layout.addWidget(vehicle_summary_frame, 3)

        self.sim_timer = QTimer(self)
        self.sim_timer.setInterval(100)
        self.sim_time = 0.0
        self.current_defs = []
        self.last_schedule_rows = []
        self.last_analysis = None
        self.last_max_finish = 0.0
        self._last_model_result_summary = None

        self.lbl_vehicle_summary = QLabel("")
        self.lbl_vehicle_summary.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.lbl_vehicle_summary.setWordWrap(True)
        self.lbl_vehicle_summary.setTextFormat(Qt.RichText)
        self.lbl_vehicle_summary.setTextInteractionFlags(Qt.TextBrowserInteraction)
        self.lbl_vehicle_summary.setOpenExternalLinks(False)
        self.lbl_vehicle_summary.setMinimumHeight(175)
        self.lbl_vehicle_summary.setMaximumHeight(225)
        self.lbl_vehicle_summary.setStyleSheet(
            "background: #f7f9fc;"
            "border: 1px dashed #cbd5e1;"
            "border-radius: 8px;"
            "padding: 10px;"
            "color: #334155;"
            "font-size: 13px;"
            "line-height: 1.5;"
        )

        self.btn_model_result_explanation = QPushButton("结果说明", self.lbl_vehicle_summary)
        self.btn_model_result_explanation.setFixedHeight(24)
        self.btn_model_result_explanation.setMinimumWidth(80)
        self.btn_model_result_explanation.setStyleSheet(
            "QPushButton {"
            "background:#ffffff;"
            "border:1px solid #cbd5e1;"
            "border-radius:6px;"
            "padding:0 10px;"
            "color:#334155;"
            "font-size:12px;"
            "}"
            "QPushButton:hover { background:#f8fafc; }"
        )
        self.btn_model_result_explanation.raise_()

        vehicle_summary_layout.addWidget(self.lbl_vehicle_summary, 1)

        simulation_frame = QFrame(self.page_multi)
        simulation_frame.setFrameShape(QFrame.StyledPanel)
        simulation_frame.setObjectName("ticketBlock")
        simulation_frame.setStyleSheet(
            "QFrame#ticketBlock {"
            "background: #ffffff;"
            "border: 1px solid #d9dee7;"
            "border-radius: 10px;"
            "}"
        )
        simulation_layout = QVBoxLayout(simulation_frame)
        simulation_layout.setContentsMargins(12, 12, 12, 10)
        simulation_layout.setSpacing(6)
        result_layout.addWidget(simulation_frame, 7)

        sim_control_row = QHBoxLayout()
        sim_control_row.setContentsMargins(0, 0, 0, 0)
        sim_control_row.setSpacing(6)
        sim_title = QLabel("仿真回放", simulation_frame)
        sim_title.setStyleSheet("font-size: 14px; font-weight: 700; color: #223042;")
        sim_control_row.addWidget(sim_title)
        sim_control_row.addStretch(1)
        self.lbl_sim_time = QLabel("当前播放：0.0s / 0.0s")
        sim_control_row.addWidget(self.lbl_sim_time)
        sim_control_row.addSpacing(10)
        self.lbl_sim_total_wait = QLabel("当前累计实际等待：0s")
        self.lbl_sim_total_wait.hide()
        sim_control_row.addSpacing(10)
        self.sim_progress = QProgressBar(self.page_multi_result)
        self.sim_progress.setRange(0, 1000)
        self.sim_progress.setValue(0)
        self.sim_progress.setTextVisible(True)
        self.sim_progress.setMinimumWidth(180)
        self.sim_progress.setMaximumWidth(260)
        self.sim_progress.setMaximumHeight(18)
        self.sim_progress.setFormat("进度 %p%")
        sim_control_row.addWidget(self.sim_progress)
        self.lbl_realtime_takt = QLabel("近期节拍（近5个间隔）：-")
        self.lbl_realtime_takt.setToolTip(
            "当前播放时刻已下线车辆中，最近最多6台形成的最多5个相邻下线间隔平均值。"
            "仅用于观察近期下线波动，不参与达标车辆、达标率、整体节拍或风险提示计算。"
        )
        self.lbl_realtime_takt.setStyleSheet(
            "QLabel {"
            "font-size: 12px;"
            "font-weight: 700;"
            "color: #334155;"
            "background: #f8fafc;"
            "border: 1px solid #dbe3ef;"
            "border-radius: 5px;"
            "padding: 3px 8px;"
            "}"
        )
        sim_control_row.addWidget(self.lbl_realtime_takt)
        self.cmb_sim_speed = QComboBox()
        self.cmb_sim_speed.addItems(["1x", "5x", "10x", "20x", "30x", "40x", "50x", "60x", "80x", "100x"])
        self.cmb_sim_speed.setCurrentText("10x")
        self.btn_sim_play = QPushButton("播放")
        self.btn_sim_pause = QPushButton("暂停")
        self.btn_sim_reset = QPushButton("重置")
        control_h = 26
        self.cmb_sim_speed.setMinimumHeight(control_h)
        self.cmb_sim_speed.setMaximumHeight(control_h)
        self.btn_sim_play.setMinimumHeight(control_h)
        self.btn_sim_play.setMaximumHeight(control_h)
        self.btn_sim_pause.setMinimumHeight(control_h)
        self.btn_sim_pause.setMaximumHeight(control_h)
        self.btn_sim_reset.setMinimumHeight(control_h)
        self.btn_sim_reset.setMaximumHeight(control_h)
        self.btn_sim_play.setMinimumWidth(52)
        self.btn_sim_pause.setMinimumWidth(52)
        self.btn_sim_reset.setMinimumWidth(52)
        sim_control_row.addWidget(QLabel("速度："))
        sim_control_row.addWidget(self.cmb_sim_speed)
        sim_control_row.addWidget(self.btn_sim_play)
        sim_control_row.addWidget(self.btn_sim_pause)
        sim_control_row.addWidget(self.btn_sim_reset)

        self.sim_scene = QGraphicsScene(self)
        self.sim_graphics_view = QGraphicsView(self.sim_scene, simulation_frame)
        self.sim_graphics_view.setMinimumHeight(232)
        self.sim_graphics_view.setMaximumHeight(252)
        self.sim_graphics_view.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.sim_graphics_view.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.sim_graphics_view.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.sim_graphics_view.setStyleSheet(
            "background: #243244;"
            "border: 1px solid #334155;"
            "border-radius: 8px;"
        )
        simulation_layout.addLayout(sim_control_row, 0)
        simulation_layout.addWidget(self.sim_graphics_view, 0)

        sim_status_frame = QFrame(simulation_frame)
        sim_status_frame.setObjectName("simStatusFrame")
        sim_status_frame.setStyleSheet(
            "QFrame#simStatusFrame {"
            "background: #f8fafc;"
            "border: 1px solid #cbd5e1;"
            "border-radius: 6px;"
            "}"
        )
        sim_status_layout = QVBoxLayout(sim_status_frame)
        sim_status_layout.setContentsMargins(10, 6, 10, 5)
        sim_status_layout.setSpacing(0)

        self.lbl_sim_view = QLabel("仿真画面：请先点击『分析当前排程』。", sim_status_frame)
        self.lbl_sim_view.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.lbl_sim_view.setWordWrap(True)
        self.lbl_sim_view.setTextFormat(Qt.RichText)
        self.lbl_sim_view.setMinimumHeight(52)
        self.lbl_sim_view.setMaximumHeight(58)
        self.lbl_sim_view.setStyleSheet(
            "background: transparent;"
            "border: 0;"
            "padding: 0;"
            "color: #334155;"
            "font-size: 12px;"
        )
        sim_status_layout.addWidget(self.lbl_sim_view)
        sim_status_frame.setMinimumHeight(66)
        sim_status_frame.setMaximumHeight(72)
        simulation_layout.addStretch(1)
        simulation_layout.addWidget(sim_status_frame, 0)

        vehicle_log_entry = QFrame(self.page_multi_result)
        vehicle_log_entry.setMaximumHeight(44)
        vehicle_log_entry.setStyleSheet(
            "background: #ffffff;"
            "border: 1px solid #e2e8f0;"
            "border-radius: 8px;"
        )
        vehicle_log_layout = QHBoxLayout(vehicle_log_entry)
        vehicle_log_layout.setContentsMargins(12, 8, 12, 8)
        vehicle_log_layout.setSpacing(8)
        vehicle_log_title = QLabel("车辆明细 / 调试日志", vehicle_log_entry)
        vehicle_log_title.setStyleSheet("font-size: 13px; font-weight: 700; color: #334155;")
        vehicle_log_layout.addWidget(vehicle_log_title)
        vehicle_log_layout.addStretch()
        self.btn_vehicle_log = QPushButton("查看车辆日志", vehicle_log_entry)
        vehicle_log_layout.addWidget(self.btn_vehicle_log)
        result_layout.addWidget(vehicle_log_entry, 0)

        self.txt_schedule_debug = QPlainTextEdit(self.page_multi_result)
        self.txt_schedule_debug.setReadOnly(True)
        self.txt_schedule_debug.setMaximumBlockCount(300)
        self.txt_schedule_debug.setMaximumHeight(160)
        self.txt_schedule_debug.setPlaceholderText("排程运行日志：点击『分析当前排程』后显示前 200 条 rows 明细。")
        self.txt_schedule_debug.setStyleSheet(
            "background: #0f172a;"
            "border: 1px solid #334155;"
            "border-radius: 8px;"
            "padding: 8px;"
            "color: #e2e8f0;"
            "font-family: Menlo, Consolas, monospace;"
            "font-size: 11px;"
        )
        self.txt_schedule_debug.hide()
        result_layout.addWidget(self.txt_schedule_debug, 0)

        # ---------- Tab2：单工程组合票 ----------
        self.page_single = QWidget(self)
        page_single_layout = QVBoxLayout(self.page_single)
        page_single_layout.setContentsMargins(8, 8, 8, 8)
        page_single_layout.setSpacing(8)

        # 顶部基本信息
        row_info = QHBoxLayout()
        page_single_layout.addLayout(row_info)

        row_info.addWidget(QLabel("工程名称："))
        self.ed_sw_project = QLineEdit(self.page_single)
        self.ed_sw_project.setPlaceholderText("例如：前轴调整工位")
        self.ed_sw_project.setFixedWidth(200)
        row_info.addWidget(self.ed_sw_project)

        row_info.addSpacing(12)
        row_info.addWidget(QLabel("品番·品名："))
        self.ed_sw_part = QLineEdit(self.page_single)
        self.ed_sw_part.setPlaceholderText("例如：XXXX-XXXXX 前轮定位")
        self.ed_sw_part.setFixedWidth(220)
        row_info.addWidget(self.ed_sw_part)

        row_info.addSpacing(12)
        row_info.addWidget(QLabel("作业者："))
        self.ed_sw_worker = QLineEdit(self.page_single)
        self.ed_sw_worker.setPlaceholderText("例如：张三")
        self.ed_sw_worker.setFixedWidth(120)
        row_info.addWidget(self.ed_sw_worker)

        row_info.addSpacing(12)
        row_info.addWidget(QLabel("节拍TT(秒)："))
        self.spn_sw_takt = QSpinBox(self.page_single)
        self.spn_sw_takt.setRange(1, 9999)
        self.spn_sw_takt.setValue(118)  # 默认示例
        row_info.addWidget(self.spn_sw_takt)

        row_info.addStretch()

        # 作业手顺表（A→B 区间）
        self.tbl_sw = QTableWidget(0, 8, self.page_single)
        self.tbl_sw.setHorizontalHeaderLabels([
            "顺序", "作业名称A", "作业名称B",
            "手作业(秒)", "自动(秒)", "步行(秒)",
            "步行在前/后", "自动在前/后"
        ])
        self.tbl_sw.horizontalHeader().setStretchLastSection(True)
        self.tbl_sw.verticalHeader().setVisible(False)
        page_single_layout.addWidget(self.tbl_sw, 1)

        # 底部按钮栏（单工程组合票）
        row_btn_sw = QHBoxLayout()
        page_single_layout.addLayout(row_btn_sw)
        row_btn_sw.addStretch()

        self.btn_sw_add = QPushButton("添加作业行", self.page_single)
        self.btn_sw_del = QPushButton("删除选中行", self.page_single)
        self.btn_sw_export = QPushButton("导出标准作业组合票", self.page_single)

        row_btn_sw.addWidget(self.btn_sw_add)
        row_btn_sw.addWidget(self.btn_sw_del)
        row_btn_sw.addWidget(self.btn_sw_export)

        self.tabs.addTab(self.page_single, "单工程组合票")

        # 状态栏（用于显示导出进度 / 完成信息）
        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self._on_tab_changed(self.tabs.currentIndex())
        self._update_mode_ui()

    def _connect_signals(self):
        self.btn_add_row.clicked.connect(self.add_row)
        self.btn_del_row.clicked.connect(self.del_row)
        self.btn_fill_sample.clicked.connect(self.fill_sample)
        self.btn_import_matrix.clicked.connect(self._import_station_matrix)
        self.btn_export_matrix.clicked.connect(self._export_station_matrix)
        self.cmb_launch_mode.currentIndexChanged.connect(self._update_mode_ui)
        self.cmb_launch_mode.currentIndexChanged.connect(self._invalidate_frozen_vehicle_sequence)
        self.cmb_seq.currentIndexChanged.connect(self._update_sequence_freeze_ui)
        self.cmb_seq.currentIndexChanged.connect(self._invalidate_frozen_vehicle_sequence)
        self.spn_a_cars.valueChanged.connect(self._invalidate_frozen_vehicle_sequence)
        self.spn_b_cars.valueChanged.connect(self._invalidate_frozen_vehicle_sequence)
        self.spn_c_cars.valueChanged.connect(self._invalidate_frozen_vehicle_sequence)
        self.spn_total_cars.valueChanged.connect(self._invalidate_frozen_vehicle_sequence)
        self.spn_max_run.valueChanged.connect(self._invalidate_frozen_vehicle_sequence)
        self.btn_freeze_sequence.clicked.connect(self._freeze_vehicle_sequence)
        self.btn_go_result_page.clicked.connect(lambda: self.multi_tabs.setCurrentWidget(self.page_multi_result_scroll))
        self.btn_analyze.clicked.connect(self.do_analyze)
        self.btn_export.clicked.connect(self.do_export)
        self.btn_model_result_explanation.clicked.connect(self._show_model_result_explanation_dialog)
        self.lbl_vehicle_summary.linkActivated.connect(self._on_vehicle_summary_link_activated)
        self.act_help.triggered.connect(self.show_help)
        self.btn_sim_play.clicked.connect(self._start_simulation)
        self.btn_sim_pause.clicked.connect(self._pause_simulation)
        self.btn_sim_reset.clicked.connect(self._reset_simulation)
        self.btn_vehicle_log.clicked.connect(self._show_vehicle_log_placeholder)
        self.sim_timer.timeout.connect(self._on_simulation_tick)

        # Tab 切换时，控制多工程页内步骤按钮是否可用
        self.tabs.currentChanged.connect(self._on_tab_changed)

        # 单工程组合票 Tab
        self.btn_sw_add.clicked.connect(self.add_single_row)
        self.btn_sw_del.clicked.connect(self.del_single_row)
        self.btn_sw_export.clicked.connect(self.export_single_placeholder)

    def _update_mode_ui(self):
        """根据模式切换 A/B/C 输入含义：数量模式填写数量，比例模式填写比例 + 分析时间。"""
        is_ratio = self.cmb_launch_mode.currentIndex() == 1
        if hasattr(self, "lbl_a_cars"):
            self.lbl_a_cars.setText("A比例：" if is_ratio else "A数量：")
        if hasattr(self, "lbl_b_cars"):
            self.lbl_b_cars.setText("B比例：" if is_ratio else "B数量：")
        if hasattr(self, "lbl_c_cars"):
            self.lbl_c_cars.setText("C比例：" if is_ratio else "C数量：")
        if hasattr(self, "lbl_total_cars"):
            self.lbl_total_cars.setVisible(is_ratio)
        if hasattr(self, "spn_total_cars"):
            self.spn_total_cars.setVisible(is_ratio)
        if hasattr(self, "lbl_sequence_mode"):
            self.lbl_sequence_mode.setVisible(not is_ratio)
        if hasattr(self, "cmb_seq"):
            self.cmb_seq.setVisible(not is_ratio)
        if hasattr(self, "lbl_max_run"):
            self.lbl_max_run.setVisible(not is_ratio)
        if hasattr(self, "spn_max_run"):
            self.spn_max_run.setVisible(not is_ratio)
            self.spn_max_run.setEnabled(not is_ratio)
        self._update_sequence_freeze_ui()
        if hasattr(self, "params_tip"):
            if is_ratio:
                self.params_tip.setText(
                    "当前模式：按比例投车。A/B/C 填比例；分析时间填写 xx 分钟。"
                    "程序按目标节拍计算理论投车台数，并额外生成50台仿真缓冲车辆；"
                    "缓冲车辆不进入目标批次完成判断。"
                )
            else:
                self.params_tip.setText("当前模式：按数量投车。A/B/C 填数量；顺排按 A→B→C，交替混流可配合最大连续台数使用。")

    def _sequence_freeze_signature(self):
        return (
            self.cmb_launch_mode.currentIndex(),
            int(self.spn_a_cars.value()),
            int(self.spn_b_cars.value()),
            int(self.spn_c_cars.value()),
            int(self.spn_total_cars.value()),
            self.cmb_seq.currentIndex(),
            int(self.spn_max_run.value()),
            tickets.QUANTITY_SEQUENCE_RULE_VERSION,
        )

    def _update_sequence_freeze_ui(self, *_args):
        if not hasattr(self, "btn_freeze_sequence"):
            return
        is_quantity = self.cmb_launch_mode.currentIndex() == 0
        is_alternate = self.cmb_seq.currentIndex() == 1
        self.btn_freeze_sequence.setVisible(is_quantity)
        self.lbl_sequence_freeze_status.setVisible(is_quantity)
        self.btn_freeze_sequence.setEnabled(is_quantity and is_alternate)
        if is_quantity and not is_alternate:
            self.lbl_sequence_freeze_status.setText("顺排无需冻结")
            self.lbl_sequence_freeze_status.setStyleSheet("color: #667085; font-size: 12px;")
        elif self._frozen_vehicle_sequence is None:
            self.lbl_sequence_freeze_status.setText("未冻结")
            self.lbl_sequence_freeze_status.setStyleSheet("color: #8a5a00; font-size: 12px;")

    def _invalidate_frozen_vehicle_sequence(self, *_args):
        if self._frozen_vehicle_sequence is not None:
            self._frozen_vehicle_sequence = None
            self._frozen_vehicle_sequence_signature = None
            self._frozen_vehicle_sequence_hash = ""
            self._frozen_vehicle_sequence_generated_at = ""
        self._update_sequence_freeze_ui()

    @staticmethod
    def _max_sequence_run(vehicle_sequence):
        max_run = 0
        run = 0
        last_type = ""
        for vehicle_type in vehicle_sequence:
            if vehicle_type == last_type:
                run += 1
            else:
                last_type = vehicle_type
                run = 1
            max_run = max(max_run, run)
        return max_run

    def _freeze_vehicle_sequence(self):
        if self.cmb_launch_mode.currentIndex() != 0 or self.cmb_seq.currentIndex() != 1:
            QMessageBox.information(self, "排列冻结", "排列冻结当前用于按数量投车的交替混流。")
            return

        vehicle_counts = {
            "A": int(self.spn_a_cars.value()),
            "B": int(self.spn_b_cars.value()),
            "C": int(self.spn_c_cars.value()),
        }
        total_cars = sum(vehicle_counts.values())
        if total_cars <= 0:
            QMessageBox.warning(self, "排列冻结", "A/B/C数量合计必须大于0。")
            return

        max_consecutive = int(self.spn_max_run.value())
        vehicle_sequence = tickets.build_vehicle_sequence(
            total_cars,
            vehicle_counts,
            "alternate",
            max_consecutive,
        )
        actual_counts = {
            vehicle_type: vehicle_sequence.count(vehicle_type)
            for vehicle_type in ("A", "B", "C")
        }
        if actual_counts != vehicle_counts:
            QMessageBox.critical(self, "排列冻结", "生成排列与当前A/B/C数量不一致，未执行冻结。")
            return

        sequence_hash = tickets.vehicle_sequence_hash(vehicle_sequence)
        self._frozen_vehicle_sequence = list(vehicle_sequence)
        self._frozen_vehicle_sequence_signature = self._sequence_freeze_signature()
        self._frozen_vehicle_sequence_hash = sequence_hash
        self._frozen_vehicle_sequence_generated_at = datetime.now().isoformat(timespec="seconds")
        actual_max_run = self._max_sequence_run(vehicle_sequence)
        self.lbl_sequence_freeze_status.setText(
            f"已冻结 {len(vehicle_sequence)}台 · {sequence_hash[:8]}"
        )
        self.lbl_sequence_freeze_status.setStyleSheet("color: #087443; font-size: 12px; font-weight: 600;")
        preview = "".join(vehicle_sequence[:30])
        QMessageBox.information(
            self,
            "排列冻结完成",
            "\n".join([
                f"规则：{tickets.QUANTITY_SEQUENCE_RULE_VERSION}",
                f"车型数量：A{actual_counts['A']} / B{actual_counts['B']} / C{actual_counts['C']}",
                f"前30台：{preview}",
                f"实际最大连续台数：{actual_max_run}",
                f"SHA-256：{sequence_hash}",
            ]),
        )

    def _frozen_sequence_for_current_inputs(self, sequence_mode):
        if self.cmb_launch_mode.currentIndex() != 0 or sequence_mode != "alternate":
            return None
        if (
            self._frozen_vehicle_sequence is None
            or self._frozen_vehicle_sequence_signature != self._sequence_freeze_signature()
        ):
            self._invalidate_frozen_vehicle_sequence()
            raise ValueError("当前交替混流排列尚未冻结，请先点击“排列冻结”。")
        return list(self._frozen_vehicle_sequence)

    def do_analyze(self):
        try:
            project, cars, grid_step, wait_policy, defs, vehicle_counts, sequence_mode, max_consecutive, ratio_pattern, target_takt = self._collect_inputs()
            frozen_sequence = self._frozen_sequence_for_current_inputs(sequence_mode)
            rows, max_finish = tickets.schedule(
                defs,
                cars,
                vehicle_counts,
                sequence_mode,
                max_consecutive,
                ratio_pattern,
                launch_takt=target_takt,
                vehicle_sequence=frozen_sequence,
            )
            analysis = tickets.analyze_schedule(rows, max_finish, target_takt)
            analysis = self._apply_time_window_analysis(analysis, rows, target_takt)
            self.current_defs = list(defs or [])
            self.last_schedule_rows = rows
            self.last_analysis = analysis
            self.last_max_finish = float(max_finish or 0.0)
            if hasattr(self, "txt_schedule_debug"):
                self.txt_schedule_debug.setPlainText(self._build_schedule_debug_log(rows, limit=200))
            self.sim_time = 0.0
            self._update_sim_time_label()
            self._update_sim_total_wait_label()
            self._update_sim_view()
            self._draw_sim_scene()
            self._show_analysis_result(analysis)
            self._update_realtime_model_result()
            self.status.showMessage("排程分析完成", 6000)
        except Exception as e:
            import traceback
            detail = traceback.format_exc()
            print(detail)
            QMessageBox.warning(self, "分析失败", str(e))
    def _update_sim_total_wait_label(self):
        """刷新仿真控制栏中的关键判定信息。"""
        if not hasattr(self, "lbl_sim_total_wait"):
            return
        target_takt = float(self.spn_target_takt.value()) if hasattr(self, "spn_target_takt") else 0.0
        metrics = self._build_wait_display_metrics(
            float(getattr(self, "sim_time", 0.0) or 0.0), target_takt
        )
        actual_wait = self._fmt_analysis_num(metrics.get("total_actual_wait", 0.0))
        self.lbl_sim_total_wait.setText(f"当前累计实际等待：{actual_wait}s")
    def _sim_speed_value(self) -> float:
        """读取仿真倍速。"""
        if not hasattr(self, "cmb_sim_speed"):
            return 1.0
        text = self.cmb_sim_speed.currentText().replace("x", "").strip()
        try:
            return max(1.0, float(text))
        except Exception:
            return 1.0

    def _update_sim_time_label(self):
        """刷新仿真时间显示。"""
        if not hasattr(self, "lbl_sim_time"):
            return
        current = float(getattr(self, "sim_time", 0.0) or 0.0)
        total = float(getattr(self, "last_max_finish", 0.0) or 0.0)
        self.lbl_sim_time.setText(f"当前播放：{current:.1f}s / {total:.1f}s")
        if hasattr(self, "sim_progress"):
            if total > 0:
                progress_value = int(max(0.0, min(1.0, current / total)) * 1000)
            else:
                progress_value = 0
            self.sim_progress.setValue(progress_value)

    def _start_simulation(self):
        """启动仿真计时。当前阶段只推进时间，不绘制车辆。"""
        if float(getattr(self, "last_max_finish", 0.0) or 0.0) <= 0:
            QMessageBox.information(self, "提示", "请先点击『分析当前排程』生成仿真数据。")
            return
        self.sim_timer.start()

    def _pause_simulation(self):
        """暂停仿真计时。"""
        if hasattr(self, "sim_timer"):
            self.sim_timer.stop()

    def _reset_simulation(self):
        """重置仿真时间。"""
        if hasattr(self, "sim_timer"):
            self.sim_timer.stop()
        self.sim_time = 0.0
        self._update_sim_time_label()
        self._update_sim_view()
        self._draw_sim_scene()
        self._update_realtime_model_result()

    def _fmt_vehicle_log_value(self, value):
        if value is None:
            return ""
        try:
            num = float(value)
        except Exception:
            return str(value)
        if abs(num - round(num)) < 1e-9:
            return str(int(round(num)))
        return f"{num:.1f}"

    def _build_vehicle_log_rows(self):
        """按车辆维度整理当前 rows，仅用于只读日志弹窗。"""
        rows = getattr(self, "last_schedule_rows", []) or []
        if not rows:
            return [], []

        def _to_int(value, default=0):
            try:
                return int(float(value))
            except Exception:
                return default

        def _row_station(row):
            return str(row.get("step_display", row.get("station", row.get("group", ""))) or "")

        columns = [
            "车辆编号", "车型", "工序序号", "工序名称", "线别", "资源标识",
            "start", "加工完成", "离开工程", "结束时间", "加工工时", "完工后等待", "投入等待",
            "下一工序", "下一工序 start", "备注",
        ]

        output = []
        car_rows = self._sim_car_rows()
        sorted_car_items = sorted(
            car_rows.items(),
            key=lambda item: _to_int(item[0], 999999),
        )
        for car, segments in sorted_car_items:
            sorted_segments = sorted(
                list(segments or []),
                key=lambda row: (
                    self._sim_row_start(row),
                    _to_int(row.get("step_seq", 0)),
                ),
            )
            for idx, row in enumerate(sorted_segments):
                next_row = sorted_segments[idx + 1] if idx + 1 < len(sorted_segments) else None
                step_seq = _to_int(row.get("step_seq", ""), 0)
                next_seq = _to_int(next_row.get("step_seq", ""), 0) if next_row else 0
                note = ""
                if next_row and step_seq > 0 and next_seq > step_seq + 1:
                    skipped = ", ".join(str(seq) for seq in range(step_seq + 1, next_seq))
                    note = f"跳过中间岗位：{skipped}"

                output.append([
                    str(car),
                    str(row.get("car_type", row.get("duration_source", row.get("vehicle_type", ""))) or ""),
                    str(row.get("step_seq", "")),
                    _row_station(row),
                    str(row.get("line_no", row.get("line", "")) or ""),
                    str(row.get("resource_key", "") or ""),
                    self._fmt_vehicle_log_value(row.get("start", row.get("start_time", ""))),
                    self._fmt_vehicle_log_value(row.get("svc_finish", row.get("finish", ""))),
                    self._fmt_vehicle_log_value(row.get("depart", row.get("end", ""))),
                    self._fmt_vehicle_log_value(row.get("end", "")),
                    self._fmt_vehicle_log_value(row.get("dur", row.get("duration", ""))),
                    self._fmt_vehicle_log_value(row.get("block_wait", 0.0)),
                    self._fmt_vehicle_log_value(row.get("launch_wait", 0.0)),
                    _row_station(next_row) if next_row else "无",
                    self._fmt_vehicle_log_value(next_row.get("start", next_row.get("start_time", ""))) if next_row else "",
                    note,
                ])

        return columns, output

    def _copy_vehicle_log_to_clipboard(self, columns, rows):
        lines = ["\t".join(columns)]
        lines.extend("\t".join(str(value) for value in row) for row in rows)
        QApplication.clipboard().setText("\n".join(lines))
        self.status.showMessage("车辆日志已复制到剪贴板", 3000)

    def _filter_vehicle_log_rows(self, rows, car_text="", car_type="全部", station="全部", wait_filter="全部"):
        """筛选现有结构化日志，不重新运行排程。"""
        query = str(car_text or "").strip().lower().replace("car#", "").replace("#", "")
        selected_type = str(car_type or "全部").strip().upper()
        selected_station = str(station or "全部").strip()
        selected_wait = str(wait_filter or "全部").strip()

        excess_wait_keys = set()
        if selected_wait == "有节拍外等待":
            _, cause_rows = self._build_wait_cause_log_rows()
            for cause_row in cause_rows:
                car_no = str(cause_row[0]).replace("Car#", "").strip()
                waiting_station = str(cause_row[2]).strip()
                excess_wait_keys.add((car_no, waiting_station))

        filtered = []
        for row in rows or []:
            car_no = str(row[0]).strip()
            vehicle_type = str(row[1]).strip().upper()
            row_station = str(row[3]).strip()
            if query and query != car_no.lower():
                continue
            if selected_type != "全部" and vehicle_type != selected_type:
                continue
            if selected_station != "全部" and row_station != selected_station:
                continue
            if selected_wait == "有实际等待":
                try:
                    actual_wait = float(row[11] or 0.0) + float(row[12] or 0.0)
                except Exception:
                    actual_wait = 0.0
                if actual_wait <= 1e-9:
                    continue
            elif selected_wait == "有节拍外等待" and (car_no, row_station) not in excess_wait_keys:
                continue
            filtered.append(row)
        return filtered

    def _export_vehicle_log_csv(self, columns, rows, parent=None):
        if not rows:
            QMessageBox.information(parent or self, "导出车辆日志", "当前筛选结果为空，无可导出内容。")
            return
        path, _ = QFileDialog.getSaveFileName(
            parent or self,
            "导出当前筛选结果",
            "M-Line车辆日志_筛选结果.csv",
            "CSV (*.csv)",
        )
        if not path:
            return
        with open(path, "w", encoding="utf-8-sig", newline="") as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(columns)
            writer.writerows(rows)
        self.status.showMessage(f"车辆日志已导出：{path}", 5000)

    def _build_compact_vehicle_log_csv_rows(self, schedule_rows):
        """将当前筛选命中的车辆整理为一车一行CSV。"""
        if not schedule_rows:
            return [], []

        def _to_float(row, *keys):
            for key in keys:
                value = row.get(key)
                if value not in (None, ""):
                    try:
                        return float(value)
                    except Exception:
                        continue
            return 0.0

        def _to_int(value, default=0):
            try:
                return int(float(value))
            except Exception:
                return default

        def _fmt_seconds(value):
            return f"{float(value or 0.0):.1f}"

        def _station(row):
            return str(row.get("step_display", row.get("station", row.get("group", "岗位"))) or "岗位")

        grouped = {}
        for row in schedule_rows:
            car = row.get("car", row.get("car_no", row.get("car_index", "")))
            grouped.setdefault(car, []).append(row)

        try:
            target_takt = float(self.spn_target_takt.value())
        except Exception:
            target_takt = 0.0
        capacity_by_car = {
            int(item["car"]): item
            for item in compute_car_capacity_results(schedule_rows, target_takt)
        }

        columns = ["CAR", "TYPE", "IN(s)", "OUT(s)", "WAIT(s)", "FLOW(s)", "能力判断", "SEGMENTS"]
        output = []
        for car, car_rows in sorted(grouped.items(), key=lambda item: _to_int(item[0], 999999)):
            car_rows.sort(key=lambda row: (
                _to_float(row, "start", "start_time"),
                _to_int(row.get("step_seq", row.get("seq", 0))),
            ))
            first_start = _to_float(car_rows[0], "start", "start_time")
            final_finish = _to_float(car_rows[-1], "depart", "end", "svc_finish", "finish")
            total_wait = sum(
                _to_float(row, "launch_wait") + _to_float(row, "block_wait")
                for row in car_rows
            )
            car_type = str(
                car_rows[0].get("car_type", car_rows[0].get("duration_source", car_rows[0].get("vehicle_type", "")))
                or ""
            )
            capacity_result = capacity_by_car.get(_to_int(car), {})
            capacity_text = str(capacity_result.get("capacity_status", "未设定目标") or "未设定目标")
            if capacity_text == "能力超目标":
                capacity_text += f"：{capacity_result.get('over_capacity_station_text', '无')}"

            segments = []
            for row in car_rows:
                step_seq = row.get("step_seq", row.get("seq", ""))
                step_label = f"ST{step_seq}" if step_seq not in (None, "") else "ST?"
                segments.append(
                    f"{step_label} {_station(row)}("
                    f"开:{_fmt_seconds(_to_float(row, 'start', 'start_time'))}s "
                    f"加:{_fmt_seconds(_to_float(row, 'dur', 'duration'))}s "
                    f"等前:{_fmt_seconds(_to_float(row, 'launch_wait'))}s "
                    f"等后:{_fmt_seconds(_to_float(row, 'block_wait'))}s)"
                )

            output.append([
                f"Car#{car}",
                car_type,
                _fmt_seconds(first_start),
                _fmt_seconds(final_finish),
                _fmt_seconds(total_wait),
                _fmt_seconds(max(0.0, final_finish - first_start)),
                capacity_text,
                " | ".join(segments),
            ])
        return columns, output

    def _on_vehicle_summary_link_activated(self, link):
        if str(link or "") == "wait-cause-details":
            self._show_vehicle_log_placeholder(initial_tab="cause")

    def _build_wait_cause_log_rows(self):
        """整理当前统计范围内的节拍外等待真因明细。"""
        if not getattr(self, "last_schedule_rows", None):
            return [], []
        realtime = self._build_realtime_model_result(float(getattr(self, "sim_time", 0.0) or 0.0))
        details = realtime.get("cause_chain_details", []) or []
        columns = [
            "等待车辆", "车型", "等待发生工程", "等待开始(s)", "等待结束(s)", "本段等待(s)",
            "直接阻挡车辆", "直接阻挡工程", "直接阻挡资源", "真因车辆", "等待真因", "阻挡证据链",
        ]
        output = []
        for item in sorted(details, key=lambda value: (float(value.get("wait_start", 0.0)), int(value.get("car", 0) or 0))):
            chain_parts = []
            for node in item.get("chain", []) or []:
                waiting_car = node.get("waiting_car")
                blocker_car = node.get("blocker_car")
                station = str(node.get("blocked_station", "") or "未知工程")
                if blocker_car is None:
                    chain_parts.append(f"Car#{waiting_car} 等待 {station}（未解析）")
                else:
                    chain_parts.append(f"Car#{waiting_car} 等 Car#{blocker_car} 释放 {station}")
            output.append([
                f"Car#{item.get('car', '')}",
                str(item.get("car_type", "") or ""),
                str(item.get("waiting_station", "") or ""),
                self._fmt_vehicle_log_value(item.get("wait_start")),
                self._fmt_vehicle_log_value(item.get("wait_end")),
                self._fmt_vehicle_log_value(item.get("wait_time")),
                "" if item.get("direct_blocker_car") is None else f"Car#{item.get('direct_blocker_car')}",
                str(item.get("direct_blocking_station", "") or ""),
                str(item.get("direct_blocking_resource", "") or ""),
                "" if item.get("terminal_car") is None else f"Car#{item.get('terminal_car')}",
                str(item.get("terminal_cause", "") or ""),
                " → ".join(chain_parts),
            ])
        return columns, output

    def _show_vehicle_log_placeholder(self, initial_tab="vehicle"):
        rows = getattr(self, "last_schedule_rows", []) or []
        if not rows:
            QMessageBox.information(self, "车辆明细 / 调试日志", "请先点击“分析当前排程”生成车辆日志。")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("车辆明细 / 调试日志")
        dialog.resize(1000, 650)

        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        info = QLabel(
            "车辆明细 / 调试日志：车辆过程保持一台车一行，连续显示完整工序流转与等待；"
            "筛选用于快速定位车辆，CSV导出与界面一致，保持一台车一行。"
            "其中投入等待表示进入首工程前的等待，完工后等待表示加工完成后等待下一工程接收。"
        )
        info.setWordWrap(True)
        info.setStyleSheet("color:#475569;font-size:12px;")
        layout.addWidget(info)

        tabs = QTabWidget(dialog)
        layout.addWidget(tabs, 1)

        vehicle_page = QWidget(dialog)
        vehicle_layout = QVBoxLayout(vehicle_page)
        vehicle_layout.setContentsMargins(0, 0, 0, 0)
        vehicle_layout.setSpacing(8)

        filter_row = QHBoxLayout()
        filter_row.setSpacing(6)
        filter_row.addWidget(QLabel("车号：", vehicle_page))
        car_filter = QLineEdit(vehicle_page)
        car_filter.setPlaceholderText("例如 11 或 Car#11")
        car_filter.setMaximumWidth(150)
        filter_row.addWidget(car_filter)
        filter_row.addWidget(QLabel("车型：", vehicle_page))
        type_filter = QComboBox(vehicle_page)
        type_filter.addItems(["全部", "A", "B", "C"])
        filter_row.addWidget(type_filter)
        filter_row.addWidget(QLabel("工位：", vehicle_page))
        station_filter = QComboBox(vehicle_page)
        vehicle_columns, all_vehicle_rows = self._build_vehicle_log_rows()
        station_filter.addItem("全部")
        station_filter.addItems(sorted({str(item[3]) for item in all_vehicle_rows if str(item[3]).strip()}))
        station_filter.setMinimumWidth(170)
        filter_row.addWidget(station_filter)
        filter_row.addWidget(QLabel("等待：", vehicle_page))
        wait_filter = QComboBox(vehicle_page)
        wait_filter.addItems(["全部", "有实际等待", "有节拍外等待"])
        filter_row.addWidget(wait_filter)
        result_count = QLabel(vehicle_page)
        result_count.setStyleSheet("color:#475569;font-size:12px;")
        filter_row.addWidget(result_count)
        filter_row.addStretch()
        vehicle_layout.addLayout(filter_row)

        text_edit = QPlainTextEdit(vehicle_page)
        text_edit.setReadOnly(True)
        text_edit.setStyleSheet(
            "QPlainTextEdit{"
            "font-family: Menlo, Monaco, Consolas, monospace;"
            "font-size: 12px;"
            "color: #0f172a;"
            "background: #f8fafc;"
            "border: 1px solid #cbd5e1;"
            "}"
        )
        vehicle_layout.addWidget(text_edit, 1)
        tabs.addTab(vehicle_page, "车辆过程")

        cause_columns, cause_rows = self._build_wait_cause_log_rows()
        cause_table = QTableWidget(len(cause_rows), len(cause_columns), dialog)
        cause_table.setHorizontalHeaderLabels(cause_columns)
        cause_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        cause_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        cause_table.setAlternatingRowColors(True)
        cause_table.verticalHeader().setVisible(False)
        cause_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        cause_table.horizontalHeader().setStretchLastSection(True)
        for row_index, values in enumerate(cause_rows):
            for column_index, value in enumerate(values):
                cause_table.setItem(row_index, column_index, QTableWidgetItem(str(value)))
        tabs.addTab(cause_table, f"等待真因（{len(cause_rows)}段）")
        if initial_tab == "cause":
            tabs.setCurrentWidget(cause_table)

        button_row = QHBoxLayout()
        button_row.addStretch()
        btn_export = QPushButton("导出当前筛选", dialog)
        btn_copy = QPushButton("复制当前结果", dialog)
        btn_close = QPushButton("关闭", dialog)
        button_row.addWidget(btn_export)
        button_row.addWidget(btn_copy)
        button_row.addWidget(btn_close)
        layout.addLayout(button_row)

        current_vehicle_rows = []
        current_vehicle_text = ""

        def _refresh_vehicle_log():
            nonlocal current_vehicle_rows, current_vehicle_text
            current_vehicle_rows = self._filter_vehicle_log_rows(
                all_vehicle_rows,
                car_filter.text(),
                type_filter.currentText(),
                station_filter.currentText(),
                wait_filter.currentText(),
            )
            selected_cars = {str(item[0]) for item in current_vehicle_rows}
            selected_rows = [
                row for row in rows
                if str(row.get("car", row.get("car_no", row.get("car_index", "")))) in selected_cars
            ]
            current_vehicle_text = self._build_schedule_debug_log(selected_rows, limit=9999)
            text_edit.setPlainText(current_vehicle_text)
            result_count.setText(f"{len(selected_cars)}台")

        def _copy_current_tab():
            if tabs.currentWidget() is cause_table:
                self._copy_vehicle_log_to_clipboard(cause_columns, cause_rows)
            else:
                QApplication.clipboard().setText(current_vehicle_text)
                self.status.showMessage("车辆日志已复制到剪贴板", 3000)

        def _export_current_tab():
            if tabs.currentWidget() is cause_table:
                self._export_vehicle_log_csv(cause_columns, cause_rows, dialog)
            else:
                selected_cars = {str(item[0]) for item in current_vehicle_rows}
                selected_rows = [
                    row for row in rows
                    if str(row.get("car", row.get("car_no", row.get("car_index", "")))) in selected_cars
                ]
                compact_columns, compact_rows = self._build_compact_vehicle_log_csv_rows(selected_rows)
                self._export_vehicle_log_csv(compact_columns, compact_rows, dialog)

        car_filter.textChanged.connect(_refresh_vehicle_log)
        type_filter.currentTextChanged.connect(_refresh_vehicle_log)
        station_filter.currentTextChanged.connect(_refresh_vehicle_log)
        wait_filter.currentTextChanged.connect(_refresh_vehicle_log)
        btn_export.clicked.connect(_export_current_tab)
        btn_copy.clicked.connect(_copy_current_tab)
        btn_close.clicked.connect(dialog.accept)

        _refresh_vehicle_log()

        dialog.exec()

    def _on_simulation_tick(self):
        """仿真计时推进。后续车辆绘制会基于 sim_time 刷新画面。"""
        total = float(getattr(self, "last_max_finish", 0.0) or 0.0)
        if total <= 0:
            self._pause_simulation()
            return
        self.sim_time = min(total, float(getattr(self, "sim_time", 0.0) or 0.0) + 0.1 * self._sim_speed_value())
        self._update_sim_time_label()
        self._update_sim_total_wait_label()
        self._update_sim_view()
        self._draw_sim_scene()
        self._update_realtime_model_result()
        if self.sim_time >= total:
            self._pause_simulation()


    def _sim_row_value(self, row: dict, *keys, default=None):
        """从排程行中兼容读取字段。"""
        for key in keys:
            if key in row and row.get(key) is not None:
                return row.get(key)
        return default

    def _sim_row_start(self, row: dict) -> float:
        value = self._sim_row_value(row, "start", "start_time", "begin", "in", "in_time", default=0.0)
        try:
            return float(value or 0.0)
        except Exception:
            return 0.0

    def _sim_row_end(self, row: dict) -> float:
        value = self._sim_row_value(row, "end", "finish", "finish_time", "out", "out_time", default=None)
        if value is not None:
            try:
                return float(value or 0.0)
            except Exception:
                pass
        start = self._sim_row_start(row)
        dur = self._sim_row_value(row, "dur", "duration", "process", "process_time", default=0.0)
        try:
            return start + float(dur or 0.0)
        except Exception:
            return start

    def _sim_row_service_finish(self, row: dict) -> float:
        """读取当前工程加工完成时间，优先使用 svc_finish。"""
        value = self._sim_row_value(
            row,
            "svc_finish",
            "finish",
            "finish_time",
            "end",
            default=None,
        )
        if value is not None:
            try:
                return float(value or 0.0)
            except Exception:
                pass
        return self._sim_row_end(row)

    def _sim_row_depart(self, row: dict) -> float:
        """读取车辆可离开当前工程的时间，优先使用 depart。"""
        value = self._sim_row_value(
            row,
            "depart",
            "end",
            "out",
            "out_time",
            "svc_finish",
            "finish",
            default=None,
        )
        if value is not None:
            try:
                return float(value or 0.0)
            except Exception:
                pass
        return max(self._sim_row_service_finish(row), self._sim_row_end(row))

    def _sim_row_wait_end(self, row: dict, next_row: dict | None = None) -> float:
        """
        等待阶段的显示终点。

        规则：
        - 优先停留在当前工程所在位置；
        - depart/end 只作为“仍停留在当前工程”的参考；
        - 如果下一工程尚未开始，current < next_start 时不允许提前跳到 next_row。
        """
        wait_start = self._sim_row_service_finish(row)
        wait_end = max(wait_start, self._sim_row_depart(row))
        if next_row is not None:
            wait_end = max(wait_end, self._sim_row_start(next_row))
        return wait_end

    def _sim_row_station(self, row: dict) -> str:
        return str(self._sim_row_value(row, "step_display", "station", "display", "name", "group", default="岗位") or "岗位")

    def _sim_row_car_label(self, row: dict) -> str:
        car = self._sim_row_value(row, "car", "car_no", "car_index", "idx", default="?")
        car_type = str(self._sim_row_value(row, "car_type", "type", "vehicle_type", default="") or "")
        if car_type:
            return f"Car#{car}({car_type})"
        return f"Car#{car}"
    
    def _sim_row_run_mode(self, row: dict) -> str:
        """兼容读取岗位运行方式。"""
        return str(self._sim_row_value(row, "run_mode", "mode", default="") or "")

    def _sim_row_line_no(self, row: dict) -> str:
        """兼容读取排程行线别，用于 v2-4 线别验证。"""
        return str(self._sim_row_value(row, "line_no", "line", "line_scope", default="") or "")

    def _sim_row_resource_key(self, row: dict) -> str:
        """兼容读取排程行资源 key。"""
        return str(self._sim_row_value(row, "resource_key", default="") or "")

    def _sim_row_block_wait(self, row: dict) -> float:
        """读取当前工程后方阻塞等待时间，仅用于 UI 保守显示。"""
        value = self._sim_row_value(row, "block_wait", default=0.0)
        try:
            return float(value or 0.0)
        except Exception:
            return 0.0

    def _sim_wait_is_blocking(self, row: dict, next_row: dict | None, occupied_resources: set[str]) -> bool:
        """
        UI 安全模式下，默认不把等待解释为“阻塞”。

        说明：
        - 当前项目的稳定诉求是保留新画布结构，但收缩显示层解释；
        - finish -> next_start 之间的时间差，不应由 UI 轻易推断成“被谁阻塞”；
        - process_over_takt_root_text 属于模型结果说明，也不用于实时阻塞判定。

        因此这里保守返回 False，等待文案统一走“等待Xs”。
        未来如果 rows 中新增了明确、可靠的实时阻塞字段，再单独恢复更细判断。
        """
        return False

    def _sim_wait_label(self, row: dict, next_row: dict | None, current: float, wait_end: float, occupied_resources: set[str]) -> str:
        """统一生成等待文案；安全模式下默认只显示等待时长。"""
        wait_remain = max(0.0, wait_end - current)
        return f"等待{wait_remain:.1f}s"

    def _build_realtime_blocking_hint(self, current_time: float) -> str:
        """按已发生 block_wait 汇总实时阻塞工程，根因优先归属下一有效工程。"""
        try:
            current = float(current_time or 0.0)
        except Exception:
            current = 0.0

        blocking_by_station = {}
        for _, car_segments in self._sim_car_rows().items():
            if not car_segments:
                continue
            ordered = sorted(
                car_segments,
                key=lambda row: (
                    self._sim_row_start(row),
                    self._sim_row_depart(row),
                    self._sim_row_station(row),
                ),
            )
            for idx, row in enumerate(ordered):
                block_wait = self._sim_row_block_wait(row)
                if block_wait <= 0:
                    continue

                wait_start = self._sim_row_service_finish(row)
                wait_end = self._sim_row_depart(row)
                next_row = ordered[idx + 1] if idx + 1 < len(ordered) else None
                if next_row is not None:
                    wait_end = max(wait_end, self._sim_row_start(next_row))

                if current <= wait_start:
                    occurred = 0.0
                elif current < wait_end:
                    occurred = min(block_wait, max(0.0, current - wait_start))
                else:
                    occurred = block_wait

                if occurred <= 0:
                    continue

                blocking_station = self._sim_row_station(next_row) if next_row is not None else self._sim_row_station(row)
                blocking_by_station[blocking_station] = blocking_by_station.get(blocking_station, 0.0) + occurred

        if not blocking_by_station:
            return "阻塞工程：无"

        def _fmt_seconds(value):
            try:
                seconds = float(value or 0.0)
            except Exception:
                seconds = 0.0
            if seconds >= 10 or abs(seconds - round(seconds)) < 1e-9:
                return f"{int(round(seconds))}s"
            return f"{seconds:.1f}s"

        sorted_items = sorted(
            blocking_by_station.items(),
            key=lambda item: float(item[1]),
            reverse=True,
        )
        parts = [
            f"{station} {_fmt_seconds(seconds)}"
            for station, seconds in sorted_items[:2]
        ]
        suffix = " 等" if len(sorted_items) > 2 else ""
        return f"阻塞工程：{'，'.join(parts)}{suffix}"

    def _build_schedule_debug_log(self, rows, limit: int = 200) -> str:
        """按车辆维度输出一车一行的排程流转日志。"""
        if not rows:
            return "暂无排程 rows。"

        def _fmt_sec(value):
            try:
                return f"{float(value or 0.0):.1f}"
            except Exception:
                return "0.0"

        def _float_value(row, *keys, default=0.0):
            for key in keys:
                if key in row and row.get(key) not in (None, ""):
                    try:
                        return float(row.get(key) or 0.0)
                    except Exception:
                        return default
            return default

        def _int_sort_value(value):
            try:
                return 0, int(value)
            except Exception:
                return 1, str(value)

        def _row_car(row):
            return row.get("car", row.get("car_no", row.get("car_index", "?")))

        def _row_car_type(row):
            return str(row.get("car_type", row.get("duration_source", row.get("vehicle_type", ""))) or "")

        def _row_step_seq(row):
            return row.get("step_seq", row.get("seq", ""))

        def _row_station(row):
            return str(row.get("step_display", row.get("station", row.get("group", "岗位"))) or "岗位")

        def _row_start(row):
            return _float_value(row, "start", "start_time")

        def _row_depart(row):
            return _float_value(row, "depart", "end", "svc_finish", "finish")

        def _row_dur(row):
            return _float_value(row, "dur", "duration")

        def _row_launch_wait(row):
            return _float_value(row, "launch_wait")

        def _row_block_wait(row):
            return _float_value(row, "block_wait")

        grouped = {}
        for row in rows or []:
            grouped.setdefault(_row_car(row), []).append(row)

        for car_rows in grouped.values():
            car_rows.sort(key=lambda r: (_row_start(r), _int_sort_value(_row_step_seq(r))))

        try:
            target_takt = float(self.spn_target_takt.value()) if hasattr(self, "spn_target_takt") else 0.0
        except Exception:
            target_takt = 0.0

        capacity_results = compute_car_capacity_results(rows, target_takt)
        capacity_by_car = {int(item["car"]): item for item in capacity_results}

        lines = [
            "车辆明细 / 调试日志（按车辆维度）",
            "说明：能力判断按车型工时÷有效设备能力与目标节拍比较；工时为0的流转节点不参加判断。",
            "字段说明：IN=实际投车时间，OUT=下线完成时间，WAIT=单车总等待，FLOW=单车贯通时间；"
            "SEGMENTS中的等前=投入等待，等后=加工完成后等待下一工程接收。",
            "-" * 120,
            f"{'CAR':<8}{'TYPE':<6}{'IN(s)':>10}{'OUT(s)':>10}{'WAIT(s)':>10}{'FLOW(s)':>10}  {'能力判断':<18}  SEGMENTS",
            "-" * 120,
        ]

        sorted_car_items = sorted(grouped.items(), key=lambda item: _int_sort_value(item[0]))
        for car, car_rows in sorted_car_items[:limit]:
            if not car_rows:
                continue

            car_type = _row_car_type(car_rows[0])
            first_start = _row_start(car_rows[0])
            final_finish = _row_depart(car_rows[-1])
            total_wait = sum(_row_launch_wait(row) + _row_block_wait(row) for row in car_rows)
            flow_time = max(0.0, final_finish - first_start)
            try:
                car_number = int(float(car))
            except Exception:
                car_number = 0
            capacity_result = capacity_by_car.get(car_number, {})
            car_result = str(capacity_result.get("capacity_status", "未设定目标") or "未设定目标")
            over_station_text = str(capacity_result.get("over_capacity_station_text", "无") or "无")
            if car_result == "能力超目标":
                car_result = f"能力超目标：{over_station_text}"

            step_parts = []
            remarks = []
            previous_seq = None
            for row in car_rows:
                step_seq = _row_step_seq(row)
                station = _row_station(row)
                start = _row_start(row)
                dur = _row_dur(row)
                launch_wait = _row_launch_wait(row)
                block_wait = _row_block_wait(row)
                step_label = f"ST{step_seq}" if step_seq not in (None, "") else "ST?"
                step_parts.append(
                    f"{step_label} {station}(开:{_fmt_sec(start)}s 加:{_fmt_sec(dur)}s "
                    f"等前:{_fmt_sec(launch_wait)}s 等后:{_fmt_sec(block_wait)}s)"
                )

                try:
                    current_seq = int(step_seq)
                except Exception:
                    current_seq = None
                if previous_seq is not None and current_seq is not None and current_seq - previous_seq > 1:
                    skipped = ",".join(str(seq) for seq in range(previous_seq + 1, current_seq))
                    remarks.append(f"跳过中间岗位：{skipped}")
                if current_seq is not None:
                    previous_seq = current_seq

                for key in ("remark", "remarks", "note", "备注"):
                    value = str(row.get(key, "") or "").strip()
                    if value:
                        remarks.append(value)

            unique_remarks = []
            for remark in remarks:
                if remark not in unique_remarks:
                    unique_remarks.append(remark)
            remark_text = f" 备注：{'；'.join(unique_remarks)}" if unique_remarks else ""

            lines.append(
                f"{'Car#' + str(car):<8}"
                f"{(car_type or '—'):<6}"
                f"{_fmt_sec(first_start):>10}"
                f"{_fmt_sec(final_finish):>10}"
                f"{_fmt_sec(total_wait):>10}"
                f"{_fmt_sec(flow_time):>10}"
                f"  {car_result:<18}  "
                + " | ".join(step_parts)
                + remark_text
            )

        if len(sorted_car_items) > limit:
            lines.append(f"……仅显示前 {limit} 台车，共 {len(sorted_car_items)} 台。")

        return "\n".join(lines)
    
    def _sim_car_key(self, row: dict):
        """按车辆编号聚合排程段。"""
        return self._sim_row_value(row, "car", "car_no", "car_index", "idx", default="?")
    
    def _sim_car_rows(self):
        """将排程行按车辆聚合，并按开始时间排序。"""
        grouped = {}
        for row in getattr(self, "last_schedule_rows", []) or []:
            key = self._sim_car_key(row)
            grouped.setdefault(key, []).append(row)
        for key in grouped:
            grouped[key].sort(key=lambda r: self._sim_row_start(r))
        return grouped

    def _sim_station_names(self):
        """优先按岗位矩阵 seq 顺序提取岗位名。"""
        defs = getattr(self, "current_defs", None) or []
        if defs:
            ordered = []
            seen = set()
            sortable_defs = []
            for idx, item in enumerate(defs):
                try:
                    seq_value = int(item.get("seq", idx + 1) or (idx + 1))
                except Exception:
                    seq_value = idx + 1
                name = str(item.get("display") or item.get("group") or "岗位").strip() or "岗位"
                sortable_defs.append((seq_value, idx, name))
            sortable_defs.sort(key=lambda x: (x[0], x[1]))
            for _, _, name in sortable_defs:
                if name not in seen:
                    ordered.append(name)
                    seen.add(name)
            if ordered:
                return ordered

        names = []
        seen = set()
        for row in getattr(self, "last_schedule_rows", []) or []:
            name = self._sim_row_station(row)
            if name not in seen:
                names.append(name)
                seen.add(name)
        return names

    def _sim_station_defs(self):
        """返回动画展示使用的岗位顺序与标签，优先使用 current_defs 的 seq。"""
        defs = getattr(self, "current_defs", None) or []
        if defs:
            ordered = []
            seen = set()
            sortable_defs = []
            for idx, item in enumerate(defs):
                try:
                    seq_value = int(item.get("seq", idx + 1) or (idx + 1))
                except Exception:
                    seq_value = idx + 1
                name = str(item.get("display") or item.get("group") or "岗位").strip() or "岗位"
                sortable_defs.append((seq_value, idx, name, item))
            sortable_defs.sort(key=lambda x: (x[0], x[1]))
            for seq_value, _, name, source in sortable_defs:
                if name not in seen:
                    ordered.append({
                        "seq": seq_value,
                        "name": name,
                        "device_count": source.get("device_count"),
                        "line_scope": source.get("line_scope"),
                        "run_mode": source.get("run_mode"),
                    })
                    seen.add(name)
            if ordered:
                return ordered

        names = self._sim_station_names()
        return [{"seq": idx + 1, "name": name} for idx, name in enumerate(names)]

    def _update_sim_view(self):
        """刷新简易仿真画面。当前阶段只展示当前加工中的车辆摘要。"""
        if not hasattr(self, "lbl_sim_view"):
            return

        rows = getattr(self, "last_schedule_rows", []) or []
        if not rows:
            self.lbl_sim_view.setText("仿真画面：请先点击『分析当前排程』。")
            return

        current = float(getattr(self, "sim_time", 0.0) or 0.0)

        active_rows = []
        for row in rows:
            start = self._sim_row_start(row)
            svc_finish = self._sim_row_service_finish(row)
            if start <= current < svc_finish:
                active_rows.append((start, svc_finish, row))

        active_rows.sort(key=lambda x: (self._sim_row_station(x[2]), x[0]))
        waiting_rows = []
        car_rows = self._sim_car_rows()
        for _, car_segments in car_rows.items():
            if not car_segments:
                continue
            for idx, seg in enumerate(car_segments):
                next_seg = car_segments[idx + 1] if idx + 1 < len(car_segments) else None
                wait_start = self._sim_row_service_finish(seg)
                wait_end = self._sim_row_wait_end(seg, next_seg)
                if wait_start <= current < wait_end:
                    waiting_rows.append((wait_start, wait_end, seg, next_seg))
                    break

        waiting_rows.sort(key=lambda x: (self._sim_row_station(x[2]), x[0]))

        def _escape_lines(lines):
            return [escape(str(line)) for line in lines if str(line).strip()]

        def _compress_lines(lines, max_lines):
            cleaned = [str(line).strip() for line in lines if str(line).strip()]
            if not cleaned:
                return ["无"]
            if len(cleaned) <= max_lines:
                return cleaned
            kept = cleaned[:max_lines]
            overflow = len(cleaned) - max_lines
            kept[-1] = f"{kept[-1]} / ……还有 {overflow} 台"
            return kept

        def _slot_html(lines):
            escaped_lines = _escape_lines(lines)
            while len(escaped_lines) < 2:
                escaped_lines.append("&#12288;")
            return "".join(
                f"<div style='margin:0;padding:0;line-height:1.28;'>{line}</div>"
                for line in escaped_lines[:2]
            )

        active_lines = []
        if active_rows:
            max_active_rows = 3
            for start, end, row in active_rows[:max_active_rows]:
                car_label = self._sim_row_car_label(row)
                station = self._sim_row_station(row)
                remain = max(0.0, end - current)
                line_no = self._sim_row_line_no(row)
                detail_parts = [f"{car_label} @ {station}"]
                if line_no:
                    detail_parts.append(str(line_no))
                if remain > 0:
                    detail_parts.append(f"剩余{remain:.1f}s")
                active_lines.append("｜".join(detail_parts))
            if len(active_rows) > max_active_rows:
                active_lines.append(f"……还有 {len(active_rows) - max_active_rows} 台")
        else:
            active_lines.append("无")
        active_lines = _compress_lines(active_lines, 2)

        waiting_lines = []
        if waiting_rows:
            max_wait_rows = 2
            for wait_start, wait_end, row, next_row in waiting_rows[:max_wait_rows]:
                car_label = self._sim_row_car_label(row)
                station = self._sim_row_station(row)
                line_no = self._sim_row_line_no(row)
                detail_parts = [f"{car_label} @ {station}"]
                if line_no:
                    detail_parts.append(str(line_no))
                wait_text = self._sim_wait_label(row, next_row, current, wait_end, set())
                if wait_text:
                    detail_parts.append(wait_text)
                waiting_lines.append("｜".join(detail_parts))
            if len(waiting_rows) > max_wait_rows:
                waiting_lines.append(f"……还有 {len(waiting_rows) - max_wait_rows} 台")
        else:
            waiting_lines.append("无")
        waiting_lines = _compress_lines(waiting_lines, 2)

        model_hint_text = self._build_realtime_blocking_hint(current)
        blocking_lines = _compress_lines(model_hint_text.splitlines() or [model_hint_text], 2)

        active_html = _slot_html(active_lines)
        waiting_html = _slot_html(waiting_lines)
        blocking_html = _slot_html(blocking_lines)

        html = (
            "<table width='100%' cellspacing='0' cellpadding='0'>"
            "<tr>"
            "<td width='33%' style='width:33%;vertical-align:top;padding-right:8px;'>"
            "<div style='font-size:12px;font-weight:700;color:#334155;margin-bottom:2px;'>加工中车辆</div>"
            f"<div style='font-size:12px;line-height:1.24;color:#334155;'>{active_html}</div>"
            "</td>"
            "<td width='34%' style='width:34%;vertical-align:top;padding:0 8px;border-left:1px solid #e2e8f0;'>"
            "<div style='font-size:12px;font-weight:700;color:#334155;margin-bottom:2px;'>等待中</div>"
            f"<div style='font-size:12px;line-height:1.24;color:#334155;'>{waiting_html}</div>"
            "</td>"
            "<td width='33%' style='width:33%;vertical-align:top;padding-left:8px;border-left:1px solid #e2e8f0;'>"
            "<div style='font-size:12px;font-weight:700;color:#334155;margin-bottom:2px;'>当前等待位置</div>"
            f"<div style='font-size:12px;line-height:1.24;color:#334155;'>{blocking_html}</div>"
            "</td>"
            "</tr>"
            "</table>"
        )
        self.lbl_sim_view.setText(html)

    # ------------- 多车组合票：动作 ------------- #
    def add_row(self):
        r = self.tbl.rowCount()
        self.tbl.insertRow(r)

        # 序号
        self.tbl.setItem(r, 0, QTableWidgetItem(str(r + 1)))
        # 工程名称
        self.tbl.setItem(r, 1, QTableWidgetItem(""))

        # 设备数量：v2 主线字段。默认 2，代表双线双设备。
        device_count_cb = QComboBox(self.tbl)
        device_count_cb.addItems(["1", "2"])
        device_count_cb.setCurrentText("2")
        self.tbl.setCellWidget(r, 2, device_count_cb)

        # 所属线别：v2 主线字段。
        line_scope_cb = QComboBox(self.tbl)
        line_scope_cb.addItems(["1号线", "2号线", "双线", "双线共用"])
        line_scope_cb.setCurrentText("双线")
        self.tbl.setCellWidget(r, 3, line_scope_cb)

        # 岗位设备
        self.tbl.setItem(r, 4, QTableWidgetItem(""))
        # A / B / C 工时
        self.tbl.setItem(r, 5, QTableWidgetItem(""))
        self.tbl.setItem(r, 6, QTableWidgetItem(""))
        self.tbl.setItem(r, 7, QTableWidgetItem(""))

        def _sync_line_scope():
            if device_count_cb.currentText() == "2":
                line_scope_cb.setCurrentText("双线")
                line_scope_cb.setEnabled(False)
            else:
                line_scope_cb.setEnabled(True)
                if line_scope_cb.currentText() == "双线":
                    line_scope_cb.setCurrentText("1号线")

        device_count_cb.currentTextChanged.connect(_sync_line_scope)
        _sync_line_scope()

    STATION_MATRIX_TEMPLATE_NAME = "M-Line岗位矩阵模板"
    STATION_MATRIX_TEMPLATE_VERSION = "v2.9"
    STATION_MATRIX_HEADERS = [
        "序号", "工程名称", "设备数量", "所属线别",
        "岗位设备", "A工时", "B工时", "C工时",
    ]

    def _collect_station_matrix_rows(self):
        """读取当前岗位矩阵的原始显示值。"""
        output = []
        for row_index in range(self.tbl.rowCount()):
            values = []
            for column_index in range(8):
                widget = self.tbl.cellWidget(row_index, column_index)
                if isinstance(widget, QComboBox):
                    value = widget.currentText().strip()
                else:
                    item = self.tbl.item(row_index, column_index)
                    value = item.text().strip() if item else ""
                values.append(value)
            if any(values[1:]):
                output.append(values)
        return output

    def _validate_station_matrix_rows(self, rows):
        """校验并标准化Excel或界面中的岗位矩阵。"""
        normalized = []
        seen_names = set()
        for source_index, values in enumerate(rows or [], start=1):
            row = [str(value if value is not None else "").strip() for value in list(values)[:8]]
            row.extend([""] * (8 - len(row)))
            if not any(row):
                continue

            excel_row = source_index + 3
            try:
                seq_number = float(row[0])
                seq = int(seq_number)
                if seq_number != seq or seq <= 0:
                    raise ValueError
            except Exception:
                raise ValueError(f"第{excel_row}行『序号』必须为正整数：{row[0] or '空'}")
            expected_seq = len(normalized) + 1
            if seq != expected_seq:
                raise ValueError(f"第{excel_row}行『序号』应为{expected_seq}，当前为{seq}。")

            name = row[1]
            if not name:
                raise ValueError(f"第{excel_row}行『工程名称』不能为空。")
            if name in seen_names:
                raise ValueError(f"第{excel_row}行『工程名称』重复：{name}。")
            seen_names.add(name)

            try:
                device_number = float(row[2])
                device_count = int(device_number)
                if device_number != device_count or device_count not in (1, 2):
                    raise ValueError
            except Exception:
                raise ValueError(f"第{excel_row}行『设备数量』只能为1或2：{row[2] or '空'}")

            line_scope = row[3]
            allowed_scopes = {"1号线", "2号线", "双线", "双线共用"}
            if line_scope not in allowed_scopes:
                raise ValueError(f"第{excel_row}行『所属线别』无效：{line_scope or '空'}")
            if device_count == 2 and line_scope != "双线":
                raise ValueError(f"第{excel_row}行设备数量为2时，所属线别必须为『双线』。")
            if device_count == 1 and line_scope == "双线":
                raise ValueError(f"第{excel_row}行设备数量为1时，所属线别不能为『双线』。")

            if not row[4]:
                raise ValueError(f"第{excel_row}行『岗位设备』不能为空。")

            durations = []
            for column_index, label in zip((5, 6, 7), ("A工时", "B工时", "C工时")):
                try:
                    duration = float(row[column_index])
                except Exception:
                    raise ValueError(f"第{excel_row}行『{label}』必须为数字：{row[column_index] or '空'}")
                if duration < 0:
                    raise ValueError(f"第{excel_row}行『{label}』不能小于0。")
                durations.append(self._fmt_vehicle_log_value(duration))

            normalized.append([
                str(seq), name, str(device_count), line_scope, row[4], *durations,
            ])

        if not normalized:
            raise ValueError("岗位矩阵为空，请至少填写一行有效数据。")
        return normalized

    def _write_station_matrix_xlsx(self, path, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "岗位矩阵"
        sheet["A1"] = self.STATION_MATRIX_TEMPLATE_NAME
        sheet["B1"] = self.STATION_MATRIX_TEMPLATE_VERSION
        sheet["A2"] = "工时为0表示车辆经过该工程，但不进行作业。"
        for column_index, header in enumerate(self.STATION_MATRIX_HEADERS, start=1):
            sheet.cell(row=3, column=column_index, value=header)
        for row_index, values in enumerate(rows, start=4):
            for column_index, value in enumerate(values, start=1):
                if column_index in (1, 3):
                    cell_value = int(value)
                elif column_index in (6, 7, 8):
                    cell_value = float(value)
                else:
                    cell_value = value
                sheet.cell(row=row_index, column=column_index, value=cell_value)
        widths = [10, 24, 12, 14, 24, 12, 12, 12]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        sheet.freeze_panes = "A4"
        workbook.save(path)

    def _read_station_matrix_xlsx(self, path):
        workbook = load_workbook(path, data_only=True, read_only=True)
        if "岗位矩阵" not in workbook.sheetnames:
            raise ValueError("未找到『岗位矩阵』工作表。")
        sheet = workbook["岗位矩阵"]
        if str(sheet["A1"].value or "").strip() != self.STATION_MATRIX_TEMPLATE_NAME:
            raise ValueError("该文件不是M-Line岗位矩阵模板。")
        version = str(sheet["B1"].value or "").strip()
        if version != self.STATION_MATRIX_TEMPLATE_VERSION:
            raise ValueError(f"模板版本不支持：{version or '未标识'}。")
        headers = [str(sheet.cell(row=3, column=index).value or "").strip() for index in range(1, 9)]
        if headers != self.STATION_MATRIX_HEADERS:
            raise ValueError("岗位矩阵表头不完整或顺序不正确。")

        raw_rows = []
        blank_seen = False
        for row_index in range(4, sheet.max_row + 1):
            values = [sheet.cell(row=row_index, column=index).value for index in range(1, 9)]
            if not any(value not in (None, "") for value in values):
                blank_seen = bool(raw_rows)
                continue
            if blank_seen:
                raise ValueError(f"第{row_index}行之前存在空白行，请删除中间空行后重试。")
            raw_rows.append(values)
        return version, self._validate_station_matrix_rows(raw_rows)

    def _apply_station_matrix_rows(self, rows):
        self.tbl.setRowCount(0)
        for values in rows:
            self.add_row()
            row_index = self.tbl.rowCount() - 1
            self.tbl.setItem(row_index, 0, QTableWidgetItem(values[0]))
            self.tbl.setItem(row_index, 1, QTableWidgetItem(values[1]))
            device_widget = self.tbl.cellWidget(row_index, 2)
            line_widget = self.tbl.cellWidget(row_index, 3)
            device_widget.setCurrentText(values[2])
            line_widget.setCurrentText(values[3])
            self.tbl.setItem(row_index, 4, QTableWidgetItem(values[4]))
            self.tbl.setItem(row_index, 5, QTableWidgetItem(values[5]))
            self.tbl.setItem(row_index, 6, QTableWidgetItem(values[6]))
            self.tbl.setItem(row_index, 7, QTableWidgetItem(values[7]))
        self._invalidate_frozen_vehicle_sequence()

    def _station_matrix_change_counts(self, old_rows, new_rows):
        old_by_seq = {row[0]: row for row in old_rows}
        new_by_seq = {row[0]: row for row in new_rows}
        added = len(set(new_by_seq) - set(old_by_seq))
        removed = len(set(old_by_seq) - set(new_by_seq))
        changed = sum(
            old_by_seq[key] != new_by_seq[key]
            for key in set(old_by_seq) & set(new_by_seq)
        )
        return added, changed, removed

    def _confirm_station_matrix_import(self, file_name, version, rows):
        current_rows = self._collect_station_matrix_rows()
        try:
            current_rows = self._validate_station_matrix_rows(current_rows) if current_rows else []
        except Exception:
            current_rows = self._collect_station_matrix_rows()
        added, changed, removed = self._station_matrix_change_counts(current_rows, rows)

        dialog = QDialog(self)
        dialog.setWindowTitle("岗位矩阵导入预览")
        dialog.resize(920, 520)
        layout = QVBoxLayout(dialog)
        summary = QLabel(
            f"文件：{file_name}<br>"
            f"模板版本：{version}｜导入{len(rows)}行｜"
            f"新增{added}行｜修改{changed}行｜删除{removed}行"
        )
        layout.addWidget(summary)
        preview = QTableWidget(len(rows), 8, dialog)
        preview.setHorizontalHeaderLabels(self.STATION_MATRIX_HEADERS)
        preview.setEditTriggers(QAbstractItemView.NoEditTriggers)
        preview.verticalHeader().setVisible(False)
        preview.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        for row_index, values in enumerate(rows):
            for column_index, value in enumerate(values):
                preview.setItem(row_index, column_index, QTableWidgetItem(str(value)))
        layout.addWidget(preview, 1)
        button_row = QHBoxLayout()
        button_row.addStretch()
        cancel_button = QPushButton("取消", dialog)
        confirm_button = QPushButton("确认导入", dialog)
        button_row.addWidget(cancel_button)
        button_row.addWidget(confirm_button)
        layout.addLayout(button_row)
        cancel_button.clicked.connect(dialog.reject)
        confirm_button.clicked.connect(dialog.accept)
        return dialog.exec() == QDialog.Accepted

    def _export_station_matrix(self):
        try:
            rows = self._validate_station_matrix_rows(self._collect_station_matrix_rows())
        except Exception as exc:
            QMessageBox.warning(self, "导出岗位矩阵", str(exc))
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "导出岗位矩阵", "M-Line岗位矩阵_v2.9.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            self._write_station_matrix_xlsx(path, rows)
        except Exception as exc:
            QMessageBox.critical(self, "导出失败", str(exc))
            return
        self.status.showMessage(f"岗位矩阵已导出：{path}", 5000)

    def _import_station_matrix(self):
        path, _ = QFileDialog.getOpenFileName(self, "导入岗位矩阵", "", "Excel (*.xlsx)")
        if not path:
            return
        try:
            version, rows = self._read_station_matrix_xlsx(path)
        except Exception as exc:
            QMessageBox.warning(self, "导入岗位矩阵", str(exc))
            return
        if not self._confirm_station_matrix_import(os.path.basename(path), version, rows):
            return

        snapshot = self._collect_station_matrix_rows()
        try:
            self._apply_station_matrix_rows(rows)
        except Exception as exc:
            self._apply_station_matrix_rows(snapshot)
            QMessageBox.critical(self, "导入失败", f"已恢复导入前的岗位矩阵。\n{exc}")
            return
        self.status.showMessage(f"已导入岗位矩阵：{os.path.basename(path)}", 5000)

    def _choose_color(self, row: int):
        dlg_col = QColorDialog.getColor(parent=self)
        if dlg_col.isValid():
            hex_code = dlg_col.name()
            btn = self.tbl.cellWidget(row, self.COL_C_TIME)
            btn.setStyleSheet(f"background:{hex_code};")
            self.tbl.item(row, self.COL_C_TIME).setData(Qt.UserRole, hex_code)

    def del_row(self):
        r = self.tbl.currentRow()
        if r >= 0:
            self.tbl.removeRow(r)

    # -------- 单人标准作业组合票：行操作 --------
    def add_single_row(self):
        """在单人作业手顺表中新增一行"""
        if not hasattr(self, "tbl_sw"):
            return

        current_rows = self.tbl_sw.rowCount()
        max_steps = getattr(self, "MAX_SINGLE_STEPS", 23)

        if current_rows >= max_steps:
            QMessageBox.warning(
                self,
                "已到模板上限",
                f"当前单人标准作业组合票模板最多支持 {max_steps} 行。\n"
                f"你现在已经添加了 {current_rows} 行，不能再继续新增。\n\n"
                "请合并部分区间或拆分为多张组合票后再导出。"
            )
            return

        r = current_rows
        self.tbl_sw.insertRow(r)
        # 顺序默认递增（组合票行号）
        self.tbl_sw.setItem(r, 0, QTableWidgetItem(str(r + 1)))
        # 作业名称A / B 先留空，让你填写
        self.tbl_sw.setItem(r, 1, QTableWidgetItem(""))
        self.tbl_sw.setItem(r, 2, QTableWidgetItem(""))
        # 手作业 / 自动 / 步行，默认 0
        self.tbl_sw.setItem(r, 3, QTableWidgetItem("0"))
        self.tbl_sw.setItem(r, 4, QTableWidgetItem("0"))
        self.tbl_sw.setItem(r, 5, QTableWidgetItem("0"))
        # 步行位置：默认“后置”
        pos_cb = QComboBox(self.tbl_sw)
        pos_cb.addItem("后置", userData="after")
        pos_cb.addItem("前置", userData="before")
        self.tbl_sw.setCellWidget(r, 6, pos_cb)
        # 自动在前/后（默认后置）
        auto_cb = QComboBox(self.tbl_sw)
        auto_cb.addItem("后置", userData="after")
        auto_cb.addItem("前置", userData="before")
        self.tbl_sw.setCellWidget(r, 7, auto_cb)

    def del_single_row(self):
        """删除单人作业手顺表中的选中行"""
        if not hasattr(self, "tbl_sw"):
            return
        r = self.tbl_sw.currentRow()
        if r >= 0:
            self.tbl_sw.removeRow(r)
        # 重写顺序列，让它保持 1,2,3,...
        for i in range(self.tbl_sw.rowCount()):
            item = self.tbl_sw.item(i, 0)
            if item is None:
                item = QTableWidgetItem()
                self.tbl_sw.setItem(i, 0, item)
            item.setText(str(i + 1))

    # -------- 单人标准作业组合票：数据收集 --------
    def _collect_single_inputs(self):
        """
        从单人作业手顺 Tab 中读取数据，并计算时间汇总。
        返回：
          project, part, worker, takt_sec, steps, totals
        其中：
          steps: [{seq, name, name_a, name_b, manual, auto, walk, walk_pos, auto_pos, duration, start, end}, ...]
          totals: {"manual": x, "auto": y, "walk": z, "total": t}
        """
        if not hasattr(self, "tbl_sw"):
            raise ValueError("单人作业手顺表尚未初始化")

        project = (self.ed_sw_project.text().strip() or "工程")
        part = self.ed_sw_part.text().strip()
        worker = self.ed_sw_worker.text().strip()
        takt_sec = int(self.spn_sw_takt.value())

        steps = []
        cur_time = 0.0
        total_manual = 0.0
        total_auto = 0.0
        total_walk = 0.0

        for r in range(self.tbl_sw.rowCount()):
            # 作业名称 A / B
            name_a_item = self.tbl_sw.item(r, 1)
            name_b_item = self.tbl_sw.item(r, 2)
            name_a = name_a_item.text().strip() if name_a_item else ""
            name_b = name_b_item.text().strip() if name_b_item else ""

            if not name_a and not name_b:
                # 两个都没填，当作空行，跳过
                continue

            # 导出时使用的显示名（A→B / 单独一个）
            if name_a and name_b:
                name = f"{name_a} → {name_b}"
            else:
                name = name_a or name_b

            def _get_time(col_idx: int) -> float:
                item = self.tbl_sw.item(r, col_idx)
                txt = item.text().strip() if item else ""
                if not txt:
                    return 0.0
                try:
                    return float(txt)
                except Exception:
                    raise ValueError(f"第 {r + 1} 行时间列（第 {col_idx + 1} 列）不是有效数字：{txt}")

            # 手作业 / 自动 / 步行时间列：3, 4, 5
            manual = _get_time(3)
            auto = _get_time(4)
            walk = _get_time(5)

            # 步行位置：前置/后置（默认后置）
            walk_pos = "after"
            pos_widget = self.tbl_sw.cellWidget(r, 6)
            if isinstance(pos_widget, QComboBox):
                walk_pos_data = pos_widget.currentData()
                if walk_pos_data in ("before", "after"):
                    walk_pos = walk_pos_data

            # 自动在前/后（默认后置）
            auto_pos = "after"
            auto_widget = self.tbl_sw.cellWidget(r, 7)
            if isinstance(auto_widget, QComboBox):
                auto_pos_data = auto_widget.currentData()
                if auto_pos_data in ("before", "after"):
                    auto_pos = auto_pos_data

            duration = manual + auto + walk
            if duration <= 0:
                raise ValueError(f"第 {r + 1} 行『{name}』的时间合计为 0，请填写手作业/自动/步行时间。")

            start = cur_time
            end = cur_time + duration
            cur_time = end

            total_manual += manual
            total_auto += auto
            total_walk += walk

            # 顺序列（如果用户改过，我们尽量读取）
            seq_item = self.tbl_sw.item(r, 0)
            try:
                seq = int(seq_item.text()) if seq_item and seq_item.text().strip() else len(steps) + 1
            except Exception:
                seq = len(steps) + 1

            steps.append({
                "seq": seq,
                "name": name,       # A→B 组合显示名（保留）
                "name_a": name_a,   # 原始作业名称A
                "name_b": name_b,   # 原始作业名称B
                "manual": manual,
                "auto": auto,
                "walk": walk,
                "walk_pos": walk_pos,  # 步行在前/后
                "auto_pos": auto_pos,  # 自动在前/后
                "duration": duration,
                "start": start,
                "end": end,
            })

        # 行数上限检查：防止超过模板预留的行数
        if len(steps) > self.MAX_SINGLE_STEPS:
            raise ValueError(
                f"当前单人标准作业组合票共有 {len(steps)} 行，已超过模板最多支持的 {self.MAX_SINGLE_STEPS} 行。\n"
                "请合并部分区间或拆分为多张组合票后再导出。"
            )

        if not steps:
            raise ValueError("请至少填写一行有效的作业（需有作业名称和时间）。")

        totals = {
            "manual": total_manual,
            "auto": total_auto,
            "walk": total_walk,
            "total": total_manual + total_auto + total_walk,
        }
        return project, part, worker, takt_sec, steps, totals

    # -------- 单人标准作业组合票：写入模板 --------
    def _export_single_to_excel(self, path, project, part, worker, takt_sec, steps, totals):
        """
        根据单人作业手顺（A→B 区间）将数据写入《组合票标准版.xlsx》模板：
        - 模板文件需放在与本文件同一目录下，文件名：组合票标准版.xlsx
        - 仅填充左侧步骤表区域和基本信息，不修改模板中的其他格式/图表
        """
        # 定位模板文件：与本 .py 同目录
        base_dir = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(base_dir, "组合票标准版.xlsx")
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"未找到模板文件：{template_path}")

        wb = load_workbook(template_path)
        try:
            ws = wb["④标准作业组合票"]
        except KeyError:
            ws = wb.active

        def _set_value(coord, value):
            """安全写入单元格：若目标是合并单元格，从其合并区域左上角写入"""
            cell = ws[coord]
            if isinstance(cell, MergedCell):
                for mr in ws.merged_cells.ranges:
                    if cell.coordinate in mr:
                        ws.cell(row=mr.min_row, column=mr.min_col).value = value
                        break
            else:
                cell.value = value

        def _set_fill(row, col, fill):
            """安全设置单元格填充：若目标是合并单元格，则写到其合并区域左上角"""
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                for mr in ws.merged_cells.ranges:
                    if cell.coordinate in mr:
                        ws.cell(row=mr.min_row, column=mr.min_col).fill = fill
                        break
            else:
                cell.fill = fill

        def _set_border(row, col, border: Border):
            """
            安全设置单元格边框：若目标是合并单元格，则写到其合并区域左上角；
            与已有边框合并（只改指定方向的线型）。
            """
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                for mr in ws.merged_cells.ranges:
                    if cell.coordinate in mr:
                        cell = ws.cell(row=mr.min_row, column=mr.min_col)
                        break

            old = cell.border or Border()

            def merge_side(new_side, old_side):
                if getattr(new_side, "style", None):
                    return new_side
                return old_side

            cell.border = Border(
                left=merge_side(border.left, old.left),
                right=merge_side(border.right, old.right),
                top=merge_side(border.top, old.top),
                bottom=merge_side(border.bottom, old.bottom),
                diagonal=old.diagonal,
                diagonal_direction=old.diagonal_direction,
                outline=old.outline,
                vertical=old.vertical,
                horizontal=old.horizontal,
            )

        def _clear_top_border(row, col):
            """
            清除单元格的上边框（保留其余边框），合并单元格时操作左上角单元格。
            """
            cell = ws.cell(row=row, column=col)
            # Handle merged cells: always operate on effective top-left cell
            if isinstance(cell, MergedCell):
                for mr in ws.merged_cells.ranges:
                    if cell.coordinate in mr:
                        cell = ws.cell(row=mr.min_row, column=mr.min_col)
                        break
            old = cell.border or Border()
            cell.border = Border(
                left=old.left,
                right=old.right,
                top=Side(style=None),
                bottom=old.bottom,
                diagonal=old.diagonal,
                diagonal_direction=old.diagonal_direction,
                outline=old.outline,
                vertical=old.vertical,
                horizontal=old.horizontal,
            )

        def _clear_left_border(row, col):
            """
            清除单元格的左边框（保留其余边框）；合并单元格时操作左上角单元格。
            """
            cell = ws.cell(row=row, column=col)
            # Handle merged cells: always operate on effective top-left cell
            if isinstance(cell, MergedCell):
                for mr in ws.merged_cells.ranges:
                    if cell.coordinate in mr:
                        cell = ws.cell(row=mr.min_row, column=mr.min_col)
                        break
            old = cell.border or Border()
            cell.border = Border(
                left=Side(style=None),
                right=old.right,
                top=old.top,
                bottom=old.bottom,
                diagonal=old.diagonal,
                diagonal_direction=old.diagonal_direction,
                outline=old.outline,
                vertical=old.vertical,
                horizontal=old.horizontal,
            )

        def _clear_right_border(row, col):
            """
            清除单元格的右边框（保留其余边框）；合并单元格时操作左上角单元格。
            """
            cell = ws.cell(row=row, column=col)
            # Handle merged cells: always operate on effective top-left cell
            if isinstance(cell, MergedCell):
                for mr in ws.merged_cells.ranges:
                    if cell.coordinate in mr:
                        cell = ws.cell(row=mr.min_row, column=mr.min_col)
                        break
            old = cell.border or Border()
            cell.border = Border(
                left=old.left,
                right=Side(style=None),
                top=old.top,
                bottom=old.bottom,
                diagonal=old.diagonal,
                diagonal_direction=old.diagonal_direction,
                outline=old.outline,
                vertical=old.vertical,
                horizontal=old.horizontal,
            )

        # 1) 清空左侧原有数据区域
        start_row = 9
        row_span = 3
        max_steps = getattr(self, "MAX_SINGLE_STEPS", 23)
        end_row = start_row + max_steps * row_span - 1
        for row in range(start_row, end_row + 1):
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None

        # 清空右侧时间轴区域填充（F列开始，按总时间估算范围）
        time_start_col = 6  # F列
        max_time = int(round(totals.get("total", 0))) if isinstance(totals, dict) else 0
        if max_time < 0:
            max_time = 0
        time_end_col = time_start_col + max_time + 5
        for row in range(start_row, end_row + 1):
            for col in range(time_start_col, time_end_col + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                cell.fill = PatternFill()

          # 2) 写入步骤：每步占 3 行（A9:A11, A12:A14, ...）
        row_span = 3
        time_start_col = 6  # F列，时间轴起始列
        time_fill = PatternFill(fill_type="solid", fgColor="000000")
        # 自动：加粗虚线（仅画上边框，不填充）
        auto_side = Side(style="mediumDashed", color="000000")
        h_auto_border = Border(top=auto_side)

        # segments：按“步骤”记录每一行的黑条起止（即手作业+自动的时间段）
        segments = []  # [{"mid_row": int, "bar_start": int, "bar_end": int}, ...]

        for idx, s in enumerate(steps):
            base_row = start_row + idx * row_span

            # 序号
            ws.cell(row=base_row, column=1).value = s["seq"]

            # 作业名称 A/B：B 列两行
            name_a = s.get("name_a") or s.get("name") or ""
            name_b = s.get("name_b") or ""
            ws.cell(row=base_row, column=2).value = name_a
            if name_b:
                ws.cell(row=base_row + 2, column=2).value = name_b

            # 时间数值（C~E）
            ws.cell(row=base_row, column=3).value = s["manual"]
            ws.cell(row=base_row, column=4).value = s["auto"]
            ws.cell(row=base_row, column=5).value = s["walk"]

            # ===== 时间轴绘制（手作业=黑填充；自动=加粗虚线；步行仅用折线表示） =====
            start_sec = int(round(s["start"]))
            manual = float(s["manual"])
            auto = float(s["auto"])
            walk = float(s["walk"])
            walk_pos = s.get("walk_pos", "after")
            auto_pos = s.get("auto_pos", "after")

            # 起点：若步行在前，整体右移
            if walk_pos == "before":
                bar_start_sec = int(round(start_sec + walk))
            else:
                bar_start_sec = start_sec

            mid_row = base_row + 1

            # 决定绘制顺序：自动在前/后
            draw_seq = []
            if auto_pos == "before":
                if auto > 0:
                    draw_seq.append(("auto", auto))
                if manual > 0:
                    draw_seq.append(("manual", manual))
            else:
                if manual > 0:
                    draw_seq.append(("manual", manual))
                if auto > 0:
                    draw_seq.append(("auto", auto))

            seg_start = bar_start_sec
            for kind, length in draw_seq:
                seg_end = int(round(seg_start + length))
                if seg_end > seg_start:
                    for sec in range(seg_start, seg_end):
                        col = time_start_col + sec
                        if kind == "manual":
                            _set_fill(mid_row, col, time_fill)        # 手作业：黑色填充
                        else:
                            _set_border(mid_row, col, h_auto_border)  # 自动：加粗虚线（上边框）
                seg_start = seg_end

            # 记录该步的整体开始/结束（不包含步行在后）
            bar_end_sec = seg_start
            segments.append(
                {
                    "mid_row": mid_row,
                    "bar_start": bar_start_sec,
                    "bar_end": bar_end_sec,
                }
            )

        # 2.5) 相邻「步骤」之间画连接线：
        #      - 有间隔：步行 → 实折线，从黑条末端右边一格开始，先竖后横
        #      - 无间隔：直接接续 → 加粗实直线
        if len(segments) >= 2:
            solid_side = Side(style="medium", color="000000")   # 加粗实线
            walk_side  = Side(style="medium", color="000000")   # 步行：加粗实线

            h_walk_border = Border(top=walk_side)               # 步行横线
            v_walk_left   = Border(left=walk_side)              # 竖线（当前列左边）
            v_walk_right  = Border(right=walk_side)             # 竖线镜像（前一列右边）
            v_solid_right_border = Border(right=solid_side)     # 无间隔直连竖线（边界线上）

            for i in range(len(segments) - 1):
                curr = segments[i]
                nxt = segments[i + 1]

                mid_row_curr = curr["mid_row"]
                mid_row_nxt = nxt["mid_row"]
                bar_end_curr = curr["bar_end"]
                bar_start_nxt = nxt["bar_start"]

                # 注意：bar_end / bar_start 是“时间（秒）”，还没加上 F 列偏移
                if bar_start_nxt > bar_end_curr:
                    # 有间隔：步行 → 实折线
                    # 连接策略：
                    #   - 竖线画在“上一列的右边界”，并且只画到下一段所在行的上一行（不进入下一段单元格）
                    #   - 横线从拐点所在列开始，沿下一段所在行的上边框一直画到下一段条形左侧
                    first_blank_col = time_start_col + bar_end_curr        # 上一段末尾右侧的第一格
                    next_bar_first_col = time_start_col + bar_start_nxt    # 下一段条形开始列

                    # 1) 竖线：用上一列（first_blank_col - 1）的『右边界』画，恰好停在下一段顶边
                    grid_col_for_right_edge = first_blank_col - 1
                    row_vert_start = mid_row_curr + 1
                    row_vert_end_exclusive = mid_row_nxt  # 不包含下一段所在行，避免出现“下垂尾巴”
                    if grid_col_for_right_edge >= time_start_col and row_vert_start < row_vert_end_exclusive:
                        for row in range(row_vert_start, row_vert_end_exclusive):
                            _set_border(row, grid_col_for_right_edge, v_walk_right)
                    # 保底清理下一行该列的右边界，避免‘下垂尾巴’
                    _clear_right_border(mid_row_nxt, grid_col_for_right_edge)

                    # 2) 横线：从拐点所在列开始（不跳空），一直到下一段左侧列
                    start_h_col = first_blank_col  # 不留缺口
                    if start_h_col < next_bar_first_col:
                        for col in range(start_h_col, next_bar_first_col):
                            _set_border(mid_row_nxt, col, h_walk_border)
                else:
                    # 无间隔：在上一段最后一秒所在列的“右边界”连线，
                    # 竖线落在列缝而不是下一段条形内部，且不覆盖下一段所在行
                    boundary_col = time_start_col + bar_end_curr
                    row_top = min(mid_row_curr, mid_row_nxt)
                    row_bottom = max(mid_row_curr, mid_row_nxt) - 1
                    if row_top <= row_bottom:
                        for row in range(row_top, row_bottom + 1):
                            _set_border(row, boundary_col - 1, v_solid_right_border)
                    # 保底清理下一行该列的右边界，避免‘下垂尾巴’
                    _clear_right_border(mid_row_nxt, boundary_col - 1)

        # 3) 合计行：B79 总时间，C79 手作业时间，D79 自动时间，E79 步行时间
        if isinstance(totals, dict):
            total_sec = totals.get("total", 0.0)
            manual_sec = totals.get("manual", 0.0)
            auto_sec = totals.get("auto", 0.0)
            walk_sec = totals.get("walk", 0.0)
        else:
            total_sec = manual_sec = auto_sec = walk_sec = 0.0

        def _fmt_sec(v):
            """把秒数统一转成整数秒写入单元格"""
            try:
                return int(round(float(v)))
            except Exception:
                return v

        _set_value("B79", _fmt_sec(total_sec))   # 合计下面：总时间
        _set_value("C79", _fmt_sec(manual_sec))  # 手作业合计
        _set_value("D79", _fmt_sec(auto_sec))    # 自动合计
        _set_value("E79", _fmt_sec(walk_sec))    # 步行合计

        # 4) 在上方空白处写入工程信息
        _set_value("B2", project)
        _set_value("B3", part)
        _set_value("B4", worker)
        _set_value("E2", takt_sec)

        # 5) 保存
        wb.save(path)

    def export_single_placeholder(self):
        """
        单工程组合票导出流程：
        1. 读取 Tab2 中 A→B 区间作业数据并校验
        2. 选择保存路径
        3. 使用固定 Excel 模板导出标准作业组合票
        """
        try:
            project, part, worker, takt_sec, steps, totals = self._collect_single_inputs()
        except Exception as e:
            QMessageBox.warning(self, "输入有误", str(e))
            return

        default_name = f"{project}_单人组合票.xlsx" if project else "单人组合票.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "导出标准作业组合票",
            default_name,
            "Excel (*.xlsx)",
        )
        if not path:
            return

        try:
            self._export_single_to_excel(path, project, part, worker, takt_sec, steps, totals)
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))
            return

        msg = (
            f"已导出标准作业组合票：\n{path}\n\n"
            f"工程名称：{project}\n"
            f"品番·品名：{part or '（未填写）'}\n"
            f"作业者：{worker or '（未填写）'}\n\n"
            f"节拍 TT：{takt_sec} 秒\n"
            f"总时间：{totals['total']:.1f} 秒\n"
            f"  其中 手作业：{totals['manual']:.1f} 秒\n"
            f"       自动：{totals['auto']:.1f} 秒\n"
            f"       步行：{totals['walk']:.1f} 秒\n\n"
            f"步骤数：{len(steps)} 步"
        )
        QMessageBox.information(self, "单工程组合票 - 导出完成", msg)

    # -------- 多车组合票：数据收集 & 导出 --------
    def fill_sample(self):
        """
        排程模型 v2 示例：设备数量 + 所属线别 + A/B/C 工时。
        工时 > 0 表示该车型在该岗位作业；工时 = 0 表示经过该岗位但不作业。
        """
        self.tbl.setRowCount(0)
        sample_rows = [
            # 序号, 工程名称,   设备数量, 所属线别, 岗位设备,       A工时, B工时, C工时
            ("1",  "电集",       "2",    "双线",     "电集",         "100", "100", "100"),
            ("2",  "空悬+快充",  "1",    "1号线",    "空悬快充设备", "0",   "200", "0"),
            ("3",  "四轮定位",   "2",    "双线",     "四轮定位",     "110", "110", "110"),
        ]
        for row in sample_rows:
            self.add_row()
            r = self.tbl.rowCount() - 1

            self.tbl.setItem(r, 0, QTableWidgetItem(str(row[0])))
            self.tbl.setItem(r, 1, QTableWidgetItem(str(row[1])))

            device_count_widget = self.tbl.cellWidget(r, 2)
            if isinstance(device_count_widget, QComboBox):
                device_count_widget.setCurrentText(str(row[2]))

            line_scope_widget = self.tbl.cellWidget(r, 3)
            if isinstance(line_scope_widget, QComboBox):
                line_scope_widget.setCurrentText(str(row[3]))

            self.tbl.setItem(r, 4, QTableWidgetItem(str(row[4])))
            self.tbl.setItem(r, 5, QTableWidgetItem(str(row[5])))
            self.tbl.setItem(r, 6, QTableWidgetItem(str(row[6])))
            self.tbl.setItem(r, 7, QTableWidgetItem(str(row[7])))

        if not self.ed_project.text().strip():
            self.ed_project.setText("排程模型v2示例")
        self.spn_a_cars.setValue(4)
        self.spn_b_cars.setValue(2)
        self.spn_c_cars.setValue(0)
        self.spn_total_cars.setValue(60)
        self.cmb_grid.setCurrentText("1.0")
        self.cmb_wait.setCurrentText("开始前等待")
        self.cmb_launch_mode.setCurrentText("按数量投车")
        self.cmb_seq.setCurrentText("顺排(A→B→C)")
        self.spn_max_run.setValue(10)

    def _collect_multi_raw_inputs(self):
        """Collect multi-project UI values as plain data for later parser migration."""
        project = self.ed_project.text().strip() or "工程"
        cars_a = int(self.spn_a_cars.value())
        cars_b = int(self.spn_b_cars.value())
        cars_c = int(self.spn_c_cars.value())
        analysis_minutes = int(self.spn_total_cars.value()) if hasattr(self, "spn_total_cars") else 0
        target_takt = float(self.spn_target_takt.value()) if hasattr(self, "spn_target_takt") else 0.0
        is_ratio_mode = self.cmb_launch_mode.currentIndex() == 1
        sequence_mode_index = self.cmb_seq.currentIndex() if hasattr(self, "cmb_seq") else 0
        max_consecutive = int(self.spn_max_run.value()) if hasattr(self, "spn_max_run") else 10

        station_rows = []
        for r in range(self.tbl.rowCount()):
            seq = (self.tbl.item(r, 0).text().strip() if self.tbl.item(r, 0) else "")
            name = (self.tbl.item(r, 1).text().strip() if self.tbl.item(r, 1) else "")

            device_count_widget = self.tbl.cellWidget(r, 2)
            device_count_text = device_count_widget.currentText().strip() if isinstance(device_count_widget, QComboBox) else "2"

            line_scope_widget = self.tbl.cellWidget(r, 3)
            line_scope = line_scope_widget.currentText().strip() if isinstance(line_scope_widget, QComboBox) else "双线"

            grp = (self.tbl.item(r, 4).text().strip() if self.tbl.item(r, 4) else "")
            dur_a = (self.tbl.item(r, 5).text().strip() if self.tbl.item(r, 5) else "")
            dur_b = (self.tbl.item(r, 6).text().strip() if self.tbl.item(r, 6) else "")
            dur_c = (self.tbl.item(r, 7).text().strip() if self.tbl.item(r, 7) else "")

            station_rows.append({
                "seq": seq,
                "display": name,
                "device_count": device_count_text,
                "line_scope": line_scope,
                "group": grp,
                "duration_a": dur_a,
                "duration_b": dur_b,
                "duration_c": dur_c,
                "color": "",
            })

        return {
            "project": project,
            "cars_a": cars_a,
            "cars_b": cars_b,
            "cars_c": cars_c,
            "analysis_minutes": analysis_minutes,
            "target_takt": target_takt,
            "is_ratio_mode": is_ratio_mode,
            "sequence_mode_index": sequence_mode_index,
            "max_consecutive": max_consecutive,
            "station_rows": station_rows,
        }

    def _parse_multi_inputs_from_raw(self):
        """
        通过 core.input_parser 解析多工程原始输入。
        当前阶段仅作为后续替换 _collect_inputs() 的桥接函数。
        暂不接入现有分析/导出流程。
        """
        raw_inputs = self._collect_multi_raw_inputs()
        return core_parse_multi_project_inputs(raw_inputs)

    def _collect_inputs_from_parser_tuple(self):
        """
        使用 core.input_parser 解析多工程输入，并转换为 _collect_inputs() 兼容的旧 tuple。
        当前阶段仅作为后续替换 _collect_inputs() 的桥接函数。
        暂不接入现有分析/导出流程。
        """
        parsed = self._parse_multi_inputs_from_raw()

        self.current_analysis_time_seconds = parsed.get("analysis_time_seconds")
        self.current_theoretical_launch_count = parsed.get("theoretical_launch_count")
        self.current_simulation_buffer_count = parsed.get("simulation_buffer_count", 0)
        self.current_simulation_vehicle_count = parsed.get("simulation_vehicle_count", parsed.get("cars", 0))

        return (
            parsed["project"],
            parsed["cars"],
            parsed["grid_step"],
            parsed["wait_policy"],
            parsed["defs"],
            parsed["vehicle_counts"],
            parsed["sequence_mode"],
            parsed["max_consecutive"],
            parsed["ratio_pattern"],
            parsed["target_takt"],
        )

    def _collect_inputs(self):
        return self._collect_inputs_from_parser_tuple()

    def do_export(self):
        try:
            project, cars, grid_step, wait_policy, defs, vehicle_counts, sequence_mode, max_consecutive, ratio_pattern, target_takt = self._collect_inputs()
            frozen_sequence = self._frozen_sequence_for_current_inputs(sequence_mode)
        except Exception as e:
            QMessageBox.warning(self, "输入有误", str(e))
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            "导出位置",
            f"{project}_组合票.xlsx",
            "Excel (*.xlsx)",
        )
        if not path:
            return
        self.dst_path = path

        worker = Worker(
            tickets.schedule_and_export,
            defs, cars, grid_step, wait_policy, project, self.dst_path,
            vehicle_counts, sequence_mode, max_consecutive, ratio_pattern, target_takt, frozen_sequence,
        )
        worker.signals.error.connect(self._on_error)
        worker.signals.finished.connect(self._on_export_finished)
        self.thread_pool.start(worker)
        self.status.showMessage("正在生成组合票...", 5000)

    def _on_export_finished(self, *args):
        self.status.showMessage("导出完成", 6000)
        QMessageBox.information(self, "完成", f"已导出：\n{self.dst_path}")

    def _fmt_analysis_num(self, value):
        try:
            v = float(value)
        except Exception:
            return str(value)
        if abs(v - round(v)) < 1e-9:
            return str(int(round(v)))
        return f"{v:.1f}"
    
    def _apply_time_window_analysis(self, analysis, rows, target_takt):
        """v2-6D：按比例投车的时间窗口产能分析薄封装。"""
        return core_apply_time_window_analysis(
            analysis=analysis,
            rows=rows,
            target_takt=target_takt,
            analysis_time_seconds=getattr(self, "current_analysis_time_seconds", None),
            theoretical_launch_count=getattr(self, "current_theoretical_launch_count", None),
            simulation_buffer_count=getattr(self, "current_simulation_buffer_count", 0),
        )

    def _show_analysis_result(self, analysis):
        summary = analysis.get("summary", {}) if isinstance(analysis, dict) else {}

        total_cars = summary.get("total_cars", 0)
        max_finish = self._fmt_analysis_num(summary.get("max_finish", 0.0))
        total_wait = self._fmt_analysis_num(summary.get("total_wait", 0.0))
        avg_wait = self._fmt_analysis_num(summary.get("avg_wait", 0.0))
        blocking_result = summary.get("blocking_result", "无阻塞") or "无阻塞"
        blocking_time = self._fmt_analysis_num(summary.get("total_wait", 0.0))
        overflow_vehicle_count = self._fmt_analysis_num(summary.get("overflow_vehicle_count", 0.0))
        blocking_station_text = summary.get("blocking_station_text", "无") or "无"
        batch_overrun_time = self._fmt_analysis_num(summary.get("batch_overrun_time", 0.0))
        batch_overrun_cars = self._fmt_analysis_num(summary.get("batch_overrun_cars", 0.0))
        process_root_text = summary.get("process_over_takt_root_text", "无") or "无"
        takt_result = summary.get("takt_result", "未设定") or "未设定"
        over_count = summary.get("over_takt_station_count", 0)
        process_root_text = summary.get("process_over_takt_root_text", "无") or "无"
        batch_overrun_raw = float(summary.get("batch_overrun_time", 0.0) or 0.0)
        text = (
            "结果分析："
            f"累计阻塞 {blocking_time} 秒 ｜ "
            f"溢出工时 {batch_overrun_time} 秒 / {batch_overrun_cars} 台 ｜ "
            f"阻塞工程 {blocking_station_text}"
        )


        if hasattr(self, "lbl_analysis"):
            self.lbl_analysis.setText(text)
        else:
            QMessageBox.information(self, "排程分析完成", text)

        if hasattr(self, "_show_station_analysis"):
            self._show_station_analysis(analysis)

        if hasattr(self, "_show_vehicle_summary"):
            self._show_vehicle_summary(analysis)


    def _on_error(self, err_msg):
        self.status.showMessage("导出失败", 6000)
        QMessageBox.critical(self, "导出失败", str(err_msg))


    # ---------- 帮助弹窗 ----------
    def show_help(self):
        msg = (
            "<h3>组合票操作指南</h3>"
            "<ol>"
            "<li>多工程组合票按『设备数量 + 所属线别 + 岗位设备 + A/B/C 工时』录入</li>"
            "<li>设备数量可选：1 / 2；设备数量为 2 时，所属线别固定为『双线』</li>"
            "<li>设备数量为 1 时，所属线别可选：1号线 / 2号线 / 双线共用</li>"
            "<li>A/B/C 工时大于 0 表示该车型在该岗位作业；工时为 0 表示经过该岗位但不作业</li>"
            "<li>参与投车的车型，工时不能为空；未参与投车的车型，工时可以为空</li>"
            "<li>投车模式支持：按数量投车 / 按比例投车；按数量投车下可选择顺排(A→B→C)或交替混流</li>"
            "<li>顺排/交替混流模式下，A/B/C 填数量；按比例投车模式下，A/B/C 填比例，并用分析时间与目标节拍计算理论投车台数</li>"
            "<li>最大连续台数默认 10；填 1 表示尽量强制交替</li>"
            "<li>填写完点击『分析当前排程』可查看结果；点击『生成并导出组合票』即可生成 Excel</li>"
            "</ol>"
        )
        QMessageBox.information(self, "帮助", msg)


    def _show_station_analysis(self, analysis):
        if not hasattr(self, "tbl_station_analysis"):
            return

        station_stats = analysis.get("station_stats", []) if isinstance(analysis, dict) else []
        self.tbl_station_analysis.setRowCount(0)

        for item in station_stats:
            r = self.tbl_station_analysis.rowCount()
            self.tbl_station_analysis.insertRow(r)

            values = [
                str(item.get("station", "")),
                str(item.get("count", 0)),
                self._fmt_analysis_num(item.get("total_process", 0.0)),
                self._fmt_analysis_num(item.get("blocking_time", item.get("overflow_wait_time", item.get("total_block_wait", 0.0)))),
                self._fmt_analysis_num(item.get("avg_process", 0.0)),
                self._fmt_analysis_num(item.get("avg_overflow_wait", item.get("avg_block_wait", 0.0))),
                str(item.get("takt_result", "未设定")),
                str(item.get("over_takt_types", "—")),
            ]

            for c, value in enumerate(values):
                table_item = QTableWidgetItem(value)
                if c > 0:
                    table_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                else:
                    table_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                self.tbl_station_analysis.setItem(r, c, table_item)

        # 固定列宽，不再自动根据内容调整，避免每次分析后表格宽度跳动
        # self.tbl_station_analysis.resizeColumnsToContents()

    def _build_wait_display_metrics(self, cutoff_time: float, target_takt: float) -> dict:
        """按当前统计截止时间汇总实际等待、节拍外等待及其真因。"""
        rows = getattr(self, "last_schedule_rows", []) or []
        try:
            cutoff = max(0.0, float(cutoff_time or 0.0))
        except Exception:
            cutoff = 0.0
        try:
            target = max(0.0, float(target_takt or 0.0))
        except Exception:
            target = 0.0

        actual_by_station = {}
        excess_by_station = {}
        cause_groups = {}
        cause_details = []
        total_actual = 0.0
        total_excess = 0.0
        incomplete_time = 0.0

        for row in rows:
            wait_start = self._sim_row_service_finish(row)
            wait_end = self._sim_row_depart(row)
            if cutoff <= wait_start + 1e-9 or wait_end <= wait_start + 1e-9:
                continue
            occurred_end = min(cutoff, wait_end)
            actual_wait = max(0.0, occurred_end - wait_start)
            if actual_wait <= 1e-9:
                continue

            station = self._sim_row_station(row) or "未知工程"
            total_actual += actual_wait
            actual_by_station[station] = actual_by_station.get(station, 0.0) + actual_wait

            try:
                capacity = max(1, int(float(row.get("capacity", row.get("device_count", 1)) or 1)))
            except Exception:
                capacity = 1
            holding_limit = capacity * target
            excess_start = wait_start + holding_limit
            excess_wait = max(0.0, occurred_end - excess_start)
            if excess_wait <= 1e-9:
                continue
            total_excess += excess_wait
            excess_by_station[station] = excess_by_station.get(station, 0.0) + excess_wait

            event_keys = set()
            covered = 0.0
            for cause_slice in row.get("wait_cause_slices", []) or []:
                slice_start = max(excess_start, float(cause_slice.get("start", excess_start) or excess_start))
                slice_end = min(occurred_end, float(cause_slice.get("end", occurred_end) or occurred_end))
                duration = max(0.0, slice_end - slice_start)
                if duration <= 1e-9:
                    continue
                covered += duration
                chain = list(cause_slice.get("chain", []) or [])
                direct = chain[0] if chain else {}
                direct_station = str(direct.get("blocked_station", "") or "未知工程")
                terminal_type = str(cause_slice.get("terminal_type", "") or "unresolved")
                terminal_station = str(cause_slice.get("terminal_station", "") or "未知工程")
                terminal_car_type = str(cause_slice.get("terminal_car_type", "") or "")
                complete = bool(cause_slice.get("chain_complete", False))
                if not complete:
                    terminal_cause = "原因链不完整"
                elif terminal_type == "over_takt_processing":
                    suffix = f" {terminal_car_type}" if terminal_car_type else ""
                    terminal_cause = f"{terminal_station}{suffix}超节拍加工占用"
                else:
                    terminal_cause = f"{terminal_station}前车占用"

                key = (station, direct_station, terminal_cause)
                item = cause_groups.setdefault(key, {
                    "waiting_station": station,
                    "direct_blocking_station": direct_station,
                    "terminal_cause": terminal_cause,
                    "event_count": 0,
                    "wait_time": 0.0,
                    "chain_complete": complete,
                })
                item["wait_time"] += duration
                event_keys.add(key)
                cause_details.append({
                    "car": row.get("car"),
                    "car_type": str(row.get("car_type", "") or ""),
                    "waiting_station": station,
                    "wait_start": slice_start,
                    "wait_end": slice_end,
                    "wait_time": duration,
                    "direct_blocker_car": direct.get("blocker_car"),
                    "direct_blocking_station": direct_station,
                    "direct_blocking_resource": str(direct.get("blocked_resource", "") or ""),
                    "terminal_car": cause_slice.get("terminal_car"),
                    "terminal_station": terminal_station,
                    "terminal_resource": str(cause_slice.get("terminal_resource", "") or ""),
                    "terminal_cause": terminal_cause,
                    "chain_complete": complete,
                    "chain": chain,
                })
            for key in event_keys:
                cause_groups[key]["event_count"] += 1
            if covered + 1e-9 < excess_wait:
                missing = excess_wait - covered
                incomplete_time += missing
                key = (station, "未知工程", "原因链不完整")
                item = cause_groups.setdefault(key, {
                    "waiting_station": station,
                    "direct_blocking_station": "未知工程",
                    "terminal_cause": "原因链不完整",
                    "event_count": 0,
                    "wait_time": 0.0,
                    "chain_complete": False,
                })
                item["event_count"] += 1
                item["wait_time"] += missing

        def _station_items(values):
            return [
                {"station": station, "wait_time": wait_time}
                for station, wait_time in sorted(values.items(), key=lambda item: (-item[1], item[0]))
            ]

        cause_summary = sorted(
            cause_groups.values(),
            key=lambda item: (-float(item["wait_time"]), -int(item["event_count"]), item["waiting_station"]),
        )
        coverage = (total_excess - incomplete_time) / total_excess if total_excess > 0 else 1.0
        return {
            "total_actual_wait": total_actual,
            "actual_wait_by_station": _station_items(actual_by_station),
            "total_excess_wait": total_excess,
            "excess_wait_by_station": _station_items(excess_by_station),
            "cause_chain_summary": cause_summary,
            "cause_chain_details": cause_details,
            "cause_chain_coverage_rate": coverage,
            "cause_chain_incomplete_time": incomplete_time,
        }

    def _build_realtime_model_result(self, current_time: float) -> dict:
        rows = getattr(self, "last_schedule_rows", []) or []
        try:
            current = float(current_time or 0.0)
        except Exception:
            current = 0.0

        try:
            target_takt = float(self.spn_target_takt.value()) if hasattr(self, "spn_target_takt") else 0.0
        except Exception:
            target_takt = 0.0

        car_rows = self._sim_car_rows()

        def _car_sort_key(car_key):
            try:
                return (0, int(float(car_key)))
            except Exception:
                return (1, str(car_key))

        analysis = getattr(self, "last_analysis", None)
        summary = analysis.get("summary", {}) if isinstance(analysis, dict) else {}
        is_ratio_mode = False
        if hasattr(self, "cmb_launch_mode"):
            is_ratio_mode = self.cmb_launch_mode.currentIndex() == 1
        else:
            is_ratio_mode = bool(summary.get("analysis_time_seconds"))

        analysis_time_seconds = 0.0
        if is_ratio_mode:
            try:
                analysis_time_seconds = float(summary.get("analysis_time_seconds", 0.0) or 0.0)
            except Exception:
                analysis_time_seconds = 0.0
            if analysis_time_seconds <= 0:
                try:
                    analysis_time_seconds = float(getattr(self, "current_analysis_time_seconds", 0.0) or 0.0)
                except Exception:
                    analysis_time_seconds = 0.0
            if analysis_time_seconds <= 0 and hasattr(self, "spn_total_cars"):
                try:
                    analysis_time_seconds = float(self.spn_total_cars.value()) * 60.0
                except Exception:
                    analysis_time_seconds = 0.0

        current_finished = []
        all_vehicles = []
        for car_key, segments in car_rows.items():
            if not segments:
                continue
            ordered = sorted(segments, key=lambda row: (self._sim_row_start(row), self._sim_row_depart(row)))
            car_in = self._sim_row_start(ordered[0])
            car_out = self._sim_row_depart(ordered[-1])
            car_type = str(
                self._sim_row_value(ordered[0], "car_type", "type", "vehicle_type", default="")
                or ""
            ).strip().upper()
            all_vehicles.append({
                "car_key": car_key,
                "car_type": car_type,
                "car_in": car_in,
                "car_out": car_out,
                "flow": max(0.0, car_out - car_in),
                "station_count": max(1, len(ordered)),
                "segments": ordered,
            })
            if car_out <= current + 1e-9:
                current_finished.append({
                    "car_key": car_key,
                    "car_type": car_type,
                    "car_in": car_in,
                    "car_out": car_out,
                    "flow": max(0.0, car_out - car_in),
                    "station_count": max(1, len(ordered)),
                    "segments": ordered,
                })

        all_vehicles.sort(key=lambda item: (item["car_out"], _car_sort_key(item["car_key"])))
        current_finished.sort(key=lambda item: (item["car_out"], _car_sort_key(item["car_key"])))
        if is_ratio_mode and analysis_time_seconds > 0:
            target_scope_vehicles = [
                item for item in all_vehicles
                if item["car_out"] <= analysis_time_seconds + 1e-9
            ]
        else:
            target_scope_vehicles = list(all_vehicles)
        denominator_vehicle_count = len(target_scope_vehicles)
        output_vehicle_count = denominator_vehicle_count

        def _recent_out_interval_takt(vehicles):
            if len(vehicles) < 2:
                return None
            recent = vehicles[-6:]
            gaps = [
                recent[idx]["car_out"] - recent[idx - 1]["car_out"]
                for idx in range(1, len(recent))
            ]
            gaps = [gap for gap in gaps if gap >= 0]
            if not gaps:
                return None
            return sum(gaps) / len(gaps)

        realtime_takt = _recent_out_interval_takt(current_finished)
        if denominator_vehicle_count >= 2:
            overall_takt = (
                target_scope_vehicles[-1]["car_out"] - target_scope_vehicles[0]["car_out"]
            ) / (denominator_vehicle_count - 1)
        else:
            overall_takt = None

        qualified_vehicle_count = 0
        capacity_over_station_map = {}
        capacity_results = (
            analysis.get("car_capacity_results", [])
            if isinstance(analysis, dict)
            else []
        )
        if not capacity_results:
            capacity_results = compute_car_capacity_results(rows, target_takt)
        capacity_by_car = {
            int(item.get("car", 0) or 0): item
            for item in capacity_results
            if int(item.get("car", 0) or 0) > 0
        }
        if denominator_vehicle_count > 0 and target_takt > 0:
            for item in target_scope_vehicles:
                try:
                    car_number = int(float(item.get("car_key", 0) or 0))
                except Exception:
                    car_number = 0
                capacity_result = capacity_by_car.get(car_number, {})
                car_type = str(capacity_result.get("car_type", item.get("car_type", "")) or "").upper()
                if bool(capacity_result.get("meets_capacity_target", False)):
                    qualified_vehicle_count += 1
                for over_item in capacity_result.get("over_capacity_stations", []) or []:
                    station_name = str(over_item.get("station", "") or "未知工程")
                    rec = capacity_over_station_map.setdefault(station_name, {
                        "station": station_name,
                        "vehicle_count": 0,
                        "by_type": {"A": 0, "B": 0, "C": 0},
                        "max_over": 0.0,
                    })
                    rec["vehicle_count"] += 1
                    if car_type in rec["by_type"]:
                        rec["by_type"][car_type] += 1
                    rec["max_over"] = max(
                        rec["max_over"], float(over_item.get("over_time", 0.0) or 0.0)
                    )
        capacity_over_stations = sorted(
            capacity_over_station_map.values(),
            key=lambda item: (-int(item.get("vehicle_count", 0) or 0), str(item.get("station", ""))),
        )
        over_takt_vehicle_count = denominator_vehicle_count - qualified_vehicle_count if target_takt > 0 else 0

        if target_takt <= 0 or denominator_vehicle_count <= 0 or overall_takt is None:
            result = "未判定"
            qualified_rate = None
        else:
            qualified_rate = qualified_vehicle_count / denominator_vehicle_count
            if (
                qualified_vehicle_count == denominator_vehicle_count
                and qualified_rate == 1.0
                and overall_takt <= target_takt
            ):
                result = "OK"
            else:
                result = "NG"

        blocking_time = current
        if is_ratio_mode and analysis_time_seconds > 0:
            blocking_time = analysis_time_seconds
        elif all_vehicles:
            blocking_time = all_vehicles[-1]["car_out"]

        total_blocking_so_far = 0.0
        for row in rows:
            block_wait = self._sim_row_block_wait(row)
            if block_wait <= 0:
                continue
            svc_finish = self._sim_row_service_finish(row)
            depart = self._sim_row_depart(row)
            if blocking_time <= svc_finish:
                continue
            if blocking_time < depart:
                total_blocking_so_far += min(max(0.0, blocking_time - svc_finish), block_wait)
            else:
                total_blocking_so_far += block_wait
        blocking_hint = self._build_realtime_blocking_hint(blocking_time)
        wait_metrics = self._build_wait_display_metrics(blocking_time, target_takt)

        return {
            "result": result,
            "output_vehicle_count": output_vehicle_count,
            "qualified_vehicle_count": qualified_vehicle_count,
            "denominator_vehicle_count": denominator_vehicle_count,
            "qualified_rate": qualified_rate,
            "over_takt_vehicle_count": over_takt_vehicle_count,
            "capacity_over_stations": capacity_over_stations,
            "realtime_takt": realtime_takt,
            "overall_takt": overall_takt,
            "target_takt": target_takt,
            "total_blocking_so_far": total_blocking_so_far,
            "total_actual_wait": wait_metrics.get("total_actual_wait", 0.0),
            "actual_wait_by_station": wait_metrics.get("actual_wait_by_station", []),
            "total_excess_wait": wait_metrics.get("total_excess_wait", 0.0),
            "excess_wait_by_station": wait_metrics.get("excess_wait_by_station", []),
            "cause_chain_summary": wait_metrics.get("cause_chain_summary", []),
            "cause_chain_details": wait_metrics.get("cause_chain_details", []),
            "cause_chain_coverage_rate": wait_metrics.get("cause_chain_coverage_rate", 1.0),
            "cause_chain_incomplete_time": wait_metrics.get("cause_chain_incomplete_time", 0.0),
            "blocking_hint": blocking_hint,
            "current_time": current,
            "all_vehicles": all_vehicles,
            "target_scope_vehicles": target_scope_vehicles,
        }

    def _update_realtime_model_result(self):
        if not getattr(self, "last_schedule_rows", None):
            if hasattr(self, "lbl_realtime_takt"):
                self.lbl_realtime_takt.setText("近期节拍（近5个间隔）：-")
            self._last_model_result_summary = None
            return
        realtime = self._build_realtime_model_result(float(getattr(self, "sim_time", 0.0) or 0.0))
        if hasattr(self, "lbl_realtime_takt"):
            realtime_takt = realtime.get("realtime_takt")
            target_takt = self._fmt_analysis_num(realtime.get("target_takt", 0.0))
            if realtime_takt is None:
                self.lbl_realtime_takt.setText("近期节拍（近5个间隔）：-")
            else:
                self.lbl_realtime_takt.setText(
                    f"近期节拍（近5个间隔）：{self._fmt_analysis_num(realtime_takt)}/{target_takt}"
                )
        analysis = getattr(self, "last_analysis", None)
        if isinstance(analysis, dict):
            self._show_vehicle_summary(analysis)

    def _show_vehicle_summary(self, analysis):
        """在车型数据摘要区展示当前排程的车型构成与关键结果。"""
        if not hasattr(self, "lbl_vehicle_summary"):
            return

        summary = analysis.get("summary", {}) if isinstance(analysis, dict) else {}
        car_stats = analysis.get("car_stats", []) if isinstance(analysis, dict) else []
        car_type_summary = analysis.get("car_type_summary", analysis.get("type_stats", [])) if isinstance(analysis, dict) else []

        type_counts = {"A": 0, "B": 0, "C": 0}
        if car_stats:
            for item in car_stats:
                vt = str(item.get("car_type", "") or "").strip().upper()
                if vt in type_counts:
                    type_counts[vt] += 1
        else:
            for item in car_type_summary:
                vt = str(item.get("car_type", "") or "").strip().upper()
                if vt in type_counts:
                    type_counts[vt] += int(item.get("count", 0) or 0)

        total_cars = int(summary.get("total_cars", 0) or 0)
        if total_cars <= 0:
            total_cars = sum(type_counts.values())

        # v2-5 blocking analysis fields
        target_takt = self._fmt_analysis_num(summary.get("target_takt", 0.0))
        max_finish = self._fmt_analysis_num(summary.get("max_finish", 0.0))
        total_wait = self._fmt_analysis_num(summary.get("total_wait", 0.0))
        avg_wait = self._fmt_analysis_num(summary.get("avg_wait", 0.0))
        takt_result = summary.get("takt_result", "未设定") or "未设定"
        over_count = summary.get("over_takt_station_count", 0)
        blocking_result = summary.get("blocking_result", "无阻塞") or "无阻塞"
        blocking_station_text = summary.get("blocking_station_text", "无") or "无"
        blocking_station_count = summary.get("blocking_station_count", 0)
        blocking_time = self._fmt_analysis_num(summary.get("total_wait", 0.0))
        overflow_vehicle_count = self._fmt_analysis_num(summary.get("overflow_vehicle_count", 0.0))
        batch_overrun_time = self._fmt_analysis_num(summary.get("batch_overrun_time", 0.0))
        batch_overrun_cars = self._fmt_analysis_num(summary.get("batch_overrun_cars", 0.0))
        process_root_text = summary.get("process_over_takt_root_text", "无") or "无"

        batch_overrun_raw = float(summary.get("batch_overrun_time", 0.0) or 0.0)

        time_window_result = summary.get("time_window_result", "") or ""
        final_result = time_window_result or "OK"

        actual_output_count_raw = int(summary.get("actual_output_count_in_window", 0) or 0)
        display_actual_output_count = self._fmt_analysis_num(summary.get(
            "display_actual_output_count_in_window",
            actual_output_count_raw,
        ))
        planned_output_count = self._fmt_analysis_num(summary.get("planned_output_count_in_window", 0.0))
        theoretical_launch_count = int(summary.get("theoretical_launch_count", total_cars) or total_cars)
        achievement_rate = float(summary.get("achievement_rate", 0.0) or 0.0) * 100
        target_takt_display = self._fmt_analysis_num(summary.get("target_takt", self.spn_target_takt.value() if hasattr(self, "spn_target_takt") else 0))

        def _fmt_optional_seconds(value):
            if value is None:
                return "—"
            return self._fmt_analysis_num(value)

        planned_n_finish_time = _fmt_optional_seconds(summary.get("planned_n_finish_time"))
        actual_n_finish_time = _fmt_optional_seconds(summary.get("actual_n_finish_time"))
        finish_delta_raw = summary.get("finish_delta")
        finish_delta_num = None
        if finish_delta_raw is None:
            finish_delta = "—"
        else:
            try:
                finish_delta_num = float(finish_delta_raw)
                finish_delta = f"{finish_delta_num:+.1f}" if abs(finish_delta_num - round(finish_delta_num)) >= 1e-9 else f"{int(round(finish_delta_num)):+d}"
            except Exception:
                finish_delta = str(finish_delta_raw)
        actual_line_takt_in_window = _fmt_optional_seconds(summary.get("actual_line_takt_in_window"))

        def _metric_card(title, value, value_color="#0f172a"):
            return (
                "<td width='20%' style='"
                "width:20%;"
                "border:1px solid #dbe3ef;"
                "background:#ffffff;"
                "padding:8px 9px;"
                "vertical-align:top;"
                "'>"
                f"<div style='font-size:10px;color:#64748b;'>{title}</div>"
                f"<div style='font-size:14px;font-weight:700;color:{value_color};margin-top:3px;'>{value}</div>"
                "</td>"
            )

        is_ratio = hasattr(self, "cmb_launch_mode") and self.cmb_launch_mode.currentIndex() == 1
        if is_ratio:
            ratio_a = int(self.spn_a_cars.value()) if hasattr(self, "spn_a_cars") else 0
            ratio_b = int(self.spn_b_cars.value()) if hasattr(self, "spn_b_cars") else 0
            ratio_c = int(self.spn_c_cars.value()) if hasattr(self, "spn_c_cars") else 0
            analysis_minutes = int(self.spn_total_cars.value()) if hasattr(self, "spn_total_cars") else 0
            theoretical_launch_count = int(summary.get("theoretical_launch_count", 0) or 0)
            simulation_buffer_count = int(summary.get("simulation_buffer_count", 0) or 0)
            summary_lines = [
                (
                    ("模式", "按比例投车"),
                    ("比例", f"{ratio_a}:{ratio_b}:{ratio_c}"),
                ),
                (
                    ("时间", f"{analysis_minutes}分钟"),
                    ("目标节拍", f"{target_takt_display}s"),
                ),
                (
                    ("理论投车", f"{theoretical_launch_count}台"),
                    ("仿真缓冲", f"{simulation_buffer_count}台"),
                ),
            ]
        else:
            seq_text = self.cmb_seq.currentText() if hasattr(self, "cmb_seq") else "—"
            summary_lines = [
                (
                    ("模式", "按数量投车"),
                    ("A/B/C", f"A{type_counts['A']}/B{type_counts['B']}/C{type_counts['C']}"),
                ),
                (
                    ("排列", seq_text),
                    ("目标节拍", f"{target_takt_display}s"),
                ),
                (
                    ("投车台数", f"{total_cars}台"),
                ),
            ]

        def _summary_line(items):
            return "｜".join(
                f"<span style='color:#64748b;'>{label}：</span><b style='color:#0f172a;'>{value}</b>"
                for label, value in items
            )

        left_html = (
            "<div style='font-weight:700;color:#334155;margin-bottom:6px;'>基础排程摘要</div>"
            "<div style='font-size:13px;color:#1e293b;line-height:1.75;'>"
            + "<br>".join(_summary_line(line) for line in summary_lines)
            + "</div>"
        )
        realtime = self._build_realtime_model_result(float(getattr(self, "sim_time", 0.0) or 0.0))
        output_vehicle_count = int(realtime.get("output_vehicle_count", 0) or 0)
        qualified_vehicle_count = int(realtime.get("qualified_vehicle_count", 0) or 0)
        denominator_vehicle_count = int(realtime.get("denominator_vehicle_count", output_vehicle_count) or 0)
        qualified_rate = realtime.get("qualified_rate")
        qualified_rate_text = "—" if qualified_rate is None else f"{float(qualified_rate) * 100:.1f}%"
        target_takt_value = realtime.get("target_takt", summary.get("target_takt", 0.0))
        target_takt_text = self._fmt_analysis_num(target_takt_value)
        overall_takt = realtime.get("overall_takt")
        overall_takt_text = "—" if overall_takt is None else self._fmt_analysis_num(overall_takt)
        actual_wait = float(realtime.get("total_actual_wait", 0.0) or 0.0)
        excess_wait = float(realtime.get("total_excess_wait", 0.0) or 0.0)
        actual_wait_text = self._fmt_analysis_num(actual_wait)
        excess_wait_text = self._fmt_analysis_num(excess_wait)
        risk_parts = []
        capacity_over_stations = realtime.get("capacity_over_stations", []) or []
        if capacity_over_stations:
            process_parts = []
            for item in capacity_over_stations:
                station = str(item.get("station", "") or "岗位")
                by_type = item.get("by_type", {}) or {}
                types = [vehicle_type for vehicle_type in ("A", "B", "C") if int(by_type.get(vehicle_type, 0) or 0) > 0]
                type_text = "/".join(types)
                over_text = self._fmt_analysis_num(item.get("max_over", 0.0))
                process_parts.append(f"{station} {type_text} {over_text}s/台".strip())
            risk_parts.append(f"超节拍工程：{'；'.join(process_parts)}")
        excess_station_items = realtime.get("excess_wait_by_station", []) or []
        if excess_station_items:
            station_text = "；".join(
                f"{item.get('station', '未知工程')} {self._fmt_analysis_num(item.get('wait_time', 0.0))}s"
                for item in excess_station_items[:3]
            )
            risk_parts.insert(0, f"节拍外等待发生工程：{station_text}")
        risk_hint_text = "｜".join(risk_parts) if risk_parts else "暂无明显风险"

        cause_chain_items = realtime.get("cause_chain_summary", []) or []
        cause_chain_parts = []
        for item in cause_chain_items[:3]:
            cause_chain_parts.append(
                f"{item.get('terminal_cause', '原因链不完整')} → "
                f"{item.get('direct_blocking_station', '未知工程')}无法放行 → "
                f"{item.get('waiting_station', '未知工程')}发生等待｜"
                f"{int(item.get('event_count', 0) or 0)}次｜"
                f"{self._fmt_analysis_num(item.get('wait_time', 0.0))}s"
            )
        cause_chain_text = "；".join(cause_chain_parts) if cause_chain_parts else "暂无节拍外等待真因"

        target_scope_vehicles = list(realtime.get("target_scope_vehicles", []) or [])
        all_vehicles = list(realtime.get("all_vehicles", []) or [])
        last_output_vehicle = target_scope_vehicles[-1] if target_scope_vehicles else None
        next_output_vehicle = None
        if last_output_vehicle is not None:
            try:
                last_out_value = float(last_output_vehicle.get("car_out", 0.0) or 0.0)
            except Exception:
                last_out_value = 0.0
            for item in all_vehicles:
                try:
                    candidate_out = float(item.get("car_out", 0.0) or 0.0)
                except Exception:
                    candidate_out = 0.0
                if candidate_out > last_out_value + 1e-9:
                    next_output_vehicle = item
                    break

        analysis_time_seconds = summary.get("analysis_time_seconds")
        try:
            analysis_time_seconds = float(analysis_time_seconds or 0.0)
        except Exception:
            analysis_time_seconds = 0.0
        analysis_time_minutes = 0.0
        if analysis_time_seconds > 0:
            analysis_time_minutes = analysis_time_seconds / 60.0
        elif hasattr(self, "spn_total_cars"):
            try:
                analysis_time_minutes = float(self.spn_total_cars.value() or 0.0)
                analysis_time_seconds = analysis_time_minutes * 60.0
            except Exception:
                analysis_time_minutes = 0.0
                analysis_time_seconds = 0.0

        self._last_model_result_summary = {
            "analysis_time_minutes": analysis_time_minutes,
            "analysis_time_seconds": analysis_time_seconds,
            "output_count": output_vehicle_count,
            "qualified_count": qualified_vehicle_count,
            "qualified_rate": qualified_rate,
            "qualified_rate_percent": None if qualified_rate is None else qualified_rate * 100.0,
            "target_takt": target_takt_value,
            "first_out": target_scope_vehicles[0].get("car_out") if target_scope_vehicles else None,
            "last_out": target_scope_vehicles[-1].get("car_out") if target_scope_vehicles else None,
            "last_output_car_no": last_output_vehicle.get("car_key") if last_output_vehicle else None,
            "last_output_car_out": last_output_vehicle.get("car_out") if last_output_vehicle else None,
            "next_car_no": next_output_vehicle.get("car_key") if next_output_vehicle else None,
            "next_car_out": next_output_vehicle.get("car_out") if next_output_vehicle else None,
            "overall_takt": overall_takt,
            "total_block_wait": actual_wait,
            "total_actual_wait": actual_wait,
            "total_excess_wait": excess_wait,
            "excess_wait_by_station": excess_station_items,
            "cause_chain_summary": cause_chain_items,
            "cause_chain_coverage_rate": realtime.get("cause_chain_coverage_rate", 1.0),
            "risk_text": risk_hint_text or "暂无明显风险",
            "blocking_station_text": risk_hint_text,
        }

        model_cards = [
            _metric_card("下线车辆", f"{output_vehicle_count}台"),
            _metric_card("达标车辆", f"{qualified_vehicle_count}/{denominator_vehicle_count}"),
            _metric_card("达标率", qualified_rate_text),
            _metric_card("整体节拍", f"{overall_takt_text}/{target_takt_text}"),
            _metric_card("累计节拍外等待", f"{excess_wait_text}s"),
        ]
        if is_ratio:
            result_scope_title = f"模型结果（分析窗口终值：{analysis_minutes:g}分钟）"
            result_scope_note = f"统计范围：前{analysis_minutes:g}分钟终值"
        else:
            result_scope_title = f"模型结果（目标批次终值：{denominator_vehicle_count}台）"
            result_scope_note = f"统计范围：目标批次{denominator_vehicle_count}台终值"

        right_html = (
            f"<div style='font-size:13px;font-weight:700;color:#334155;margin-bottom:4px;padding-right:92px;'>{result_scope_title}</div>"
            "<div style='margin-bottom:2px;'>"
            "<table width='100%' cellspacing='2' cellpadding='0' style='width:100%;'>"
            "<tr>"
            + "".join(model_cards)
            + "</tr></table>"
            "</div>"
            f"<div style='font-size:11px;color:#334155;line-height:1.3;margin-top:2px;margin-bottom:0;'>"
            f"{result_scope_note}｜累计实际等待 {actual_wait_text}s"
            "</div>"
            f"<div style='font-size:12px;color:#334155;line-height:1.4;margin-top:3px;'>"
            f"<span style='font-weight:700;color:#0f172a;'>风险提示：</span>{risk_hint_text}"
            "</div>"
            f"<div style='font-size:11px;color:#475569;line-height:1.35;margin-top:3px;'>"
            f"<span style='font-weight:700;color:#0f172a;'>等待真因：</span>{cause_chain_text} "
            "<a href='wait-cause-details' style='color:#2563eb;text-decoration:none;'>查看车辆明细</a>"
            "</div>"
        )
        
        html = (
            "<table width='100%' cellspacing='0' cellpadding='0'>"
            "<tr>"
            f"<td width='30%' valign='top' style='width:30%;padding-right:12px;border-right:1px solid #dbe3ef;'>{left_html}</td>"
            f"<td width='70%' valign='top' style='width:70%;padding-left:12px;'>{right_html}</td>"
            "</tr>"
            "</table>"
        )
        self.lbl_vehicle_summary.setText(html)
        self._position_model_result_explanation_button()

    def _position_model_result_explanation_button(self):
        if not hasattr(self, "btn_model_result_explanation") or not hasattr(self, "lbl_vehicle_summary"):
            return
        button = self.btn_model_result_explanation
        label = self.lbl_vehicle_summary
        margin = 14
        x = max(margin, label.width() - button.width() - margin)
        y = 6
        button.move(x, y)
        button.raise_()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._position_model_result_explanation_button()

    def _get_model_result_explanation_anchor_widget(self):
        if hasattr(self, "lbl_vehicle_summary"):
            return self.lbl_vehicle_summary
        return self

    def _build_model_result_explanation_text(self, summary):
        if not summary:
            return ""

        def _fmt_num(value, default="—"):
            if value is None:
                return default
            return self._fmt_analysis_num(value)

        def _fmt_minutes(value):
            try:
                minutes = float(value or 0.0)
            except Exception:
                return "0"
            if abs(minutes - round(minutes)) < 1e-9:
                return str(int(round(minutes)))
            return f"{minutes:.1f}"

        output_count = int(summary.get("output_count", 0) or 0)
        qualified_count = int(summary.get("qualified_count", 0) or 0)
        qualified_rate_percent = summary.get("qualified_rate_percent")
        target_takt = summary.get("target_takt", 0.0)
        first_out = summary.get("first_out")
        last_out = summary.get("last_out")
        last_output_car_no = summary.get("last_output_car_no")
        last_output_car_out = summary.get("last_output_car_out")
        overall_takt = summary.get("overall_takt")
        total_actual_wait = summary.get("total_actual_wait", summary.get("total_block_wait", 0.0))
        total_excess_wait = summary.get("total_excess_wait", 0.0)
        risk_text = str(summary.get("risk_text", "") or "暂无明显风险")
        analysis_time_minutes = summary.get("analysis_time_minutes", 0.0)
        analysis_time_seconds = summary.get("analysis_time_seconds", 0.0)

        if output_count <= 0:
            output_calc = "当前分析时间内暂无车辆完成下线，所以当前下线车辆为 0台。"
        elif last_output_car_no is not None and last_output_car_out is not None and analysis_time_seconds:
            output_calc = (
                f"第{escape(str(last_output_car_no))}台车辆下线完成时间 {_fmt_num(last_output_car_out)}s ≤ "
                f"分析时间 {_fmt_minutes(analysis_time_minutes)}分钟（{_fmt_num(analysis_time_seconds)}s），"
                f"所以当前下线车辆为 {output_count}台。"
            )
        else:
            output_calc = f"当前分析时间范围内共统计到 {output_count}台已下线车辆。"

        if output_count <= 0:
            qualified_calc = "当前暂无下线车辆，因此达标车辆暂显示为 0/0。"
        else:
            qualified_calc = (
                f"本次下线车辆 {output_count}台，其中 {qualified_count}台满足工位能力节拍要求，"
                f"所以达标车辆为 {qualified_count}/{output_count}。"
            )

        if output_count <= 0 or qualified_rate_percent is None:
            rate_calc = "当前暂无有效下线车辆，因此达标率暂显示为 —。"
        else:
            rate_calc = f"{qualified_count} ÷ {output_count} × 100% = {qualified_rate_percent:.1f}%。"

        if output_count < 2 or first_out is None or last_out is None or overall_takt is None:
            overall_calc = "当前下线车辆不足 2台，无法计算相邻下线间隔平均值，所以整体节拍暂显示为 -。"
        else:
            overall_calc = (
                f"（{_fmt_num(last_out)} - {_fmt_num(first_out)}）÷（{output_count} - 1）≈ "
                f"{_fmt_num(overall_takt)}s/台；目标节拍为 {_fmt_num(target_takt)}s/台，"
                f"所以整体节拍显示为 {_fmt_num(overall_takt)}/{_fmt_num(target_takt)}。"
            )

        wait_calc = (
            f"本次所有车辆在工程完成后实际停留等待合计为 {_fmt_num(total_actual_wait)}s；"
            f"扣除各等待工程可接纳上限后，累计节拍外等待为 {_fmt_num(total_excess_wait)}s。"
        )
        risk_calc = f"当前风险提示显示为：{escape(risk_text)}。"

        return f"""
<div style="font-size:13px; line-height:1.3; color:#0f172a;">
  <div style="font-size:16px; font-weight:700; margin-bottom:8px;">模型结果说明</div>
  <div style="margin-bottom:10px; color:#334155;">
    以下内容用于解释当前模型结果的计算来源，帮助理解当前排程表现。
  </div>

  <div style="margin-top:8px;"><b>1. 下线车辆</b></div>
  <div>表示：在设定分析时间内，已经完成最后一道工序并下线的车辆数量。</div>
  <div>计算口径：下线完成时间 ≤ 分析时间。</div>
  <div>本次计算：{output_calc}</div>

  <div style="margin-top:8px;"><b>2. 达标车辆</b></div>
  <div>表示：下线车辆中，经过的所有有效工位，其工位能力节拍均不超过目标节拍的车辆数量。</div>
  <div>计算口径：工位能力节拍 = 当前车型该工位工时 ÷ 该工位有效设备数。</div>
  <div>补充说明：工时为 0 的经过节点不参与工位能力超节拍判断。</div>
  <div>本次计算：{qualified_calc}</div>

  <div style="margin-top:8px;"><b>3. 达标率</b></div>
  <div>表示：达标车辆在下线车辆中的占比。</div>
  <div>计算口径：达标率 = 达标车辆 ÷ 下线车辆 × 100%。</div>
  <div>本次计算：{rate_calc}</div>

  <div style="margin-top:8px;"><b>4. 整体节拍</b></div>
  <div>表示：已下线车辆在当前分析时间内的整体下线节奏。</div>
  <div>计算口径：整体节拍 =（最后一台下线完成时间 - 第一台下线完成时间）÷（下线车辆数 - 1），即全部相邻下线间隔的平均值。</div>
  <div>本次计算：{overall_calc}</div>

  <div style="margin-top:8px;"><b>5. 累计实际等待与累计节拍外等待</b></div>
  <div>表示：累计实际等待是车辆加工完成后真实停留的总时间；累计节拍外等待是其中超过当前工程可接纳上限、无法在目标节拍内吸收的部分。</div>
  <div>计算口径：可接纳上限 = 有效设备数 × 目标节拍；节拍外等待 = max（0，实际等待 - 可接纳上限）。</div>
  <div>本次计算：{wait_calc}</div>

  <div style="margin-top:8px;"><b>6. 风险提示</b></div>
  <div>表示：提示节拍外等待发生在哪里，以及哪些车型在具体工程的加工工时超过工程能力上限。</div>
  <div>计算口径：等待发生工程与超节拍工程分开显示；整体节拍已在结果卡片中显示，不在此重复。</div>
  <div>本次计算：{risk_calc}</div>

  <div style="margin-top:10px; color:#475569;">
    说明：以上内容用于帮助理解当前模型结果，不代表最终业务判定。
  </div>
</div>
""".strip()

    def _show_model_result_explanation_dialog(self):
        summary = getattr(self, "_last_model_result_summary", None)
        if not summary:
            QMessageBox.information(self, "提示", "请先完成一次分析，再查看结果说明。")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("模型结果说明")
        dialog.resize(760, 560)

        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        browser = QTextBrowser(dialog)
        browser.setReadOnly(True)
        browser.setOpenExternalLinks(False)
        browser.setStyleSheet(
            "QTextBrowser {"
            "background:#ffffff;"
            "border:1px solid #dbe3ef;"
            "border-radius:8px;"
            "padding:8px;"
            "font-size:12px;"
            "}"
        )
        browser.setHtml(self._build_model_result_explanation_text(summary))
        layout.addWidget(browser, 1)

        button_row = QHBoxLayout()
        button_row.addStretch()
        btn_close = QPushButton("关闭", dialog)
        btn_close.clicked.connect(dialog.accept)
        button_row.addWidget(btn_close)
        layout.addLayout(button_row)

        anchor = self._get_model_result_explanation_anchor_widget()
        if anchor is not None:
            global_pos = anchor.mapToGlobal(anchor.rect().bottomLeft())
            x = global_pos.x()
            y = global_pos.y() + 8
            screen = self.screen().availableGeometry() if self.screen() else None
            parent_geo = self.frameGeometry()
            if screen is not None:
                if x + dialog.width() > screen.right() - 12:
                    x = max(screen.left() + 12, screen.right() - dialog.width() - 12)
                if y + dialog.height() > screen.bottom() - 12:
                    x = max(screen.left() + 12, min(parent_geo.center().x() - dialog.width() // 2, screen.right() - dialog.width() - 12))
                    y = max(screen.top() + 12, min(parent_geo.center().y() - dialog.height() // 2, screen.bottom() - dialog.height() - 12))
            dialog.move(x, y)

        dialog.exec()

    def _active_sim_rows(self):
        """返回当前仿真时间正在加工的排程段。"""
        rows = getattr(self, "last_schedule_rows", []) or []
        current = float(getattr(self, "sim_time", 0.0) or 0.0)
        active = []
        for row in rows:
            start = self._sim_row_start(row)
            end = self._sim_row_end(row)
            if start <= current < end:
                active.append((start, end, row))
        active.sort(key=lambda x: (self._sim_row_station(x[2]), x[0]))
        return active

    def _draw_sim_scene(self):
        """绘制横向双轨坐标系的生产线仿真底图。"""
        if not hasattr(self, "sim_scene"):
            return

        self.sim_scene.clear()
        rows = getattr(self, "last_schedule_rows", []) or []
        if not rows:
            self.sim_scene.addText("请先点击『分析当前排程』生成仿真数据。")
            return

        station_defs = self._sim_station_defs()
        station_names = [item["name"] for item in station_defs]
        if not station_names:
            self.sim_scene.addText("暂无岗位数据。")
            return

        station_count = len(station_names)
        view_w = 0
        if hasattr(self, "sim_graphics_view") and self.sim_graphics_view is not None:
            try:
                view_w = int(self.sim_graphics_view.viewport().width() or 0)
            except Exception:
                view_w = 0
        view_w = max(920, view_w or 0)

        margin_x = 20
        line_label_w = 64
        track_lead_w = 14
        zone_top = 24
        line1_y = 88
        line2_y = 148
        lane_gap = line2_y - line1_y
        title_band_h = 34

        def _get_station_metrics(count, available_width):
            if count <= 4:
                pref_station_w, pref_gap = 208, 24
            elif count == 5:
                pref_station_w, pref_gap = 176, 14
            elif count == 6:
                pref_station_w, pref_gap = 154, 12
            elif count == 7:
                pref_station_w, pref_gap = 134, 10
            else:
                pref_station_w, pref_gap = 118, 8

            if count <= 1:
                return pref_station_w, pref_gap

            pref_total = count * pref_station_w + (count - 1) * pref_gap
            if pref_total <= available_width:
                return pref_station_w, pref_gap

            min_gap = 6
            min_station_w = 96 if count >= 8 else 108
            fit_station_w = int((available_width - (count - 1) * min_gap) / max(1, count))
            fit_station_w = max(min_station_w, fit_station_w)
            return fit_station_w, min_gap

        usable_track_w = max(620, view_w - margin_x * 2 - line_label_w - track_lead_w - 28)
        station_w, station_gap = _get_station_metrics(station_count, usable_track_w)
        compact_mode = station_w <= 134
        slot_h = 38 if compact_mode else 40
        slot_w = max(82, station_w - 24)
        zone_h = int((line2_y + slot_h / 2 + 18) - zone_top)
        track_start_x = margin_x + line_label_w
        first_station_x = track_start_x + track_lead_w
        last_station_x = first_station_x + (station_count - 1) * (station_w + station_gap)
        track_end_x = last_station_x + station_w + 18
        scene_w = max(view_w - 6, track_end_x + margin_x)
        scene_h = zone_top + zone_h + 14

        self.sim_scene.addRect(
            0,
            0,
            scene_w,
            scene_h,
            QPen(QColor("#243244")),
            QBrush(QColor("#243244")),
        )

        def _station_device_mode(station_info):
            scope = str(station_info.get("line_scope") or "").strip()
            try:
                device_count = int(float(station_info.get("device_count") or 0))
            except Exception:
                device_count = 0

            if scope in ("1号线", "2号线"):
                return "single"
            if scope == "双线共用":
                return "shared"
            if scope == "双线" and device_count == 1:
                return "shared"
            return "parallel"

        def _line_slots(line_scope):
            scope = str(line_scope or "").strip()
            if scope == "1号线":
                return [("1号线", True), ("2号线", False)]
            if scope == "2号线":
                return [("1号线", False), ("2号线", True)]
            return [("1号线", True), ("2号线", True)]

        def _line_key(line_no):
            text = str(line_no or "").strip()
            if "2" in text:
                return "2号线"
            return "1号线"

        def _build_sim_track_layout():
            layout = {}
            for idx, station_info in enumerate(station_defs):
                x = first_station_x + idx * (station_w + station_gap)
                slot_x = x + max(0, (station_w - slot_w) / 2)
                layout[station_info["name"]] = {
                    "x": x,
                    "title_x": x + 12,
                    "slot_x": slot_x,
                    "1号线": {
                        "x": slot_x,
                        "y": line1_y - slot_h / 2,
                        "w": slot_w,
                        "h": slot_h,
                    },
                    "2号线": {
                        "x": slot_x,
                        "y": line2_y - slot_h / 2,
                        "w": slot_w,
                        "h": slot_h,
                    },
                }
            return layout

        def _draw_track_headers():
            for label, y in (("1号线", line1_y), ("2号线", line2_y)):
                head_item = self.sim_scene.addText(label)
                head_item.setDefaultTextColor(QColor("#cbd5e1"))
                font = QFont()
                font.setPointSize(11)
                font.setBold(True)
                head_item.setFont(font)
                head_item.setPos(margin_x, y - 15)
                self.sim_scene.addLine(
                    track_start_x - 12,
                    y,
                    track_start_x - 1,
                    y,
                    QPen(QColor("#64748b"), 1),
                )

        def _draw_lane_lines():
            channel_h = 22
            channel_fill = QBrush(QColor("#334155"))
            channel_border = QPen(QColor(100, 116, 139, 150), 1)
            texture_pen = QPen(QColor(71, 85, 105, 115), 1)
            for idx in range(station_count - 1):
                left_x = first_station_x + idx * (station_w + station_gap) + station_w - 1
                right_x = first_station_x + (idx + 1) * (station_w + station_gap) + 1
                if right_x <= left_x:
                    continue
                for y in (line1_y, line2_y):
                    self.sim_scene.addRect(
                        left_x,
                        y - channel_h / 2,
                        right_x - left_x,
                        channel_h,
                        QPen(Qt.NoPen),
                        channel_fill,
                    )
                    self.sim_scene.addLine(
                        left_x,
                        y - channel_h / 2,
                        right_x,
                        y - channel_h / 2,
                        channel_border,
                    )
                    self.sim_scene.addLine(
                        left_x,
                        y + channel_h / 2,
                        right_x,
                        y + channel_h / 2,
                        channel_border,
                    )
                    texture_x = left_x + 8
                    while texture_x < right_x - 4:
                        self.sim_scene.addLine(
                            texture_x,
                            y - channel_h / 2 + 3,
                            texture_x,
                            y + channel_h / 2 - 3,
                            texture_pen,
                        )
                        texture_x += 10

        def _draw_station_zone(station_info, layout_info):
            x = layout_info["x"]
            self.sim_scene.addRect(
                x,
                zone_top,
                station_w,
                zone_h,
                QPen(QColor(148, 163, 184, 105), 1),
                QBrush(QColor(203, 213, 225, 90)),
            )

            title_item = self.sim_scene.addText(f"ST-{station_info['seq']} {station_info['name']}")
            title_item.setDefaultTextColor(QColor("#e5e7eb"))
            title_font = QFont()
            title_font.setPointSize(9 if station_count >= 8 else (10 if compact_mode else 11))
            title_font.setBold(True)
            title_item.setFont(title_font)
            title_item.document().setDocumentMargin(0)
            title_item.setTextWidth(max(74, station_w - 20))
            title_item.setPos(layout_info["title_x"] - 2, zone_top + 11)

            if _station_device_mode(station_info) == "shared":
                badge_w = 56
                badge_h = 18
                badge_x = x + station_w - badge_w - 12
                badge_y = zone_top + 12
                self.sim_scene.addRect(
                    badge_x,
                    badge_y,
                    badge_w,
                    badge_h,
                    QPen(QColor("#f59e0b"), 1),
                    QBrush(QColor("#fffbeb")),
                )
                badge_item = self.sim_scene.addText("共用设备")
                badge_item.setDefaultTextColor(QColor("#b45309"))
                badge_item.setPos(badge_x + 6, badge_y + 1)

        def _draw_station_track_slot(slot, enabled):
            if enabled:
                pen = QPen(QColor(148, 163, 184, 95), 1)
                brush = QBrush(QColor(248, 250, 252, 175))
            else:
                pen = QPen(QColor(156, 163, 175, 75), 1, Qt.DashLine)
                brush = QBrush(QColor(203, 213, 225, 70))
            self.sim_scene.addRect(
                slot["x"],
                slot["y"],
                slot["w"],
                slot["h"],
                pen,
                brush,
            )
            status_text = "空闲" if enabled else "不适用"
            status_item = self.sim_scene.addText(status_text)
            slot_font = QFont()
            slot_font.setPointSize(9 if compact_mode else 10)
            status_item.setFont(slot_font)
            status_item.setDefaultTextColor(QColor("#94a3b8" if enabled else "#9ca3af"))
            status_item.setPos(slot["x"] + 10, slot["y"] + max(8, (slot["h"] - 18) / 2))

        def _draw_slot_waiting_outline(slot):
            self.sim_scene.addRect(
                slot["x"],
                slot["y"],
                slot["w"],
                slot["h"],
                QPen(QColor("#f97316"), 2),
            )

        def _draw_overflow_badge(x, y, count):
            badge_w = 52
            badge_h = 18
            self.sim_scene.addRect(
                x,
                y,
                badge_w,
                badge_h,
                QPen(QColor("#64748b"), 1),
                QBrush(QColor("#f8fafc")),
            )
            badge_item = self.sim_scene.addText(f"还有{count}台")
            badge_item.setDefaultTextColor(QColor("#334155"))
            badge_item.setPos(x + 4, y + 1)

        def _vehicle_block_label(row):
            car_type = str(
                self._sim_row_value(row, "car_type", "type", "vehicle_type", default="")
                or ""
            ).upper()
            car_no = str(self._sim_row_value(row, "car", "car_no", "car_id", default="") or "")
            return f"{car_type or '车'}-{car_no or '?'}"

        def _slot_status_text(status, seconds_text):
            if status == "移动":
                return ""
            if status == "等待":
                return seconds_text if str(seconds_text).startswith("等待") else f"等待 {seconds_text}"
            return f"{status} {seconds_text}".strip()

        def _draw_slot_status_label(slot, status, seconds_text):
            label_text = _slot_status_text(status, seconds_text)
            if not label_text:
                return
            if status == "等待":
                color = QColor("#ef4444")
            else:
                color = QColor("#22c55e")
            label_item = self.sim_scene.addText(label_text)
            label_font = QFont()
            label_font.setPointSize(8 if compact_mode else 9)
            label_font.setBold(True)
            label_item.setFont(label_font)
            label_item.setDefaultTextColor(color)
            label_item.document().setDocumentMargin(0)
            label_item.setTextWidth(slot["w"])
            label_item.setPos(slot["x"] + 2, max(zone_top + title_band_h - 2, slot["y"] - 15))
            label_item.setZValue(4)

        def _draw_vehicle_capsule(x, y, w, h, row, status, seconds_text, over_takt=False):
            car_type = str(
                self._sim_row_value(row, "car_type", "type", "vehicle_type", default="")
                or ""
            ).upper()
            if status == "等待":
                border = QColor("#dc2626")
                fill = QColor("#fee2e2")
            elif status == "移动":
                border = QColor("#38bdf8")
                fill = QColor("#e0f2fe")
            else:
                border = QColor("#16a34a")
                fill = QColor("#dcfce7")

            if car_type == "A":
                type_bar = QColor("#2563eb")
            elif car_type == "B":
                type_bar = QColor("#ea580c")
            elif car_type == "C":
                type_bar = QColor("#16a34a")
            else:
                type_bar = QColor("#64748b")

            shadow_offset = 2
            self.sim_scene.addRect(
                x + shadow_offset,
                y + shadow_offset,
                w,
                h,
                QPen(QColor(15, 23, 42, 80), 1),
                QBrush(QColor(15, 23, 42, 60)),
            )
            self.sim_scene.addRect(
                x,
                y,
                w,
                h,
                QPen(border, 2),
                QBrush(fill),
            )
            self.sim_scene.addRect(
                x,
                y,
                5,
                h,
                QPen(type_bar, 1),
                QBrush(type_bar),
            )

            nose_w = 8 if compact_mode else 10
            self.sim_scene.addRect(
                x + w - nose_w - 3,
                y + 4,
                nose_w,
                max(8, h - 8),
                QPen(border, 1),
                QBrush(QColor("#f8fafc")),
            )
            self.sim_scene.addLine(
                x + 9,
                y + 4,
                x + w - nose_w - 8,
                y + 4,
                QPen(QColor("#ffffff"), 1),
            )

            wheel_r = 4 if compact_mode else 5
            wheel_y = y + h - 1
            for wheel_x in (x + w * 0.25, x + w * 0.76):
                self.sim_scene.addEllipse(
                    wheel_x - wheel_r,
                    wheel_y - wheel_r,
                    wheel_r * 2,
                    wheel_r * 2,
                    QPen(QColor("#0f172a"), 1),
                    QBrush(QColor("#0f172a")),
                )

            label_item = self.sim_scene.addText(_vehicle_block_label(row))
            label_font = QFont()
            label_font.setPointSize(10 if compact_mode else 11)
            label_font.setBold(True)
            label_item.setFont(label_font)
            label_item.setDefaultTextColor(QColor("#0f172a"))
            label_item.document().setDocumentMargin(0)
            label_item.setTextWidth(max(40, w - 18))
            label_item.setPos(x + 12, y + max(3, (h - 17) / 2))
            label_item.setZValue(5)

        station_layout = _build_sim_track_layout()
        _draw_track_headers()
        _draw_lane_lines()

        slot_rects = {}
        station_modes = {}
        for station_info in station_defs:
            name = station_info["name"]
            line_scope = station_info.get("line_scope")
            station_modes[name] = _station_device_mode(station_info)
            layout_info = station_layout[name]
            _draw_station_zone(station_info, layout_info)
            slots = _line_slots(line_scope)
            slot_rects[name] = {}
            for line_label, enabled in slots:
                slot = dict(layout_info[line_label])
                slot["enabled"] = enabled
                _draw_station_track_slot(slot, enabled)
                slot_rects[name][line_label] = slot

        def _slot_for_row(row):
            station_slots = slot_rects.get(self._sim_row_station(row))
            if not station_slots:
                return None, ""
            slot_line = _line_key(self._sim_row_line_no(row))
            slot = station_slots.get(slot_line)
            if not slot or not slot.get("enabled"):
                fallback = next(
                    (
                        (line_label, candidate)
                        for line_label, candidate in station_slots.items()
                        if candidate.get("enabled")
                    ),
                    None,
                )
                if fallback:
                    slot_line, slot = fallback
            return slot, slot_line

        def _vehicle_rect_from_slot(slot, block_w=None):
            default_w = min(slot["w"] - 22, 78 if compact_mode else 88)
            w = block_w if block_w is not None else default_w
            w = max(54, min(w, slot["w"] - 12))
            h = min(slot["h"] - 12, 24 if compact_mode else 26)
            return {
                "x": slot["x"] + (slot["w"] - w) / 2,
                "y": slot["y"] + (slot["h"] - h) / 2,
                "w": w,
                "h": h,
            }

        def _smoothstep(value):
            value = max(0.0, min(1.0, value))
            return value * value * (3 - 2 * value)

        def _lerp_point(start_point, end_point, ratio):
            eased = _smoothstep(ratio)
            return (
                start_point[0] + (end_point[0] - start_point[0]) * eased,
                start_point[1] + (end_point[1] - start_point[1]) * eased,
            )

        def _transition_rect(prev_row, next_row, from_slot, to_slot, move_start, move_end):
            if move_end <= move_start:
                return _vehicle_rect_from_slot(to_slot)
            ratio = (current - move_start) / (move_end - move_start)
            ratio = max(0.0, min(1.0, ratio))
            from_rect = _vehicle_rect_from_slot(from_slot)
            to_rect = _vehicle_rect_from_slot(to_slot)
            rect_w = from_rect["w"]
            rect_h = from_rect["h"]

            from_center = (
                from_rect["x"] + from_rect["w"] / 2,
                from_rect["y"] + from_rect["h"] / 2,
            )
            to_center = (
                to_rect["x"] + to_rect["w"] / 2,
                to_rect["y"] + to_rect["h"] / 2,
            )

            prev_layout = station_layout.get(self._sim_row_station(prev_row), {})
            next_layout = station_layout.get(self._sim_row_station(next_row), {})
            moving_right = to_center[0] >= from_center[0]
            if moving_right:
                prev_exit_x = float(prev_layout.get("x", from_slot["x"])) + station_w + 6
                next_entry_x = float(next_layout.get("x", to_slot["x"])) - 6
            else:
                prev_exit_x = float(prev_layout.get("x", from_slot["x"])) - 6
                next_entry_x = float(next_layout.get("x", to_slot["x"])) + station_w + 6

            prev_exit = (prev_exit_x, from_center[1])
            next_entry = (next_entry_x, to_center[1])

            if ratio < 0.25:
                center = _lerp_point(from_center, prev_exit, ratio / 0.25)
            elif ratio < 0.75:
                center = _lerp_point(prev_exit, next_entry, (ratio - 0.25) / 0.50)
            else:
                center = _lerp_point(next_entry, to_center, (ratio - 0.75) / 0.25)

            return {
                "x": center[0] - rect_w / 2,
                "y": center[1] - rect_h / 2,
                "w": rect_w,
                "h": rect_h,
            }

        current = float(getattr(self, "sim_time", 0.0) or 0.0)
        vehicle_blocks = []
        visual_move_window = 2.0
        car_rows = self._sim_car_rows()
        transition_car_keys = set()

        for car_key, car_segments in car_rows.items():
            if not car_segments:
                continue
            for idx in range(1, len(car_segments)):
                prev_seg = car_segments[idx - 1]
                next_seg = car_segments[idx]
                next_start = self._sim_row_start(next_seg)
                prev_finish = self._sim_row_service_finish(prev_seg)
                if next_start < prev_finish:
                    continue
                if next_start <= current < next_start + visual_move_window:
                    from_slot, _ = _slot_for_row(prev_seg)
                    to_slot, _ = _slot_for_row(next_seg)
                    if not from_slot or not to_slot:
                        continue
                    transition_car_keys.add(car_key)
                    vehicle_blocks.append({
                        "station": self._sim_row_station(next_seg),
                        "line": _line_key(self._sim_row_line_no(next_seg)),
                        "row": next_seg,
                        "status": "移动",
                        "seconds": f"剩余{max(0.0, next_start + visual_move_window - current):.1f}s",
                        "over_takt": False,
                        "order_time": next_start,
                        "display_rect": _transition_rect(
                            prev_seg,
                            next_seg,
                            from_slot,
                            to_slot,
                            next_start,
                            next_start + visual_move_window,
                        ),
                    })
                    break

        for row in rows:
            if self._sim_car_key(row) in transition_car_keys:
                continue
            start = self._sim_row_start(row)
            svc_finish = self._sim_row_service_finish(row)
            if start <= current < svc_finish:
                remain = max(0.0, svc_finish - current)
                vehicle_blocks.append({
                    "station": self._sim_row_station(row),
                    "line": _line_key(self._sim_row_line_no(row)),
                    "row": row,
                    "status": "加工",
                    "seconds": f"剩余{remain:.1f}s",
                    "over_takt": False,
                    "order_time": start,
                })

        for car_key, car_segments in car_rows.items():
            if car_key in transition_car_keys:
                continue
            if not car_segments:
                continue
            for idx, seg in enumerate(car_segments):
                next_seg = car_segments[idx + 1] if idx + 1 < len(car_segments) else None
                wait_start = self._sim_row_service_finish(seg)
                wait_end = self._sim_row_wait_end(seg, next_seg)
                if wait_start <= current < wait_end:
                    vehicle_blocks.append({
                        "station": self._sim_row_station(seg),
                        "line": _line_key(self._sim_row_line_no(seg)),
                        "row": seg,
                        "status": "等待",
                        "seconds": self._sim_wait_label(seg, next_seg, current, wait_end, set()),
                        "over_takt": False,
                        "order_time": wait_start,
                    })
                    break

        shared_processing = {}
        for item in sorted(vehicle_blocks, key=lambda value: value["order_time"]):
            if station_modes.get(item["station"]) != "shared":
                continue
            if item["status"] in ("等待", "移动"):
                continue
            station_name = item["station"]
            if station_name not in shared_processing:
                shared_processing[station_name] = item
                continue
            item["status"] = "等待"
            item["seconds"] = "等待"
            item["over_takt"] = False

        def _vehicle_sort_key(item):
            station_index = station_names.index(item["station"]) if item["station"] in station_names else 999
            status_rank = 1 if item["status"] == "等待" else 0
            return station_index, item["line"], status_rank, item["order_time"]

        vehicle_blocks.sort(key=_vehicle_sort_key)
        slot_groups = {}
        moving_blocks = []
        for item in vehicle_blocks:
            if item.get("display_rect"):
                moving_blocks.append(item)
                continue
            station_slots = slot_rects.get(item["station"])
            if not station_slots:
                continue
            slot_line = item["line"]
            slot = station_slots.get(item["line"])
            if not slot or not slot.get("enabled"):
                fallback = next(
                    (
                        (line_label, candidate)
                        for line_label, candidate in station_slots.items()
                        if candidate.get("enabled")
                    ),
                    None,
                )
                if fallback:
                    slot_line, slot = fallback
            if not slot:
                continue
            slot_key = (item["station"], slot_line)
            slot_groups.setdefault(slot_key, {"slot": slot, "items": []})["items"].append(item)

        for slot_group in slot_groups.values():
            slot = slot_group["slot"]
            items = slot_group["items"]
            if not items:
                continue
            main_item = items[0]
            overflow_count = max(0, len(items) - 1)
            block_w = None
            if overflow_count:
                block_w = max(54, min(slot["w"] - 66, 64 if compact_mode else 72))
            rect = _vehicle_rect_from_slot(slot, block_w)
            _draw_slot_status_label(slot, main_item["status"], main_item["seconds"])
            _draw_vehicle_capsule(
                rect["x"],
                rect["y"],
                rect["w"],
                rect["h"],
                main_item["row"],
                main_item["status"],
                main_item["seconds"],
                main_item.get("over_takt", False),
            )
            if overflow_count:
                _draw_overflow_badge(
                    slot["x"] + slot["w"] - 56,
                    slot["y"] + max(4, (slot["h"] - 18) / 2),
                    overflow_count,
                )

        for item in moving_blocks:
            rect = item["display_rect"]
            _draw_vehicle_capsule(
                rect["x"],
                rect["y"],
                rect["w"],
                rect["h"],
                item["row"],
                item["status"],
                item["seconds"],
                item.get("over_takt", False),
            )

        self.sim_scene.setSceneRect(0, 0, scene_w, scene_h)
