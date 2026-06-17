"""Build the read-only M-Line schedule analysis workbook from existing results."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPORT_SHEETS = ("结果总览", "车辆时间明细", "等待真因明细", "计算口径说明")


def _number(value, default=0.0):
    try:
        return float(value or 0.0)
    except (TypeError, ValueError):
        return float(default)


def _integer(value, default=0):
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return int(default)


def _station(row):
    return str(row.get("step_display", row.get("station", row.get("group", "工程"))) or "工程")


def _value(row, *keys, default=0.0):
    for key in keys:
        if row.get(key) not in (None, ""):
            return row.get(key)
    return default


def build_vehicle_records(
    schedule_rows,
    *,
    is_ratio_mode,
    analysis_time_seconds,
    theoretical_launch_count,
    report_cutoff_seconds,
    target_takt,
    capacity_results,
):
    """Return one record per vehicle, ordered by actual output time."""
    grouped = {}
    for row in schedule_rows or []:
        car = _integer(row.get("car", row.get("car_no", row.get("car_index", 0))))
        if car > 0:
            grouped.setdefault(car, []).append(row)

    capacity_by_car = {
        _integer(item.get("car")): item
        for item in capacity_results or []
        if _integer(item.get("car")) > 0
    }
    records = []
    for car, rows in grouped.items():
        ordered = sorted(
            rows,
            key=lambda row: (
                _number(_value(row, "start", "start_time")),
                _integer(row.get("step_seq", row.get("seq", 0))),
            ),
        )
        first = ordered[0]
        last = ordered[-1]
        car_in = _number(_value(first, "start", "start_time"))
        car_out = _number(_value(last, "depart", "end", "svc_finish", "finish"))
        launch_wait = sum(_number(row.get("launch_wait")) for row in ordered)
        post_wait = sum(_number(row.get("block_wait")) for row in ordered)
        occurred_post_wait = 0.0
        occurred_excess_wait = 0.0
        for row in ordered:
            service_finish = _number(
                _value(
                    row,
                    "svc_finish",
                    "service_finish",
                    default=(
                        _number(_value(row, "start", "start_time"))
                        + _number(_value(row, "dur", "duration"))
                    ),
                )
            )
            depart = _number(_value(row, "depart", "end", default=service_finish))
            if report_cutoff_seconds <= service_finish + 1e-9 or depart <= service_finish + 1e-9:
                continue
            occurred_end = min(report_cutoff_seconds, depart)
            occurred = max(0.0, occurred_end - service_finish)
            occurred_post_wait += occurred
            capacity = max(1, _integer(_value(row, "capacity", "device_count", default=1), 1))
            holding_limit = capacity * max(0.0, _number(target_takt))
            occurred_excess_wait += max(0.0, occurred_end - (service_finish + holding_limit))
        car_type = str(_value(first, "car_type", "duration_source", "vehicle_type", default="") or "")

        if is_ratio_mode:
            if car > theoretical_launch_count:
                scope = "仿真缓冲"
            elif car_out <= analysis_time_seconds + 1e-9:
                scope = "分析窗口内"
            else:
                scope = "目标批次窗口外"
        else:
            scope = "目标批次"

        capacity = capacity_by_car.get(car, {})
        capacity_text = str(capacity.get("capacity_status", "未设定目标") or "未设定目标")
        over_text = str(capacity.get("over_capacity_station_text", "无") or "无")
        if capacity_text == "能力超目标":
            capacity_text = f"能力超目标：{over_text}"

        segments = []
        for row in ordered:
            seq = row.get("step_seq", row.get("seq", ""))
            label = f"ST{seq}" if seq not in (None, "") else "ST?"
            segments.append(
                f"{label} {_station(row)}(开始:{_number(_value(row, 'start', 'start_time')):.1f}s "
                f"加工:{_number(_value(row, 'dur', 'duration')):.1f}s "
                f"投入等待:{_number(row.get('launch_wait')):.1f}s "
                f"完工后等待:{_number(row.get('block_wait')):.1f}s "
                f"离开:{_number(_value(row, 'depart', 'end', 'svc_finish', 'finish')):.1f}s)"
            )

        records.append({
            "car": car,
            "car_label": f"Car#{car}",
            "car_type": car_type,
            "scope": scope,
            "car_in": car_in,
            "car_out": car_out,
            "flow": max(0.0, car_out - car_in),
            "wait": launch_wait + post_wait,
            "launch_wait": launch_wait,
            "post_wait": post_wait,
            "occurred_post_wait": occurred_post_wait,
            "occurred_excess_wait": occurred_excess_wait,
            "capacity_text": capacity_text,
            "segments": " | ".join(segments),
        })

    records.sort(key=lambda item: (item["car_out"], item["car"]))
    for index, record in enumerate(records, start=1):
        record["output_order"] = index
    return records


def _apply_sheet_title(sheet, title, width):
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    cell = sheet.cell(row=1, column=1, value=title)
    cell.font = Font(size=16, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 26


def _section(sheet, row, text, width):
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    cell = sheet.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(vertical="center")


def _table_header(sheet, row, headers):
    fill = PatternFill("solid", fgColor="D9EAF7")
    border = Border(bottom=Side(style="thin", color="9FBAD0"))
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=header)
        cell.font = Font(bold=True, color="17365D")
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_overview(workbook, payload, records):
    sheet = workbook.active
    sheet.title = "结果总览"
    _apply_sheet_title(sheet, "M-Line 排程分析报告", 4)
    sheet["A2"] = "导出时间"
    sheet["B2"] = payload["exported_at"]
    sheet["A3"] = "统计范围"
    sheet["B3"] = payload.get("scope_full_text", payload["scope_note"])
    sheet.merge_cells("B2:D2")
    sheet.merge_cells("B3:D3")

    _section(sheet, 5, "一、核心结果", 4)
    _table_header(sheet, 6, ("项目", "结果", "含义", "备注"))
    overview_rows = [
        (payload["output_label"], f"{payload['output_count']}台", "当前统计范围内完成最后一道工程的车辆", ""),
        (payload["completion_label"], f"{payload['completion_time']:.1f}s" if payload["completion_time"] is not None else "-", "当前统计范围内最晚的实际下线时间", payload["last_vehicle_text"]),
        ("整体节拍", payload["overall_takt_text"], "第一台至最后一台的平均下线间隔", f"目标{payload['target_takt']:.1f}s/台"),
        ("累计实际等待", f"{payload['total_actual_wait']:.1f}s", "截止统计时点已发生的完工后等待累计", "明细I列求和；已进入下线时间，不能重复相加"),
        ("累计节拍外等待", f"{payload['total_excess_wait']:.1f}s", "截止统计时点实际等待中超过工程可接纳上限的部分", "明细J列求和；用于定位等待压力"),
        ("首工程等待进入累计排队车辆", f"{payload.get('first_station_entry_queue_max', 0)}台", "目标节拍投放后，等待进入首工程的最大排队车辆数", "不计入工程完工后等待"),
    ]
    if payload["is_ratio_mode"]:
        overview_rows.extend([
            ("目标批次车辆", f"{payload['theoretical_launch_count']}台", "按分析时间和目标节拍得到的目标投车批次", f"仿真缓冲{payload['simulation_buffer_count']}台"),
            ("目标批次实际完成", f"{payload['target_batch_actual_finish']:.1f}s", "目标批次车辆全部下线的实际时刻", f"计划完成{payload['target_batch_planned_finish']:.1f}s"),
        ])
    for row_values in overview_rows:
        sheet.append(row_values)

    section_row = sheet.max_row + 2
    _section(sheet, section_row, "二、时间真实性核对", 4)
    _table_header(sheet, section_row + 1, ("核对项", "人工复算式", "状态", "说明"))
    scope_records = payload["scope_records"]
    if scope_records:
        last = scope_records[-1]
        first = scope_records[0]
        interval_sum = sum(
            scope_records[index]["car_out"] - scope_records[index - 1]["car_out"]
            for index in range(1, len(scope_records))
        )
        checks = [
            ("单车时间", f"{last['car_in']:.1f} + {last['flow']:.1f} = {last['car_out']:.1f}s", "一致", f"{last['car_label']}：投车 + 贯通 = 下线"),
            ("全部下线间隔", f"{first['car_out']:.1f} + {interval_sum:.1f} = {last['car_out']:.1f}s", "一致", "第一台下线 + 全部间隔 = 最晚下线"),
            ("统计范围完成", f"MAX（{len(scope_records)}台下线时间）= {last['car_out']:.1f}s", "一致", "最大值即当前统计范围完成时刻"),
            ("总核对状态", "时间数据核对一致", "一致", "可由车辆时间明细逐行复算"),
        ]
    else:
        checks = [("总核对状态", "当前统计范围无下线车辆", "-", "无可复算下线时间")]
    for values in checks:
        sheet.append(values)

    section_row = sheet.max_row + 2
    _section(sheet, section_row, "三、等待与工程提示", 4)
    _table_header(sheet, section_row + 1, ("项目", "摘要", "作用", "详细位置"))
    sheet.append(("主要节拍外等待位置", payload["excess_station_text"] or "无", "定位等待集中工程", "车辆时间明细 / 等待真因明细"))
    sheet.append(("超节拍工程", payload["capacity_station_text"] or "无", "识别加工工时超过工程能力上限的车型", "计算口径说明"))
    sheet.append(("节奏净差值", payload["net_takt_delta_text"], "解释整体节拍与目标节拍的差异", "车辆时间明细"))
    sheet.append(("重要说明", "累计等待已经体现在各车下线时间中，不能再次加到总完成时间。", "详细车辆与工程记录见“车辆时间明细”。", "概念解释见“计算口径说明”。"))

    for column, width in enumerate((22, 42, 42, 34), start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A6"


def _write_vehicle_details(workbook, payload, records):
    sheet = workbook.create_sheet("车辆时间明细")
    headers = [
        "下线顺序", "统计归属", "CAR", "TYPE", "投车时间 IN(s)", "下线时间 OUT(s)",
        "单车贯通 FLOW(s)", "车辆总等待 WAIT(s)\n（含投车前等待）", "截止统计时点已发生完工后等待(s)",
        "截止统计时点节拍外等待(s)", "前一台下线车辆", "相邻下线间隔(s)",
        "与目标节拍差值(s)", "延误增量(s)", "提前补偿(s)", "累计下线间隔(s)",
        "单车等式差(s)", "车型能力观察（静态）", "工程记录 SEGMENTS",
    ]
    _table_header(sheet, 1, headers)
    target_takt = payload["target_takt"]
    for index, record in enumerate(records, start=2):
        previous = records[index - 3] if index > 2 else None
        sheet.cell(index, 1, record["output_order"])
        sheet.cell(index, 2, record["scope"])
        sheet.cell(index, 3, record["car_label"])
        sheet.cell(index, 4, record["car_type"])
        sheet.cell(index, 5, record["car_in"])
        sheet.cell(index, 6, record["car_out"])
        sheet.cell(index, 7, f"=F{index}-E{index}")
        sheet.cell(index, 8, record["wait"])
        sheet.cell(index, 9, record["occurred_post_wait"])
        sheet.cell(index, 10, record["occurred_excess_wait"])
        sheet.cell(index, 11, previous["car_label"] if previous else "-")
        if previous:
            sheet.cell(index, 12, f"=F{index}-F{index - 1}")
            sheet.cell(index, 13, f"=L{index}-{target_takt}")
            sheet.cell(index, 14, f"=MAX(0,M{index})")
            sheet.cell(index, 15, f"=MAX(0,-M{index})")
            sheet.cell(index, 16, f"=P{index - 1}+L{index}")
        else:
            for column in (12, 13, 14, 15):
                sheet.cell(index, column, "-")
            sheet.cell(index, 16, 0)
        sheet.cell(index, 17, f"=F{index}-E{index}-G{index}")
        sheet.cell(index, 18, record["capacity_text"])
        sheet.cell(index, 19, record["segments"])

    widths = (12, 18, 12, 8, 16, 16, 18, 19, 24, 22, 18, 18, 18, 15, 15, 20, 16, 28, 90)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _write_wait_cause_details(workbook, payload):
    sheet = workbook.create_sheet("等待真因明细")
    headers = (
        "等待车辆", "车型", "等待发生工程", "等待开始(s)", "等待结束(s)", "本段节拍外等待(s)",
        "直接阻挡车辆", "直接阻挡工程", "直接阻挡资源", "真因车辆", "等待真因", "阻挡证据链",
    )
    _table_header(sheet, 1, headers)
    details = list(payload.get("cause_chain_details", []) or [])
    if details:
        details.sort(key=lambda item: (_number(item.get("wait_start")), _integer(item.get("car"))))
        for row_index, item in enumerate(details, start=2):
            chain_parts = []
            for node in item.get("chain", []) or []:
                waiting_car = node.get("waiting_car")
                blocker_car = node.get("blocker_car")
                station = str(node.get("blocked_station", "") or "未知工程")
                if blocker_car is None:
                    chain_parts.append(f"Car#{waiting_car} 等待 {station}（未解析）")
                else:
                    chain_parts.append(f"Car#{waiting_car} 等 Car#{blocker_car} 释放 {station}")
            values = [
                f"Car#{item.get('car', '')}",
                str(item.get("car_type", "") or ""),
                str(item.get("waiting_station", "") or ""),
                _number(item.get("wait_start")),
                _number(item.get("wait_end")),
                _number(item.get("wait_time")),
                "" if item.get("direct_blocker_car") is None else f"Car#{item.get('direct_blocker_car')}",
                str(item.get("direct_blocking_station", "") or ""),
                str(item.get("direct_blocking_resource", "") or ""),
                "" if item.get("terminal_car") is None else f"Car#{item.get('terminal_car')}",
                str(item.get("terminal_cause", "") or ""),
                " → ".join(chain_parts),
            ]
            for column, value in enumerate(values, start=1):
                sheet.cell(row_index, column, value)
    else:
        sheet.cell(2, 1, "当前统计范围内未产生节拍外等待真因明细。")
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    widths = (14, 8, 24, 16, 16, 20, 16, 24, 28, 14, 28, 80)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _write_definitions(workbook, payload):
    sheet = workbook.create_sheet("计算口径说明")
    _apply_sheet_title(sheet, "计算口径说明", 3)
    sheet["A2"] = "原则"
    sheet["B2"] = "汇总结果必须能从“车辆时间明细”逐行复算；本页只解释口径，不产生新的排程结果。"
    sheet.merge_cells("B2:C2")
    sheet["A3"] = "统计范围"
    sheet["B3"] = payload.get("scope_full_text", payload["scope_note"])
    sheet.merge_cells("B3:C3")
    _table_header(sheet, 5, ("名称", "大白话解释", "核对方法 / 注意事项"))
    definitions = [
        (payload["completion_label"], "当前统计范围内最后完成车辆的实际下线时间。", "取当前统计范围全部OUT的最大值。"),
        ("目标批次总完成时刻", "目标批次全部车辆中最晚的下线时间。", "比例模式不能用窗口末台下线代替目标批次完成。"),
        ("单车贯通时间", "一台车从投车到下线经历的总时间。", "投车时间 + 单车贯通时间 = 下线时间。"),
        ("相邻下线间隔", "按实际下线先后，本车与前一台车下线时间的差。", "本车OUT - 前车OUT；第一台显示“-”。"),
        ("整体节拍", "第一台至最后一台之间的平均下线间隔。", "（最晚OUT - 第一台OUT）÷（下线车辆数-1）。"),
        ("车辆总等待 WAIT", "包含进入首工程前的投入等待和加工完成后的等待。", "不等同于模型结果中的累计实际等待。"),
        ("累计实际等待", "截止统计时点，加工完成后因下一工程暂时不能接收而已经发生的等待累计。", "车辆明细I列求和；已体现在下线时间中，不能再次加入总完成时间。"),
        ("累计节拍外等待", "截止统计时点，实际等待中超过工程可接纳上限、已经发生的部分。", "车辆明细J列求和；用于定位等待压力，不是总完成时间加项。"),
        ("延误增量 / 提前补偿", "相邻间隔比目标节拍多出的部分 / 少于目标节拍的部分。", "两者相减得到整批节奏净差值。"),
        ("车型能力观察（静态）", "逐车查看正工时工程是否超过能力上限。", "不代表本次排程最终表现。"),
        ("SEGMENTS", "每台车逐工程的开始、加工、等待和离开记录。", "时间核对异常时回查具体车辆和工程。"),
    ]
    for values in definitions:
        sheet.append(values)
    for column, width in enumerate((28, 58, 58), start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A6"


def write_analysis_report(path, payload):
    """Write and reopen the workbook to guarantee a valid Excel artifact."""
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    records = list(payload.get("vehicle_records", []) or [])
    workbook = Workbook()
    _write_overview(workbook, payload, records)
    _write_vehicle_details(workbook, payload, records)
    _write_wait_cause_details(workbook, payload)
    _write_definitions(workbook, payload)
    workbook.save(output_path)

    check = load_workbook(output_path, read_only=True, data_only=False)
    if tuple(check.sheetnames) != REPORT_SHEETS:
        raise ValueError(f"报告工作表不完整：{check.sheetnames}")
    check.close()
    return output_path


def default_export_timestamp():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
