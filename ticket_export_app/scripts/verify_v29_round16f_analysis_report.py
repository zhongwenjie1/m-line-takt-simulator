#!/usr/bin/env python3
"""Verify the round-16F analysis report without rerunning the schedule."""

from __future__ import annotations

import os
import sys
import zipfile
from pathlib import Path


os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

APP_DIR = Path(__file__).resolve().parents[1]
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

from openpyxl import load_workbook  # noqa: E402
from PySide6.QtWidgets import QApplication  # noqa: E402
from ui.export_ticket_window import ExportTicketWindow  # noqa: E402
from utils.analysis_report import REPORT_SHEETS, write_analysis_report  # noqa: E402


def _verify_quantity(window: ExportTicketWindow) -> dict:
    window.fill_sample()
    window.spn_target_takt.setValue(118)
    window.do_analyze()
    rows_identity = id(window.last_schedule_rows)
    rows_snapshot = list(window.last_schedule_rows)
    payload = window._build_analysis_report_payload()
    output = Path("/tmp/M-Line排程分析报告_16F_数量.xlsx")
    write_analysis_report(output, payload)

    assert id(window.last_schedule_rows) == rows_identity
    assert window.last_schedule_rows == rows_snapshot
    assert zipfile.is_zipfile(output)
    workbook = load_workbook(output, read_only=False, data_only=False)
    assert tuple(workbook.sheetnames) == REPORT_SHEETS
    overview = workbook["结果总览"]
    detail = workbook["车辆时间明细"]
    definitions = workbook["计算口径说明"]
    assert overview["A1"].value == "M-Line 排程分析报告"
    assert "目标批次" in str(overview["B3"].value)
    assert detail.max_row == 7
    assert detail["L3"].value == "=F3-F2"
    assert detail["Q2"].value == "=F2-E2-G2"
    assert abs(sum(detail.cell(row=row, column=9).value or 0 for row in range(2, detail.max_row + 1)) - payload["total_actual_wait"]) < 1e-6
    assert abs(sum(detail.cell(row=row, column=10).value or 0 for row in range(2, detail.max_row + 1)) - payload["total_excess_wait"]) < 1e-6
    assert "累计实际等待" in {definitions.cell(row=row, column=1).value for row in range(1, definitions.max_row + 1)}
    assert "达标车辆" not in " ".join(str(cell.value or "") for row in overview.iter_rows() for cell in row)
    workbook.close()
    return {"path": str(output), "size": output.stat().st_size, "rows": detail.max_row - 1}


def _verify_ratio(window: ExportTicketWindow) -> dict:
    window.cmb_launch_mode.setCurrentText("按比例投车")
    window.spn_a_cars.setValue(4)
    window.spn_b_cars.setValue(2)
    window.spn_c_cars.setValue(0)
    window.spn_total_cars.setValue(60)
    window.spn_target_takt.setValue(58)
    window.do_analyze()
    payload = window._build_analysis_report_payload()
    output = Path("/tmp/M-Line排程分析报告_16F_比例.xlsx")
    write_analysis_report(output, payload)

    workbook = load_workbook(output, read_only=False, data_only=False)
    overview = workbook["结果总览"]
    detail = workbook["车辆时间明细"]
    scope_text = str(overview["B3"].value)
    assert "按比例投车" in scope_text
    assert "A4" in scope_text and "B2" in scope_text and "C0" in scope_text
    assert "分析窗口60分钟" in scope_text
    assert "目标节拍58s/台" in scope_text
    scopes = {detail.cell(row=row, column=2).value for row in range(2, detail.max_row + 1)}
    assert "分析窗口内" in scopes
    assert "目标批次窗口外" in scopes
    assert "仿真缓冲" in scopes
    assert abs(sum(detail.cell(row=row, column=9).value or 0 for row in range(2, detail.max_row + 1)) - payload["total_actual_wait"]) < 1e-6
    assert abs(sum(detail.cell(row=row, column=10).value or 0 for row in range(2, detail.max_row + 1)) - payload["total_excess_wait"]) < 1e-6
    workbook.close()
    return {
        "path": str(output),
        "size": output.stat().st_size,
        "scope_counts": {
            scope: sum(record["scope"] == scope for record in payload["vehicle_records"])
            for scope in sorted(scopes)
        },
    }


def main() -> int:
    app = QApplication.instance() or QApplication([])
    window = ExportTicketWindow()
    assert not window.btn_export_analysis_report.isEnabled()
    quantity = _verify_quantity(window)
    assert window.btn_export_analysis_report.isEnabled()
    ratio = _verify_ratio(window)
    assert not hasattr(window, "btn_model_result_explanation")
    print({"quantity": quantity, "ratio": ratio, "assertions": "passed"})
    window.close()
    app.processEvents()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
