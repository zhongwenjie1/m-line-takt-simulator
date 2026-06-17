#!/usr/bin/env python3
"""Verify first-station entry queue hints in UI and analysis report."""

from __future__ import annotations

import importlib.util
import os
import sys
from pathlib import Path


os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

APP_DIR = Path(__file__).resolve().parents[1]
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

from openpyxl import load_workbook  # noqa: E402
from PySide6.QtWidgets import QApplication  # noqa: E402
from core import tickets  # noqa: E402
from ui.export_ticket_window import ExportTicketWindow  # noqa: E402
from utils.analysis_report import write_analysis_report  # noqa: E402


EXPECTED_MAX_ENTRY_QUEUE = 125


def _load_baseline():
    path = APP_DIR / "scripts" / "verify_v29_real_data_baseline.py"
    spec = importlib.util.spec_from_file_location("v29_baseline", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"无法加载基线脚本：{path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _prepare_window(window: ExportTicketWindow) -> None:
    window.cmb_launch_mode.setCurrentIndex(0)
    window.cmb_seq.setCurrentIndex(1)
    window.spn_a_cars.setValue(276)
    window.spn_b_cars.setValue(306)
    window.spn_c_cars.setValue(566)
    window.spn_max_run.setValue(5)
    window.spn_target_takt.setValue(58)


def main() -> int:
    baseline = _load_baseline()
    station_defs = baseline.station_defs()
    rows, max_finish = tickets.schedule(
        station_defs,
        1148,
        vehicle_counts={"A": 276, "B": 306, "C": 566},
        sequence_mode="alternate",
        max_consecutive=5,
        launch_takt=58.0,
    )
    analysis = tickets.analyze_schedule(rows, max_finish, 58.0)

    app = QApplication.instance() or QApplication([])
    window = ExportTicketWindow()
    _prepare_window(window)
    window.current_defs = station_defs
    window.last_schedule_rows = rows
    window.last_analysis = analysis
    window.last_max_finish = max_finish
    window._invalidate_sim_scene_cache()
    window._set_analysis_result_available(True)

    max_queue = window._max_first_station_entry_queue()
    assert max_queue == EXPECTED_MAX_ENTRY_QUEUE, max_queue
    assert window._first_station_entry_queue_count(1340.0) > 0

    window.sim_time = 1340.0
    window._show_vehicle_summary(analysis)
    summary_text = window.lbl_vehicle_summary.text()
    assert "首工程前最大排队" in summary_text
    assert "125台" in summary_text
    assert window._last_model_result_summary["first_station_entry_queue_max"] == EXPECTED_MAX_ENTRY_QUEUE

    payload = window._build_analysis_report_payload()
    assert payload["first_station_entry_queue_max"] == EXPECTED_MAX_ENTRY_QUEUE
    output = Path("/tmp/M-Line排程分析报告_16KC_首工程排队.xlsx")
    write_analysis_report(output, payload)
    workbook = load_workbook(output, read_only=False, data_only=True)
    overview = workbook["结果总览"]
    overview_items = {
        str(overview.cell(row=row, column=1).value or ""): str(overview.cell(row=row, column=2).value or "")
        for row in range(1, overview.max_row + 1)
    }
    assert overview_items["首工程等待进入累计排队车辆"] == "125台"
    workbook.close()

    window._draw_sim_scene()
    scene_texts = [
        item.toPlainText()
        for item in window.sim_scene.items()
        if hasattr(item, "toPlainText")
    ]
    assert any(text.startswith("等待进入：") for text in scene_texts)

    window.close()
    app.processEvents()
    print({
        "max_first_station_entry_queue": max_queue,
        "queue_at_1340": window._first_station_entry_queue_count(1340.0),
        "report_path": str(output),
        "report_size": output.stat().st_size,
        "assertions": "passed",
    })
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
